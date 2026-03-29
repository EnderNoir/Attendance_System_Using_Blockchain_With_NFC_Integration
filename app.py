from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, session, send_from_directory
from web3 import Web3
from datetime import datetime
from functools import wraps
from threading import Thread, Lock
import json, os, secrets, time, csv, io, hashlib, uuid, re, sqlite3, calendar as _cal
from collections import deque
from werkzeug.utils import secure_filename
import re
import io
import uuid
import json
import secrets as _sec
from datetime import datetime
# pdfminer is imported inside parse_registration_pdf() so a startup
# import glitch can never permanently disable PDF parsing for the session.

app = Flask(__name__)
app.secret_key = 'davs-super-secret-2024'

# ── Jinja2 custom filters ─────────────────────────────────────────────────────
import json as _json_mod
app.jinja_env.filters['from_json'] = lambda s: _json_mod.loads(s) if s else []

def _parse_cvsu_pdf_text(raw_bytes: bytes) -> str:
    """Extract plain text from a CvSU PDF. Tries pypdf, falls back to pdfminer."""
    full = ''
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(raw_bytes))
        full = '\n'.join(page.extract_text() or '' for page in reader.pages)
    except Exception as _e:
        print(f'[PDF] pypdf failed: {_e}')
    if not full.strip():
        try:
            from pdfminer.high_level import extract_text as _pm
            full = _pm(io.BytesIO(raw_bytes))
        except Exception as _e:
            print(f'[PDF] pdfminer failed: {_e}')
    return full
 
 
def _generate_cvsu_email(name: str) -> str:
    """sc.firstname.lastname@cvsu.edu.ph from a full name string."""
    clean = re.sub(r'\b[A-Za-z]\.\s*', '', name).strip()
    clean = re.sub(r'\b(JR|SR|II|III|IV)\.?\b', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s+', ' ', clean)
    words = clean.split()
    if len(words) >= 2:
        first_slug = ''.join(re.sub(r'[^a-z]', '', w.lower()) for w in words[:-1])
        last_slug  = re.sub(r'[^a-z]', '', words[-1].lower())
        if first_slug and last_slug:
            return f'sc.{first_slug}.{last_slug}@cvsu.edu.ph'
    return ''
 
 
def _surname_sort_key(student: dict) -> tuple:
    """
    (surname_lower, firstnames_lower) for alphabetical-by-surname sorting.
    Handles "First Mid Last" and "Last, First Mid" formats.
    """
    name = (student.get('name') or '').strip()
    if not name:
        return ('', '')
    if ',' in name:
        parts   = name.split(',', 1)
        surname = parts[0].strip()
        firsts  = parts[1].strip()
    else:
        parts   = name.split()
        surname = parts[-1] if len(parts) > 1 else parts[0]
        firsts  = ' '.join(parts[:-1])
    clean = lambda s: re.sub(r'[^a-z ]', '', s.lower()).strip()
    return (clean(surname), clean(firsts))
# ── Course name alias map (abbreviation ↔ full name) ──────────────────
_COURSE_ALIASES: dict[str, list[str]] = {
    'BS Information Technology':  ['BSIT', 'B.S. Information Technology', 'BS InfoTech'],
    'BS Computer Science':        ['BSCS', 'B.S. Computer Science', 'BS CompSci'],
    'BS Computer Engineering':    ['BSCpE', 'BSCOE', 'B.S. Computer Engineering'],
    'BS Information Systems':     ['BSIS', 'B.S. Information Systems'],
    'BS Electrical Engineering':  ['BSEE', 'B.S. Electrical Engineering'],
    'BS Electronics Engineering': ['BSEcE', 'BSECE', 'B.S. Electronics Engineering'],
    'BS Accountancy':             ['BSA', 'B.S. Accountancy'],
    'BS Business Administration': ['BSBA', 'B.S. Business Administration'],
    'BS Education':               ['BSEd', 'B.S. Education'],
    'BS Nursing':                 ['BSN', 'B.S. Nursing'],
    'BS Civil Engineering':       ['BSCE', 'B.S. Civil Engineering'],
    'BS Mechanical Engineering':  ['BSME', 'B.S. Mechanical Engineering'],
}

AUTO_THREAD = None
AUTO_THREAD_LOCK = Lock()
_ALIAS_TO_CANONICAL: dict[str, str] = {}
for _canon, _aliases in _COURSE_ALIASES.items():
    _ALIAS_TO_CANONICAL[_canon.lower()] = _canon
    for _a in _aliases:
        _ALIAS_TO_CANONICAL[_a.lower()] = _canon

def normalize_course_name(course: str) -> str:
    """Return canonical course name if recognized, else return unchanged."""
    return _ALIAS_TO_CANONICAL.get((course or '').strip().lower(), (course or '').strip())

def _extract_cvsu_fields(full: str) -> dict:
    """
    Parse all CvSU registration fields from extracted PDF text.
 
    CvSU pypdf layout:  LABEL LINE\\nVALUE LINE
    The value is always on the line immediately after its label.
 
    Returns dict with keys:
        student_id, name, email, contact, adviser,
        semester, school_year, course, year_level, section,
        major, date_registered, subjects
    """
    ABBR_COURSE = {
        'BSCS': 'BS Computer Science', 'BSIT': 'BS Information Technology',
        'BSIS': 'BS Information Systems', 'BSCOE': 'BS Computer Engineering',
        'BSECE': 'BS Electronics Engineering', 'BSEE': 'BS Electrical Engineering',
        'BSCE': 'BS Civil Engineering', 'BSME': 'BS Mechanical Engineering',
        'BSED': 'BS Education', 'BSN': 'BS Nursing',
        'BSA': 'BS Accountancy', 'BSBA': 'BS Business Administration',
    }
    PREFIX_COURSE = {
        'CS': 'BS Computer Science', 'IT': 'BS Information Technology',
        'IS': 'BS Information Systems', 'COE': 'BS Computer Engineering',
        'ECE': 'BS Electronics Engineering', 'EE': 'BS Electrical Engineering',
        'CE': 'BS Civil Engineering', 'ME': 'BS Mechanical Engineering',
        'ED': 'BS Education', 'N': 'BS Nursing',
        'A': 'BS Accountancy', 'BA': 'BS Business Administration',
    }
    YEAR_MAP = {
        '1': '1st Year', '2': '2nd Year', '3': '3rd Year',
        '4': '4th Year', '5': '5th Year',
        '1st': '1st Year', '2nd': '2nd Year', '3rd': '3rd Year',
        '4th': '4th Year', '5th': '5th Year',
    }
    SEM_MAP = {
        'FIRST': 'First', 'SECOND': 'Second', 'SUMMER': 'Summer',
        '1ST': 'First', '2ND': 'Second',
    }
 
    result = {
        'student_id': '', 'name': '', 'email': '', 'contact': '',
        'adviser': '', 'semester': '', 'school_year': '',
        'course': '', 'year_level': '', 'section': '',
        'major': '', 'date_registered': '', 'subjects': [],
    }
 
    def next_line(label_re):
        """Value on the line immediately after a label pattern."""
        m = re.search(label_re + r'[^\n]*\n([^\n]+)', full, re.IGNORECASE)
        if not m:
            return ''
        v = (m.group(1) or '').replace('\t', ' ').strip()
        return v
 
    # Student ID
    result['student_id'] = next_line(r'Student\s*(?:No\.?|Number|ID)')
    if not result['student_id']:
        m = re.search(r'\b(\d{4}-\d{4,6})\b', full)
        if m:
            result['student_id'] = m.group(1)
    if not result['student_id']:
        m = re.search(r'\b(\d{7,10})\b', full)
        if m:
            result['student_id'] = m.group(1)
 
    # Semester
    raw_sem  = next_line(r'Semester').upper()
    first_w  = raw_sem.split()[0] if raw_sem else ''
    result['semester'] = SEM_MAP.get(first_w, raw_sem.title()) if raw_sem else ''
 
    # School Year
    result['school_year'] = next_line(r'School\s*Year')
 
    # Student Name → also derive email
    _norm_full = full.replace('\t', ' ')
    _name_m = re.search(
        r'Student\s+Name\s*:\s*((?:[A-Z][A-Z\s.,\'\-]+?)'
        r'(?=\s*(?:Date|Course|Year|Encoder|Major|Section|Address)\s*:|$))',
        _norm_full, re.IGNORECASE | re.DOTALL
    )
    raw_name = ''
    if _name_m:
        raw_name = re.sub(r'\s+', ' ', _name_m.group(1)).strip()
        # Guard: reject if it looks like a label value rather than a name
        if re.match(r'(Date|Course|Year|Encoder)\s*:', raw_name, re.IGNORECASE):
            raw_name = ''
    if not raw_name:
        # Last-resort: grab the next non-empty line after the label
        _fb = re.search(r'Student\s*Name[^\n]*\n([^\n:]+)', _norm_full, re.IGNORECASE)
        if _fb:
            raw_name = _fb.group(1).replace('\t', ' ').strip()
    if raw_name:
        result['name']  = raw_name.title()
        result['email'] = _generate_cvsu_email(raw_name)
 
    # Adviser
    result['adviser'] = next_line(r'Adviser|Advisor')
 
    # Contact
    result['contact'] = next_line(r'Contact\s*(?:No\.?|Number)')
 
    # Date Registered (skip VALIDATION DATE lines)
    raw_date = ''
    for _m in re.finditer(r'Date[^\n]*\n([^\n]+)', full, re.IGNORECASE):
        ctx = full[max(0, _m.start() - 30):_m.start()].upper()
        if any(kw in ctx for kw in ('VALIDAT', 'PAYMENT', 'CONFIR')):
            continue
        raw_date = _m.group(1).replace('\t', ' ').strip()
        break
    if raw_date:
        dm = re.search(r'(\d{1,2})[-\s/]+([A-Za-z]+)[-\s/]+(\d{4})', raw_date)
        if dm:
            try:
                d = datetime.strptime(f"{dm.group(2)} {dm.group(3)}", "%b %Y")
                result['date_registered'] = d.strftime('%Y-%m')
            except Exception:
                pass
        if not result['date_registered']:
            dm2 = re.search(r'([A-Za-z]+)\s+(\d{4})', raw_date)
            if dm2:
                try:
                    d = datetime.strptime(f"{dm2.group(1)} {dm2.group(2)}", "%b %Y")
                    result['date_registered'] = d.strftime('%Y-%m')
                except Exception:
                    pass
 
    # Course
    raw_course = next_line(r'Course').strip().upper()
    result['course'] = ABBR_COURSE.get(raw_course, '')
    if not result['course']:
        for abbr, cname in ABBR_COURSE.items():
            if re.search(r'\b' + abbr + r'\b', full, re.IGNORECASE):
                result['course'] = cname
                break
 
    # Section code (e.g. IT4A → year=4th, sec=A, possibly course)
    raw_section = next_line(r'Section').strip()
    sec_m = re.match(r'^([A-Za-z]+)(\d)([A-Za-z])$', raw_section)
    if sec_m:
        prefix               = sec_m.group(1).upper()
        yr_digit             = sec_m.group(2)
        result['section']    = sec_m.group(3).upper()
        result['year_level'] = YEAR_MAP.get(yr_digit, yr_digit + 'th Year')
        if not result['course']:
            result['course'] = PREFIX_COURSE.get(prefix, '')
    else:
        result['section'] = raw_section[:1].upper() if raw_section else ''
        raw_year  = next_line(r'Year\s*(?:Level)?').strip()
        yr_w      = raw_year.split()[0].lower() if raw_year else ''
        result['year_level'] = (YEAR_MAP.get(yr_w) or
                                YEAR_MAP.get(yr_w.rstrip('thsrnd')) or '')
 
    # Major
    raw_major = next_line(r'Major')
    result['major'] = ('' if raw_major.upper().strip() in ('N/A', 'NA', 'NONE', '', '\u2014')
                       else raw_major.title())
 
    # Subjects — 8-line blocks after Schedule Code header, before Fees section
    subjects = []
    hdr_m  = re.search(
        r'Schedule\s*Code.*?(?:Course\s*)?Description.*?Hour[s]?\s*\n',
        full, re.DOTALL | re.IGNORECASE
    )
    fees_m = re.search(r'Laboratory\s*Fees|Total\s*Units', full, re.IGNORECASE)
    if hdr_m and fees_m and fees_m.start() > hdr_m.end():
        block = full[hdr_m.end():fees_m.start()]
        lines = [l.replace('\t', ' ').strip() for l in block.split('\n') if l.strip()]
        i = 0
        while i < len(lines):
            if re.match(r'^\d{7,12}$', lines[i]) and i + 7 < len(lines):
                code_clean  = re.sub(r'\s+', ' ', lines[i + 2]).strip()
                desc_clean  = lines[i + 3].title()
                try:
                    units_int = str(round(float(lines[i + 4])))
                except Exception:
                    units_int = '3'
                if units_int == '0':
                    units_int = '3'
                if code_clean and desc_clean:
                    subjects.append({
                        'course_code': code_clean,
                        'name':        desc_clean,
                        'units':       units_int,
                    })
                i += 8
            else:
                i += 1
    result['subjects'] = subjects
    return result

# ── Email config helpers ──────────────────────────────────────────────────
def get_email_config():
    """Load SMTP config from DB. Returns dict with all keys."""
    defaults = {
        'smtp_host':     'smtp.gmail.com',
        'smtp_port':     '587',
        'smtp_user':     '',
        'smtp_password': '',
        'smtp_from':     '',
        'enabled':       '0',
    }
    try:
        with get_db() as conn:
            rows = conn.execute('SELECT key, value FROM email_config').fetchall()
            cfg  = dict(defaults)
            for row in rows:
                cfg[row['key']] = row['value']
            return cfg
    except Exception:
        return defaults
 
def save_email_config(cfg: dict):
    """Upsert email config into DB."""
    with get_db() as conn:
        for key, value in cfg.items():
            conn.execute(
                'INSERT INTO email_config (key, value) VALUES (?, ?) '
                'ON CONFLICT(key) DO UPDATE SET value=excluded.value',
                (key, str(value))
            )
 
def _send_email(to_addrs: list, subject: str, html_body: str):
    """
    Send an HTML email via Gmail SMTP in a background thread.
    Silently logs errors — never crashes the main request.
    """
    import threading as _th
    def _worker():
        try:
            import smtplib, ssl
            from email.mime.multipart import MIMEMultipart
            from email.mime.text      import MIMEText
            cfg = get_email_config()
            if cfg.get('enabled') != '1':
                return
            if not cfg.get('smtp_user') or not cfg.get('smtp_password'):
                print('[EMAIL] SMTP credentials not configured — skipping.')
                return
            recipients = [a for a in to_addrs if a and '@' in a]
            if not recipients:
                return
            msg                    = MIMEMultipart('alternative')
            msg['Subject']         = subject
            msg['From']            = cfg.get('smtp_from') or cfg['smtp_user']
            msg['To']              = ', '.join(recipients)
            msg.attach(MIMEText(html_body, 'html'))
            ctx  = ssl.create_default_context()
            port = int(cfg.get('smtp_port', 587))
            with smtplib.SMTP(cfg['smtp_host'], port, timeout=10) as srv:
                srv.ehlo()
                srv.starttls(context=ctx)
                srv.login(cfg['smtp_user'], cfg['smtp_password'])
                srv.sendmail(msg['From'], recipients, msg.as_string())
            print(f'[EMAIL] Sent "{subject}" → {recipients}')
        except Exception as _e:
            print(f'[EMAIL] Failed to send "{subject}": {_e}')
    _th.Thread(target=_worker, daemon=True).start()
 
def send_student_attendance_receipt(
        student_name, student_email, student_id,
        subject_name, section_key, teacher_name,
        tap_time, status, tx_hash, block_num,
        sess_id=None, nfc_id=None):
    """Send attendance receipt email to student."""
    if not student_email or '@' not in student_email:
        return
    status_colors = {
        'present': ('#2D6A27', '#E8F5E9', '✓ Present'),
        'late':    ('#D4A017', '#FFF8E1', '⏱ Late'),
        'absent':  ('#C0392B', '#FFEBEE', '✕ Absent'),
        'excused': ('#2980B9', '#E3F2FD', '◎ Excused'),
    }
    clr, bg, label = status_colors.get(status, ('#333333', '#F5F5F5', status.capitalize()))
    section_display = section_key.replace('|', ' · ') if section_key else '—'
    tx_row = ''
    excuse_section = ''
    
    if status == 'absent' and sess_id and nfc_id:
        try:
            excuse_link = url_for('excuse_submit', sess_id=sess_id, nfc_id=nfc_id, _external=True)
            excuse_section = f'''
            <tr>
              <td colspan="2" style="padding:16px 32px;text-align:center;border-top:1px solid #eee;">
                <div style="font-size:13px;color:#666;margin-bottom:10px;">If this absence is valid, please submit an excuse request below:</div>
                <a href="{excuse_link}" style="display:inline-block;background:#3b82f6;color:#ffffff;text-decoration:none;padding:10px 20px;border-radius:6px;font-weight:bold;font-size:14px;">Submit Excuse Form</a>
              </td>
            </tr>
            '''
        except Exception as e:
            excuse_section = ''

    if tx_hash:
        tx_row = f'''
        <tr>
          <td style="padding:8px 12px;font-size:12px;color:#666;border-bottom:1px solid #eee;">
            Blockchain TX
          </td>
          <td style="padding:8px 12px;font-size:11px;font-family:monospace;
                     color:#2D6A27;border-bottom:1px solid #eee;word-break:break-all;">
            {tx_hash}
          </td>
        </tr>
        <tr>
          <td style="padding:8px 12px;font-size:12px;color:#666;">Block #</td>
          <td style="padding:8px 12px;font-size:12px;font-family:monospace;color:#333;">
            {block_num}
          </td>
        </tr>'''
    html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Calibri,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:32px 16px;">
    <table width="560" cellpadding="0" cellspacing="0"
           style="background:#fff;border-radius:12px;overflow:hidden;
                  box-shadow:0 2px 12px rgba(0,0,0,.1);">
      <!-- Header -->
      <tr>
        <td style="background:#1E4A1A;padding:24px 32px;">
          <div style="font-size:20px;font-weight:700;color:#F5C518;
                      letter-spacing:1px;">DAVS</div>
          <div style="font-size:11px;color:#94a3b8;margin-top:2px;">
            Decentralized Attendance Verification System
          </div>
          <div style="font-size:11px;color:#94a3b8;">
            Cavite State University — Silang Campus
          </div>
        </td>
      </tr>
      <!-- Status banner -->
      <tr>
        <td style="background:{bg};padding:20px 32px;
                   border-left:4px solid {clr};">
          <div style="font-size:28px;font-weight:700;color:{clr};">
            {label}
          </div>
          <div style="font-size:13px;color:#555;margin-top:4px;">
            Your attendance has been recorded for today's class.
          </div>
        </td>
      </tr>
      <!-- Details table -->
      <tr>
        <td style="padding:24px 32px 8px;">
          <div style="font-size:13px;font-weight:700;color:#1E4A1A;
                      text-transform:uppercase;letter-spacing:1px;
                      margin-bottom:12px;">Attendance Details</div>
          <table width="100%" cellpadding="0" cellspacing="0"
                 style="border:1px solid #eee;border-radius:8px;overflow:hidden;">
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;width:140px;">Student</td>
              <td style="padding:8px 12px;font-size:12px;font-weight:600;
                         color:#333;border-bottom:1px solid #eee;">
                {student_name}
                {f'<span style="color:#999;font-size:11px;"> · ID: {student_id}</span>'
                 if student_id else ''}
              </td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Subject</td>
              <td style="padding:8px 12px;font-size:12px;font-weight:600;
                         color:#333;border-bottom:1px solid #eee;">
                {subject_name}
              </td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Section</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;
                         border-bottom:1px solid #eee;">{section_display}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Instructor</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;
                         border-bottom:1px solid #eee;">{teacher_name}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Date & Time</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;
                         border-bottom:1px solid #eee;">{tap_time}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Status</td>
              <td style="padding:8px 12px;">
                <span style="background:{bg};color:{clr};font-weight:700;
                             font-size:12px;padding:3px 10px;border-radius:20px;
                             border:1px solid {clr};">{label}</span>
              </td>
            </tr>
            {tx_row}
          </table>
        </td>
      </tr>
      {excuse_section}
      <!-- Footer -->
      <tr>
        <td style="padding:20px 32px 28px;">
          <div style="font-size:11px;color:#94a3b8;line-height:1.6;">
            This is an automated attendance receipt from the DAVS system.<br>
            {"The TX hash above is your tamper-proof blockchain proof of attendance.<br>" if tx_hash else ""}
            Please do not reply to this email.
          </div>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</body></html>'''
    _send_email(
        [student_email],
        f'[DAVS] Attendance Receipt — {subject_name} ({label})',
        html
    )
 
def send_teacher_session_summary(
        teacher_email, teacher_name,
        subject_name, section_key, time_slot,
        started_at, ended_at,
        present_count, late_count, absent_count, excused_count,
        student_rows):
    """
    Send session summary email to teacher when session ends.
    student_rows: list of dicts with keys:
        name, student_id, status, tap_time, tx_hash, block_num
    """
    if not teacher_email or '@' not in teacher_email:
        return
    total        = present_count + late_count + absent_count + excused_count
    rate         = round((present_count + late_count) / total * 100, 1) if total else 0
    section_disp = section_key.replace('|', ' · ') if section_key else '—'
    status_colors = {
        'present': ('#2D6A27', '#E8F5E9', '✓ Present'),
        'late':    ('#D4A017', '#FFF8E1', '⏱ Late'),
        'absent':  ('#C0392B', '#FFEBEE', '✕ Absent'),
        'excused': ('#2980B9', '#E3F2FD', '◎ Excused'),
    }
    rows_html = ''
    for i, st in enumerate(student_rows):
        clr, bg, lbl = status_colors.get(st.get('status','absent'),
                                          ('#333','#f5f5f5', st.get('status','—').capitalize()))
        tx = st.get('tx_hash','')
        tx_cell = (f'<span style="font-family:monospace;font-size:10px;color:#2D6A27;">'
                   f'{tx[:20]}…</span>') if tx else '—'
        bg_row = '#F9FBF9' if i % 2 == 0 else '#FFFFFF'
        rows_html += f'''<tr style="background:{bg_row};">
          <td style="padding:7px 10px;font-size:12px;border-bottom:1px solid #eee;">
            {st.get("name","—")}
            <div style="font-size:10px;color:#999;">{st.get("student_id","")}</div>
          </td>
          <td style="padding:7px 10px;font-size:11px;color:#666;
                     border-bottom:1px solid #eee;white-space:nowrap;">
            {st.get("tap_time","—")}
          </td>
          <td style="padding:7px 10px;border-bottom:1px solid #eee;">
            <span style="background:{bg};color:{clr};font-weight:700;
                         font-size:11px;padding:2px 8px;border-radius:20px;
                         border:1px solid {clr};">{lbl}</span>
          </td>
          <td style="padding:7px 10px;font-size:11px;border-bottom:1px solid #eee;">
            {tx_cell}
          </td>
        </tr>'''
    html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Calibri,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:32px 16px;">
    <table width="640" cellpadding="0" cellspacing="0"
           style="background:#fff;border-radius:12px;overflow:hidden;
                  box-shadow:0 2px 12px rgba(0,0,0,.1);">
      <!-- Header -->
      <tr>
        <td style="background:#1E4A1A;padding:24px 32px;">
          <div style="font-size:20px;font-weight:700;color:#F5C518;">DAVS</div>
          <div style="font-size:11px;color:#94a3b8;margin-top:2px;">
            Session Summary Report — {subject_name}
          </div>
        </td>
      </tr>
      <!-- Summary stats -->
      <tr>
        <td style="padding:20px 32px 8px;">
          <div style="font-size:13px;font-weight:700;color:#1E4A1A;
                      text-transform:uppercase;letter-spacing:1px;
                      margin-bottom:12px;">Session Overview</div>
          <table width="100%" cellpadding="0" cellspacing="0"
                 style="border:1px solid #eee;border-radius:8px;
                        overflow:hidden;margin-bottom:16px;">
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;width:140px;">Subject</td>
              <td style="padding:8px 12px;font-size:12px;font-weight:600;
                         color:#333;border-bottom:1px solid #eee;">{subject_name}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Section</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;
                         border-bottom:1px solid #eee;">{section_disp}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Time Slot</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;
                         border-bottom:1px solid #eee;">{time_slot or "—"}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;
                         border-bottom:1px solid #eee;">Started</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;
                         border-bottom:1px solid #eee;">{started_at}</td>
            </tr>
            <tr>
              <td style="padding:8px 12px;font-size:12px;color:#666;">Ended</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;">{ended_at}</td>
            </tr>
          </table>
          <!-- Stat boxes -->
          <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px;">
            <tr>
              <td width="25%" style="padding:4px;">
                <div style="background:#E8F5E9;border:1px solid #2D6A27;border-radius:8px;
                            padding:12px;text-align:center;">
                  <div style="font-size:28px;font-weight:700;color:#2D6A27;">{present_count}</div>
                  <div style="font-size:11px;color:#2D6A27;font-weight:600;">Present</div>
                </div>
              </td>
              <td width="25%" style="padding:4px;">
                <div style="background:#FFF8E1;border:1px solid #D4A017;border-radius:8px;
                            padding:12px;text-align:center;">
                  <div style="font-size:28px;font-weight:700;color:#D4A017;">{late_count}</div>
                  <div style="font-size:11px;color:#D4A017;font-weight:600;">Late</div>
                </div>
              </td>
              <td width="25%" style="padding:4px;">
                <div style="background:#FFEBEE;border:1px solid #C0392B;border-radius:8px;
                            padding:12px;text-align:center;">
                  <div style="font-size:28px;font-weight:700;color:#C0392B;">{absent_count}</div>
                  <div style="font-size:11px;color:#C0392B;font-weight:600;">Absent</div>
                </div>
              </td>
              <td width="25%" style="padding:4px;">
                <div style="background:#E3F2FD;border:1px solid #2980B9;border-radius:8px;
                            padding:12px;text-align:center;">
                  <div style="font-size:28px;font-weight:700;color:#2980B9;">{excused_count}</div>
                  <div style="font-size:11px;color:#2980B9;font-weight:600;">Excused</div>
                </div>
              </td>
            </tr>
          </table>
          <div style="font-size:12px;color:#555;margin-bottom:20px;">
            Attendance rate: <strong style="color:#1E4A1A;">{rate}%</strong>
            &nbsp;·&nbsp; {total} students enrolled
          </div>
          <!-- Student list -->
          <div style="font-size:13px;font-weight:700;color:#1E4A1A;
                      text-transform:uppercase;letter-spacing:1px;
                      margin-bottom:10px;">Student Attendance List</div>
          <table width="100%" cellpadding="0" cellspacing="0"
                 style="border:1px solid #eee;border-radius:8px;overflow:hidden;">
            <thead>
              <tr style="background:#1E4A1A;">
                <th style="padding:9px 10px;font-size:11px;color:#fff;
                           text-align:left;font-weight:600;">Student</th>
                <th style="padding:9px 10px;font-size:11px;color:#fff;
                           text-align:left;font-weight:600;">Tap Time</th>
                <th style="padding:9px 10px;font-size:11px;color:#fff;
                           text-align:left;font-weight:600;">Status</th>
                <th style="padding:9px 10px;font-size:11px;color:#fff;
                           text-align:left;font-weight:600;">TX Hash</th>
              </tr>
            </thead>
            <tbody>{rows_html}</tbody>
          </table>
        </td>
      </tr>
      <!-- Footer -->
      <tr>
        <td style="padding:16px 32px 28px;">
          <div style="font-size:11px;color:#94a3b8;line-height:1.6;">
            This is an automated session summary from the DAVS system.<br>
            All TX hashes are immutable blockchain records verifiable on the Hardhat network.<br>
            Please do not reply to this email.
          </div>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</body></html>'''
    _send_email(
        [teacher_email],
        f'[DAVS] Session Summary — {subject_name} · {section_disp}',
        html
    )

@app.context_processor
def inject_globals():
    return dict(
        photos_db     = db_get_all_photos(),
        pending_count = db_pending_count(),
        fmt_time      = fmt_time,
        fmt_time_short= fmt_time_short,
    )

hardhat_url = "http://127.0.0.1:8545"
web3 = Web3(Web3.HTTPProvider(hardhat_url))

BLOCKCHAIN_ONLINE = web3.is_connected()
if BLOCKCHAIN_ONLINE:
    print("[OK] Connected to Hardhat Network")
else:
    print("[WARNING] Hardhat not running — students will load from SQLite cache.")

contract_data_path = os.path.join(os.path.dirname(__file__), 'attendance-contract.json')
try:
    with open(contract_data_path) as f:
        contract_data = json.load(f)
    contract      = web3.eth.contract(address=contract_data['address'], abi=contract_data['abi'])
    admin_account = web3.eth.accounts[0] if BLOCKCHAIN_ONLINE else None
except Exception as _ce:
    print(f"[WARNING] Could not load contract: {_ce}")
    contract      = None
    admin_account = None
    BLOCKCHAIN_ONLINE = False
    print("[INFO] Offline mode active: contract file missing or unreadable.")

BASE_DIR      = os.path.dirname(__file__)
DB_FILE       = os.path.join(BASE_DIR, 'davs.db')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
UPLOAD_FOLDER_EXCUSES = os.path.join(BASE_DIR, 'static', 'uploads', 'excuses')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_EXCUSES, exist_ok=True)

ALLOWED_EXTENSIONS_EXCUSES = {'png', 'jpg', 'jpeg', 'pdf'}

EXCUSE_REASONS = [
    ('sickness', 'Sickness / Illness'),
    ('lbm', 'LBM'),
    ('emergency', 'Family Emergency'),
    ('bereavement', 'Bereavement'),
    ('medical', 'Medical Appointment'),
    ('accident', 'Accident / Injury'),
    ('official', 'Official School Business'),
    ('weather', 'Extreme Weather / Calamity'),
    ('transport', 'Transportation Problem'),
    ('others', 'Others')
]

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def normalize_section_key(key):
    if not key:
        return key
    parts = [p.strip() for p in key.split('|')]
    if len(parts) == 3:
        course    = parts[0].strip()
        year      = parts[1].strip()
        section   = parts[2].strip().upper()
        year_map = {
            '1': '1st Year', '2': '2nd Year', '3': '3rd Year', '4': '4th Year', '5': '5th Year',
            '1st': '1st Year', '2nd': '2nd Year', '3rd': '3rd Year', '4th': '4th Year',
            '1st year': '1st Year', '2nd year': '2nd Year', '3rd year': '3rd Year',
            '4th year': '4th Year', '5th year': '5th Year',
        }
        year_normalized = year_map.get(year.lower(), year)
        course_canonical = normalize_course_name(course)
        return f"{course_canonical}|{year_normalized}|{section}"
    return key

def build_student_section_key(student):
    course     = (student.get('course') or '').strip()
    year_level = (student.get('year_level') or '').strip()
    section    = (student.get('section') or '').strip().upper()
    if not course or not year_level or not section:
        return None
    return normalize_section_key(f"{course}|{year_level}|{section}")

def generate_cvsu_email(name, provided_email=''):
    """
    Generate CVSU email pattern: sc.firstname.lastname@cvsu.edu.ph
    If provided_email is given and not empty, return it.
    Otherwise, generate from name.
    """
    provided_email = provided_email.strip()
    if provided_email:
        return provided_email

    # Generate CVSU pattern email: sc.firstname.lastname@cvsu.edu.ph
    clean = re.sub(r'\b[A-Za-z]\.\s*', '', name).strip()
    clean = re.sub(r'\b(JR|SR|II|III|IV)\.?\b', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s+', ' ', clean)
    words = clean.split()
    if len(words) >= 2:
        first_slug = ''.join(re.sub(r'[^a-z]', '', w.lower()) for w in words[:-1])
        last_slug  = re.sub(r'[^a-z]', '', words[-1].lower())
        if first_slug and last_slug:
            return f'sc.{first_slug}.{last_slug}@cvsu.edu.ph'
    return ''

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    sql = """
    CREATE TABLE IF NOT EXISTS accounts (
        username        TEXT PRIMARY KEY,
        password_hash   TEXT NOT NULL,
        role            TEXT NOT NULL DEFAULT 'teacher',
        full_name       TEXT NOT NULL DEFAULT '',
        email           TEXT NOT NULL DEFAULT '',
        status          TEXT NOT NULL DEFAULT 'pending',
        sections_json   TEXT NOT NULL DEFAULT '[]',
        my_subjects_json TEXT NOT NULL DEFAULT '[]',
        created_at      TEXT NOT NULL DEFAULT '',
        updated_at      TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS subjects (
        subject_id  TEXT PRIMARY KEY,
        name        TEXT NOT NULL,
        course_code TEXT NOT NULL DEFAULT '',
        units       TEXT NOT NULL DEFAULT '3',
        created_by  TEXT NOT NULL DEFAULT '',
        created_at  TEXT NOT NULL DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_subj_code ON subjects(course_code);
    CREATE TABLE IF NOT EXISTS students (
        nfc_id          TEXT PRIMARY KEY,
        full_name       TEXT NOT NULL DEFAULT '',
        student_id      TEXT NOT NULL DEFAULT '',
        program         TEXT NOT NULL DEFAULT '',
        year_level      TEXT NOT NULL DEFAULT '',
        section         TEXT NOT NULL DEFAULT '',
        adviser         TEXT NOT NULL DEFAULT '',
        email           TEXT NOT NULL DEFAULT '',
        contact         TEXT NOT NULL DEFAULT '',
        major           TEXT NOT NULL DEFAULT '',
        semester        TEXT NOT NULL DEFAULT '',
        school_year     TEXT NOT NULL DEFAULT '',
        date_registered TEXT NOT NULL DEFAULT '',
        raw_name        TEXT NOT NULL DEFAULT '',
        eth_address     TEXT NOT NULL DEFAULT '',
        reg_tx_hash     TEXT NOT NULL DEFAULT '',
        reg_block       INTEGER NOT NULL DEFAULT 0,
        photo_file      TEXT NOT NULL DEFAULT '',
        created_at      TEXT NOT NULL DEFAULT '',
        updated_at      TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS sessions (
            sess_id         TEXT PRIMARY KEY,
            subject_id      TEXT NOT NULL DEFAULT '',
            subject_name    TEXT NOT NULL DEFAULT '',
            course_code     TEXT NOT NULL DEFAULT '',
            units           INTEGER NOT NULL DEFAULT 3,
            time_slot       TEXT NOT NULL DEFAULT '',
            section_key     TEXT NOT NULL DEFAULT '',
            teacher_username TEXT NOT NULL DEFAULT '',
            teacher_name    TEXT NOT NULL DEFAULT '',
            started_at      TEXT NOT NULL DEFAULT '',
            late_cutoff     TEXT NOT NULL DEFAULT '',
            auto_end_at     TEXT,
            ended_at        TEXT,
            grace_period    INTEGER NOT NULL DEFAULT 15,
            schedule_id     TEXT DEFAULT NULL,
            total_enrolled  INTEGER NOT NULL DEFAULT 0,
            total_present   INTEGER NOT NULL DEFAULT 0,
            total_late      INTEGER NOT NULL DEFAULT 0,
            total_absent    INTEGER NOT NULL DEFAULT 0,
            total_excused   INTEGER NOT NULL DEFAULT 0,
            warn_log_json   TEXT NOT NULL DEFAULT '[]',
            invalid_log_json TEXT NOT NULL DEFAULT '[]'
        );
    CREATE INDEX IF NOT EXISTS idx_sess_ended   ON sessions(ended_at);
    CREATE INDEX IF NOT EXISTS idx_sess_section ON sessions(section_key);
    CREATE TABLE IF NOT EXISTS attendance_logs (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        sess_id      TEXT NOT NULL,
        nfc_id       TEXT NOT NULL,
        student_name TEXT NOT NULL DEFAULT '',
        student_id   TEXT NOT NULL DEFAULT '',
        status       TEXT NOT NULL DEFAULT 'absent',
        tap_time     TEXT NOT NULL DEFAULT '',
        tx_hash      TEXT NOT NULL DEFAULT '',
        block_number INTEGER NOT NULL DEFAULT 0,
        excuse_note  TEXT NOT NULL DEFAULT '',
        created_at   TEXT NOT NULL DEFAULT '',
        UNIQUE(sess_id, nfc_id)
    );
    CREATE INDEX IF NOT EXISTS idx_att_sess  ON attendance_logs(sess_id);
    CREATE INDEX IF NOT EXISTS idx_att_nfc   ON attendance_logs(nfc_id);
    CREATE INDEX IF NOT EXISTS idx_att_status ON attendance_logs(status);
    CREATE TABLE IF NOT EXISTS photos (
        person_id   TEXT PRIMARY KEY,
        filename    TEXT NOT NULL,
        uploaded_at TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS nfc_scanner (
        id           INTEGER PRIMARY KEY CHECK (id = 1),
        waiting      INTEGER NOT NULL DEFAULT 0,
        scanned_uid  TEXT NOT NULL DEFAULT '',
        requested_by TEXT NOT NULL DEFAULT '',
        requested_at TEXT NOT NULL DEFAULT ''
    );
    INSERT OR IGNORE INTO nfc_scanner (id, waiting, scanned_uid) VALUES (1, 0, '');
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY, password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'teacher', full_name TEXT NOT NULL DEFAULT '',
        email TEXT NOT NULL DEFAULT '', status TEXT NOT NULL DEFAULT 'pending',
        sections_json TEXT NOT NULL DEFAULT '[]',
        my_subjects_json TEXT NOT NULL DEFAULT '[]',
        created_at TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS nfc_registration (
        id INTEGER PRIMARY KEY CHECK (id = 1),
        waiting INTEGER NOT NULL DEFAULT 0,
        scanned_uid TEXT NOT NULL DEFAULT '',
        requested_by TEXT NOT NULL DEFAULT '',
        requested_at TEXT NOT NULL DEFAULT ''
    );
    INSERT OR IGNORE INTO nfc_registration (id, waiting, scanned_uid) VALUES (1, 0, '');
    CREATE TABLE IF NOT EXISTS student_overrides (
        nfc_id TEXT PRIMARY KEY, full_name TEXT DEFAULT '',
        student_id TEXT DEFAULT '', email TEXT DEFAULT '',
        contact TEXT DEFAULT '', adviser TEXT DEFAULT '',
        major TEXT DEFAULT '', semester TEXT DEFAULT '',
        school_year TEXT DEFAULT '', date_registered TEXT DEFAULT '',
        course TEXT DEFAULT '', year_level TEXT DEFAULT '',
        section TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS email_config (
        key   TEXT PRIMARY KEY,
        value TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS schedules (
        schedule_id      TEXT PRIMARY KEY,
        section_key      TEXT NOT NULL DEFAULT '',
        subject_id       TEXT NOT NULL DEFAULT '',
        subject_name     TEXT NOT NULL DEFAULT '',
        course_code      TEXT NOT NULL DEFAULT '',
        teacher_username TEXT NOT NULL DEFAULT '',
        teacher_name     TEXT NOT NULL DEFAULT '',
        day_of_week      INTEGER NOT NULL DEFAULT 1,
        start_time       TEXT NOT NULL DEFAULT '',
        end_time         TEXT NOT NULL DEFAULT '',
        grace_minutes    INTEGER NOT NULL DEFAULT 15,
        is_active        INTEGER NOT NULL DEFAULT 1,
        created_by       TEXT NOT NULL DEFAULT '',
        created_at       TEXT NOT NULL DEFAULT '',
        updated_at       TEXT NOT NULL DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_sched_teacher ON schedules(teacher_username);
    CREATE INDEX IF NOT EXISTS idx_sched_day     ON schedules(day_of_week);
    CREATE TABLE IF NOT EXISTS excuse_requests (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        sess_id         TEXT NOT NULL DEFAULT '',
        nfc_id          TEXT NOT NULL DEFAULT '',
        student_name    TEXT NOT NULL DEFAULT '',
        student_id      TEXT NOT NULL DEFAULT '',
        student_email   TEXT NOT NULL DEFAULT '',
        reason_type     TEXT NOT NULL DEFAULT '',
        reason_detail   TEXT NOT NULL DEFAULT '',
        attachment_file TEXT NOT NULL DEFAULT '',
        status          TEXT NOT NULL DEFAULT 'pending',
        reviewed_by     TEXT NOT NULL DEFAULT '',
        reviewed_at     TEXT NOT NULL DEFAULT '',
        created_at      TEXT NOT NULL DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_excuse_sess   ON excuse_requests(sess_id);
    CREATE INDEX IF NOT EXISTS idx_excuse_status ON excuse_requests(status);
    """
    with get_db() as conn:
        conn.executescript(sql)
    _migrate_add_missing_columns()
    _migrate_users_to_accounts()
    _migrate_nfc_registration()
    print("[DB] Schema ready ->", DB_FILE)

def _migrate_add_missing_columns():
    migrations = [
        ("students", "program",         "TEXT NOT NULL DEFAULT ''"),
        ("students", "full_name",       "TEXT NOT NULL DEFAULT ''"),
        ("students", "reg_tx_hash",     "TEXT NOT NULL DEFAULT ''"),
        ("students", "reg_block",       "INTEGER NOT NULL DEFAULT 0"),
        ("students", "photo_file",      "TEXT NOT NULL DEFAULT ''"),
        ("students", "updated_at",      "TEXT NOT NULL DEFAULT ''"),
        ("sessions", "teacher_username","TEXT NOT NULL DEFAULT ''"),
        ("sessions", "total_enrolled",  "INTEGER NOT NULL DEFAULT 0"),
        ("sessions", "total_present",   "INTEGER NOT NULL DEFAULT 0"),
        ("sessions", "total_late",      "INTEGER NOT NULL DEFAULT 0"),
        ("sessions", "total_absent",    "INTEGER NOT NULL DEFAULT 0"),
        ("sessions", "total_excused",   "INTEGER NOT NULL DEFAULT 0"),
        ("sessions", "warn_log_json",   "TEXT NOT NULL DEFAULT '[]'"),
        ("sessions", "invalid_log_json","TEXT NOT NULL DEFAULT '[]'"),
        ("sessions", "grace_period",    "INTEGER NOT NULL DEFAULT 15"),
        ("sessions", "auto_end_at",     "TEXT"),
        ("sessions", "schedule_id",     "TEXT DEFAULT NULL"),
        ("accounts", "updated_at",      "TEXT NOT NULL DEFAULT ''"),
        ("photos",   "uploaded_at",     "TEXT NOT NULL DEFAULT ''"),
        ("attendance_logs", "excuse_request_id", "INTEGER DEFAULT NULL"),
        ("schedules", "section_key",      "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "subject_id",       "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "subject_name",     "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "course_code",      "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "teacher_username", "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "teacher_name",     "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "day_of_week",      "INTEGER NOT NULL DEFAULT 1"),
        ("schedules", "start_time",       "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "end_time",         "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "grace_minutes",    "INTEGER NOT NULL DEFAULT 15"),
        ("schedules", "is_active",        "INTEGER NOT NULL DEFAULT 1"),
        ("schedules", "created_by",       "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "created_at",       "TEXT NOT NULL DEFAULT ''"),
        ("schedules", "updated_at",       "TEXT NOT NULL DEFAULT ''"),
    ]
    with get_db() as conn:
        # ── Ensure schedules table exists (created after initial DB may exist) ──
        existing_tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]

        if 'schedules' not in existing_tables:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS schedules (
                    schedule_id      TEXT PRIMARY KEY,
                    section_key      TEXT NOT NULL DEFAULT '',
                    subject_id       TEXT NOT NULL DEFAULT '',
                    subject_name     TEXT NOT NULL DEFAULT '',
                    course_code      TEXT NOT NULL DEFAULT '',
                    teacher_username TEXT NOT NULL DEFAULT '',
                    teacher_name     TEXT NOT NULL DEFAULT '',
                    day_of_week      INTEGER NOT NULL DEFAULT 1,
                    start_time       TEXT NOT NULL DEFAULT '',
                    end_time         TEXT NOT NULL DEFAULT '',
                    grace_minutes    INTEGER NOT NULL DEFAULT 15,
                    is_active        INTEGER NOT NULL DEFAULT 1,
                    created_by       TEXT NOT NULL DEFAULT '',
                    created_at       TEXT NOT NULL DEFAULT '',
                    updated_at       TEXT NOT NULL DEFAULT ''
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_sched_teacher ON schedules(teacher_username)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_sched_day     ON schedules(day_of_week)")
            print("[MIGRATION] Created missing table: schedules")

        if 'excuse_requests' not in existing_tables:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS excuse_requests (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    sess_id         TEXT NOT NULL DEFAULT '',
                    nfc_id          TEXT NOT NULL DEFAULT '',
                    student_name    TEXT NOT NULL DEFAULT '',
                    student_id      TEXT NOT NULL DEFAULT '',
                    student_email   TEXT NOT NULL DEFAULT '',
                    reason_type     TEXT NOT NULL DEFAULT '',
                    reason_detail   TEXT NOT NULL DEFAULT '',
                    attachment_file TEXT NOT NULL DEFAULT '',
                    status          TEXT NOT NULL DEFAULT 'pending',
                    reviewed_by     TEXT NOT NULL DEFAULT '',
                    reviewed_at     TEXT NOT NULL DEFAULT '',
                    created_at      TEXT NOT NULL DEFAULT ''
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_excuse_sess   ON excuse_requests(sess_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_excuse_status ON excuse_requests(status)")
            print("[MIGRATION] Created missing table: excuse_requests")

        for table, col, col_def in migrations:
            try:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_def}")
                print(f"[MIGRATION] Added {table}.{col}")
            except Exception:
                pass
        try:
            existing = [r[1] for r in conn.execute("PRAGMA table_info(students)").fetchall()]
            if 'course' in existing and 'program' in existing:
                conn.execute("UPDATE students SET program = course WHERE program = '' AND course != ''")
        except Exception as e:
            print(f"[MIGRATION] course->program copy: {e}")
        try:
            existing = [r[1] for r in conn.execute("PRAGMA table_info(students)").fetchall()]
            if 'name' in existing and 'full_name' in existing:
                conn.execute("UPDATE students SET full_name = name WHERE full_name = '' AND name != ''")
        except Exception as e:
            print(f"[MIGRATION] name->full_name copy: {e}")
        try:
            existing = [r[1] for r in conn.execute("PRAGMA table_info(students)").fetchall()]
            if 'tx_hash' in existing and 'reg_tx_hash' in existing:
                conn.execute("UPDATE students SET reg_tx_hash = tx_hash WHERE reg_tx_hash = '' AND tx_hash != ''")
        except Exception as e:
            print(f"[MIGRATION] tx_hash->reg_tx_hash copy: {e}")
        try:
            existing = [r[1] for r in conn.execute("PRAGMA table_info(sessions)").fetchall()]
            if 'teacher' in existing and 'teacher_username' in existing:
                conn.execute("UPDATE sessions SET teacher_username = teacher WHERE teacher_username = '' AND teacher != ''")
        except Exception as e:
            print(f"[MIGRATION] teacher->teacher_username copy: {e}")
        try:
            conn.execute("CREATE INDEX IF NOT EXISTS idx_stu_program ON students(program)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_stu_section ON students(year_level, section)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_sess_teacher ON sessions(teacher_username)")
        except Exception as e:
            print(f"[MIGRATION] Index creation: {e}")

def _migrate_users_to_accounts():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM users").fetchall()
        for r in rows:
            d = dict(r)
            conn.execute(
                "INSERT OR IGNORE INTO accounts "
                "(username,password_hash,role,full_name,email,status,"
                " sections_json,my_subjects_json,created_at,updated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (d['username'], d['password'], d.get('role','teacher'),
                 d.get('full_name',''), d.get('email',''), d.get('status','pending'),
                 d.get('sections_json','[]'), d.get('my_subjects_json','[]'),
                 d.get('created_at',''), d.get('created_at',''))
            )

def _migrate_nfc_registration():
    with get_db() as conn:
        row = conn.execute("SELECT * FROM nfc_registration WHERE id=1").fetchone()
        if row:
            conn.execute(
                "INSERT OR REPLACE INTO nfc_scanner "
                "(id,waiting,scanned_uid,requested_by,requested_at) VALUES (?,?,?,?,?)",
                (1, row['waiting'], row['scanned_uid'],
                 row['requested_by'], row['requested_at'])
            )

def _row_to_dict(row):
    if row is None: return None
    d = dict(row)
    for col in ('present_json','late_json','excused_json','warned_json','absent_json',
                'tap_log_json','warn_log_json','invalid_log_json'):
        key = col.replace('_json','')
        if col in d:
            d[key] = json.loads(d.pop(col) or '[]')
        elif key not in d:
            d[key] = []
    d['excuse_notes'] = json.loads(d.pop('excuse_notes_json','{}') or '{}') if 'excuse_notes_json' in d else {}
    d['tx_hashes']    = json.loads(d.pop('tx_hashes_json',   '{}') or '{}') if 'tx_hashes_json'    in d else {}
    if 'teacher_username' in d and 'teacher' not in d:
        d['teacher'] = d.pop('teacher_username')
    elif 'teacher_username' in d:
        d.pop('teacher_username')
    if d.get('section_key'):
        d['section_key'] = normalize_section_key(d['section_key'])
    return d

def load_sessions():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM sessions").fetchall()
        result = {}
        for r in rows:
            s = _session_row_with_logs(conn, r)
            result[r['sess_id']] = s
    return result

def load_session(sess_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM sessions WHERE sess_id=?", (sess_id,)).fetchone()
        if row is None: return None
        return _session_row_with_logs(conn, row)

def _session_row_with_logs(conn, row):
    d = dict(row)
    if 'teacher_username' in d:
        d['teacher'] = d.pop('teacher_username')
    if d.get('section_key'):
        d['section_key'] = normalize_section_key(d['section_key'])
    logs = conn.execute(
        "SELECT * FROM attendance_logs WHERE sess_id=?", (d['sess_id'],)
    ).fetchall()
    present, late, excused, absent = [], [], [], []
    tap_log, excuse_notes, tx_hashes = [], {}, {}
    for lg in logs:
        nid = lg['nfc_id']
        st  = lg['status']
        if   st == 'excused': excused.append(nid); excuse_notes[nid] = lg['excuse_note']
        elif st == 'late':    late.append(nid);    present.append(nid)
        elif st == 'present': present.append(nid)
        elif st == 'absent':  absent.append(nid)
        if lg['tap_time'] and st in ('present','late'):
            tap_log.append({
                'nfc_id': nid, 'name': lg['student_name'],
                'student_id': lg['student_id'], 'time': lg['tap_time'],
                'tx_hash': lg['tx_hash'], 'block': lg['block_number'],
                'is_late': st == 'late', 'timestamp': 0,
            })
        if lg['tx_hash']:
            tx_hashes[nid] = {'tx_hash': lg['tx_hash'],
                              'block': lg['block_number'],
                              'time': lg['tap_time']}
    d['present']      = present
    d['late']         = late
    d['excused']      = excused
    d['absent']       = absent
    d['warned']       = []
    d['tap_log']      = tap_log
    d['warn_log']     = json.loads(d.pop('warn_log_json',  '[]') or '[]')
    d['invalid_log']  = json.loads(d.pop('invalid_log_json','[]') or '[]')
    d['excuse_notes'] = excuse_notes
    d['tx_hashes']    = tx_hashes
    return d

def save_session(sess_id, s):
    sk = normalize_section_key(s.get('section_key', ''))
    teacher_uname = s.get('teacher_username') or s.get('teacher') or ''
    teacher_name  = s.get('teacher_name', '')
    schedule_id   = s.get('schedule_id')
    with get_db() as conn:
        conn.execute(
            "INSERT INTO sessions "
            "(sess_id,subject_id,subject_name,course_code,units,time_slot,"
            " section_key,teacher_username,teacher_name,started_at,late_cutoff,"
            " auto_end_at,ended_at,grace_period,schedule_id,"
            " warn_log_json,invalid_log_json) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(sess_id) DO UPDATE SET "
            "subject_id=excluded.subject_id, subject_name=excluded.subject_name, "
            "course_code=excluded.course_code, units=excluded.units, "
            "time_slot=excluded.time_slot, section_key=excluded.section_key, "
            "teacher_username=excluded.teacher_username, "
            "teacher_name=excluded.teacher_name, "
            "started_at=excluded.started_at, late_cutoff=excluded.late_cutoff, "
            "auto_end_at=excluded.auto_end_at, "
            "ended_at=excluded.ended_at, "
            "grace_period=excluded.grace_period, "
            "schedule_id=excluded.schedule_id, "
            "warn_log_json=excluded.warn_log_json, "
            "invalid_log_json=excluded.invalid_log_json",
            (sess_id, s.get('subject_id', ''), s.get('subject_name', ''),
             s.get('course_code', ''), s.get('units', 3), s.get('time_slot', ''),
             sk, teacher_uname, teacher_name,
             s.get('started_at', ''), s.get('late_cutoff', ''),
             s.get('auto_end_at'), s.get('ended_at'),
             s.get('grace_period', 15), schedule_id,
             json.dumps(s.get('warn_log', [])), json.dumps(s.get('invalid_log', [])))
        )
        counts = conn.execute(
            "SELECT status, COUNT(*) as cnt FROM attendance_logs "
            "WHERE sess_id=? GROUP BY status", (sess_id,)
        ).fetchall()
        totals = {r['status']: r['cnt'] for r in counts}
        conn.execute(
            "UPDATE sessions SET total_present=?,total_late=?,total_absent=?,total_excused=? "
            "WHERE sess_id=?",
            (totals.get('present', 0) + totals.get('late', 0),
             totals.get('late', 0), totals.get('absent', 0),
             totals.get('excused', 0), sess_id)
        )

def save_sessions(sessions_dict):
    for sid, s in sessions_dict.items():
        save_session(sid, s)

def migrate_json_to_sqlite():
    for fname, table, saver in [
        ('users.json',             'users',             lambda d: [db_save_user(u,v) for u,v in d.items()]),
        ('subjects.json',          'subjects',          lambda d: [db_save_subject(k,v) for k,v in d.items()]),
        ('student_photos.json',    'photos',            lambda d: [db_save_photo(k,v) for k,v in d.items()]),
        ('student_overrides.json', 'student_overrides', lambda d: [db_save_override(k,v) for k,v in d.items()]),
        ('sessions.json',          'sessions',          lambda d: save_sessions(d)),
    ]:
        fpath = os.path.join(BASE_DIR, fname)
        if os.path.exists(fpath):
            try:
                old = json.load(open(fpath))
                with get_db() as conn:
                    count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                if count == 0 and old:
                    print(f"[DB] Migrating {fname} -> {table}...")
                    saver(old)
                    os.rename(fpath, fpath + '.migrated')
            except Exception as e:
                print(f"[DB] Migration warning for {fname}: {e}")
    if db_get_user('admin') is None:
        db_save_user('admin', {
            'username':'admin','password':hash_password('admin123'),'role':'admin',
            'full_name':'System Administrator','email':'admin@davs.edu',
            'status':'approved','sections':[],'my_subjects':[],
            'created_at':datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    if db_get_user('superadmin') is None:
        db_save_user('superadmin', {
            'username':'superadmin','password':hash_password('Admin@DAVS2024'),
            'role':'super_admin','full_name':'Super Administrator',
            'email':'superadmin@davs.edu',
            'status':'approved','sections':[],'my_subjects':[],
            'created_at':datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        print('[DB] Default superadmin account created (change password immediately!)')

def _account_row(row):
    if row is None: return None
    d = dict(row)
    if 'password_hash' in d:
        d['password'] = d.pop('password_hash')
    raw_sections     = json.loads(d.pop('sections_json',   '[]') or '[]')
    d['my_subjects'] = json.loads(d.pop('my_subjects_json','[]') or '[]')
    d['sections']    = [normalize_section_key(s) for s in raw_sections]
    return d

def _user_row(row):
    return _account_row(row)

def db_get_all_users():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM accounts").fetchall()
    return {r['username']: _account_row(r) for r in rows}

def db_get_user(username):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM accounts WHERE username=?", (username,)).fetchone()
    return _account_row(row)

def db_save_user(username, u):
    sections = [normalize_section_key(s) for s in u.get('sections', [])]
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    pw  = u.get('password', u.get('password_hash', ''))
    with get_db() as conn:
        conn.execute(
            "INSERT INTO accounts "
            "(username,password_hash,role,full_name,email,status,"
            " sections_json,my_subjects_json,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(username) DO UPDATE SET "
            "password_hash=excluded.password_hash, role=excluded.role, "
            "full_name=excluded.full_name, email=excluded.email, "
            "status=excluded.status, sections_json=excluded.sections_json, "
            "my_subjects_json=excluded.my_subjects_json, updated_at=excluded.updated_at",
            (username, pw, u.get('role','teacher'),
             u.get('full_name',''), u.get('email',''), u.get('status','pending'),
             json.dumps(sections), json.dumps(u.get('my_subjects',[])),
             u.get('created_at', now), now)
        )
        conn.execute(
            "INSERT INTO users "
            "(username,password,role,full_name,email,status,"
            " sections_json,my_subjects_json,created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(username) DO UPDATE SET "
            "password=excluded.password, role=excluded.role, "
            "full_name=excluded.full_name, email=excluded.email, "
            "status=excluded.status, sections_json=excluded.sections_json, "
            "my_subjects_json=excluded.my_subjects_json",
            (username, pw, u.get('role','teacher'),
             u.get('full_name',''), u.get('email',''), u.get('status','pending'),
             json.dumps(sections), json.dumps(u.get('my_subjects',[])),
             u.get('created_at', now))
        )

def db_delete_user(username):
    with get_db() as conn:
        conn.execute("DELETE FROM accounts WHERE username=?", (username,))
        conn.execute("DELETE FROM users    WHERE username=?", (username,))

def db_pending_count():
    with get_db() as conn:
        return conn.execute("SELECT COUNT(*) FROM accounts WHERE status='pending'").fetchone()[0]

def _student_row(row):
    if row is None: return None
    d = dict(row)
    d['nfcId']    = d.get('nfc_id', '')
    d['name']     = d.get('full_name', '')
    d['course']   = d.get('program', '')
    d['address']  = d.get('eth_address', '')
    d['tx_hash']  = d.get('reg_tx_hash', '')
    d['section']  = (d.get('section') or '').strip().upper()
    return d

def db_save_student(s):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            "INSERT INTO students "
            "(nfc_id,full_name,student_id,program,year_level,section,"
            " adviser,email,contact,major,semester,school_year,"
            " date_registered,raw_name,eth_address,reg_tx_hash,reg_block,"
            " photo_file,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(nfc_id) DO UPDATE SET "
            "full_name=excluded.full_name, student_id=excluded.student_id, "
            "program=excluded.program, year_level=excluded.year_level, "
            "section=excluded.section, adviser=excluded.adviser, "
            "email=excluded.email, contact=excluded.contact, major=excluded.major, "
            "semester=excluded.semester, school_year=excluded.school_year, "
            "date_registered=excluded.date_registered, raw_name=excluded.raw_name, "
            "eth_address=excluded.eth_address, reg_tx_hash=excluded.reg_tx_hash, "
            "reg_block=excluded.reg_block, photo_file=excluded.photo_file, "
            "updated_at=excluded.updated_at",
            (
                s.get('nfcId', s.get('nfc_id','')),
                s.get('name',  s.get('full_name','')),
                s.get('student_id',''),
                s.get('course', s.get('program','')),
                s.get('year_level',''),
                (s.get('section') or '').strip().upper(),
                s.get('adviser',''), s.get('email',''), s.get('contact',''),
                s.get('major',''), s.get('semester',''), s.get('school_year',''),
                s.get('date_registered',''),
                s.get('raw_name',''),
                s.get('address', s.get('eth_address','')),
                s.get('tx_hash', s.get('reg_tx_hash','')),
                s.get('reg_block', 0),
                s.get('photo_file',''),
                s.get('created_at', now), now
            )
        )

def db_get_all_students():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM students ORDER BY full_name").fetchall()
    return [_student_row(r) for r in rows]

def db_get_student(nfc_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM students WHERE nfc_id=?", (nfc_id,)).fetchone()
    return _student_row(row)

def db_delete_student(nfc_id):
    with get_db() as conn:
        conn.execute("DELETE FROM students WHERE nfc_id=?", (nfc_id,))

def db_save_attendance_log(sess_id, nfc_id, student_name, student_id,
                            status, tap_time, tx_hash='', block_number=0,
                            excuse_note=''):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            "INSERT INTO attendance_logs "
            "(sess_id,nfc_id,student_name,student_id,status,tap_time,"
            " tx_hash,block_number,excuse_note,created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(sess_id,nfc_id) DO UPDATE SET "
            "status=excluded.status, tap_time=excluded.tap_time, "
            "tx_hash=excluded.tx_hash, block_number=excluded.block_number, "
            "excuse_note=excluded.excuse_note",
            (sess_id, nfc_id, student_name, student_id,
             status, tap_time, tx_hash, block_number, excuse_note, now)
        )

def db_get_session_attendance(sess_id):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM attendance_logs WHERE sess_id=? ORDER BY tap_time",
            (sess_id,)
        ).fetchall()
    return [dict(r) for r in rows]

def db_update_session_totals(sess_id):
    with get_db() as conn:
        counts = conn.execute(
            "SELECT status, COUNT(*) as cnt FROM attendance_logs "
            "WHERE sess_id=? GROUP BY status", (sess_id,)
        ).fetchall()
        totals = {r['status']: r['cnt'] for r in counts}
        conn.execute(
            "UPDATE sessions SET "
            "total_present=?, total_late=?, total_absent=?, total_excused=? "
            "WHERE sess_id=?",
            (totals.get('present',0), totals.get('late',0),
             totals.get('absent',0),  totals.get('excused',0), sess_id)
        )

def nfc_is_waiting():
    with get_db() as conn:
        row = conn.execute("SELECT waiting FROM nfc_scanner WHERE id=1").fetchone()
    return bool(row and row['waiting'])

def nfc_set_waiting(flag, requested_by=''):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            "UPDATE nfc_scanner SET waiting=?, scanned_uid='', "
            "requested_by=?, requested_at=? WHERE id=1",
            (1 if flag else 0, requested_by, now)
        )
        conn.execute(
            "UPDATE nfc_registration SET waiting=?, scanned_uid='', "
            "requested_by=?, requested_at=? WHERE id=1",
            (1 if flag else 0, requested_by, now)
        )

def nfc_set_uid(uid):
    with get_db() as conn:
        conn.execute(
            "UPDATE nfc_scanner SET waiting=0, scanned_uid=? WHERE id=1", (uid,))
        conn.execute(
            "UPDATE nfc_registration SET waiting=0, scanned_uid=? WHERE id=1", (uid,))

def nfc_get_uid():
    with get_db() as conn:
        row = conn.execute(
            "SELECT scanned_uid FROM nfc_scanner WHERE id=1").fetchone()
    return row['scanned_uid'] if row else ''

def nfc_clear():
    with get_db() as conn:
        conn.execute(
            "UPDATE nfc_scanner SET waiting=0, scanned_uid='' WHERE id=1")
        conn.execute(
            "UPDATE nfc_registration SET waiting=0, scanned_uid='' WHERE id=1")

def db_get_all_subjects():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    return {r['subject_id']: dict(r) for r in rows}

def db_get_subject(subject_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM subjects WHERE subject_id=?", (subject_id,)).fetchone()
    return dict(row) if row else None

def db_save_subject(subject_id, s):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO subjects (subject_id,name,course_code,units,created_by,created_at)"
            " VALUES (?,?,?,?,?,?)"
            " ON CONFLICT(subject_id) DO UPDATE SET"
            " name=excluded.name, course_code=excluded.course_code,"
            " units=excluded.units, created_by=excluded.created_by,"
            " created_at=excluded.created_at",
            (subject_id, s.get('name',''), s.get('course_code',''),
             str(s.get('units','3')), s.get('created_by',''), s.get('created_at',''))
        )

def db_delete_subject(subject_id):
    with get_db() as conn:
        conn.execute("DELETE FROM subjects WHERE subject_id=?", (subject_id,))

def db_get_all_photos():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM photos").fetchall()
    return {r['person_id']: r['filename'] for r in rows}

def db_get_photo(person_id):
    with get_db() as conn:
        row = conn.execute("SELECT filename FROM photos WHERE person_id=?", (person_id,)).fetchone()
    return row['filename'] if row else None

def db_save_photo(person_id, filename):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO photos (person_id,filename) VALUES (?,?)"
            " ON CONFLICT(person_id) DO UPDATE SET filename=excluded.filename",
            (person_id, filename)
        )

def db_delete_photo(person_id):
    with get_db() as conn:
        conn.execute("DELETE FROM photos WHERE person_id=?", (person_id,))

def db_rename_photo_key(old_key, new_key):
    with get_db() as conn:
        conn.execute("UPDATE photos SET person_id=? WHERE person_id=?", (new_key, old_key))

def db_get_override(nfc_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM student_overrides WHERE nfc_id=?", (nfc_id,)).fetchone()
    return dict(row) if row else {}

def db_save_override(nfc_id, fields):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO student_overrides (nfc_id,full_name,student_id,email,contact,"
            "adviser,major,semester,school_year,date_registered,course,year_level,section)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)"
            " ON CONFLICT(nfc_id) DO UPDATE SET"
            " full_name=excluded.full_name, student_id=excluded.student_id,"
            " email=excluded.email, contact=excluded.contact,"
            " adviser=excluded.adviser, major=excluded.major,"
            " semester=excluded.semester, school_year=excluded.school_year,"
            " date_registered=excluded.date_registered,"
            " course=excluded.course, year_level=excluded.year_level,"
            " section=excluded.section",
            (nfc_id, fields.get('full_name',''), fields.get('student_id',''),
             fields.get('email',''), fields.get('contact',''),
             fields.get('adviser',''), fields.get('major',''),
             fields.get('semester',''), fields.get('school_year',''),
             fields.get('date_registered',''), fields.get('course',''),
             fields.get('year_level',''), fields.get('section','').upper())
        )

# ── Schedule DB helpers ────────────────────────────────────────────────────

def db_get_all_schedules():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM schedules WHERE is_active=1 ORDER BY day_of_week, start_time"
        ).fetchall()
    return [dict(r) for r in rows]

def db_get_schedules_for_teacher(username):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM schedules WHERE teacher_username=? AND is_active=1 ORDER BY day_of_week, start_time",
            (username,)
        ).fetchall()
    return [dict(r) for r in rows]

def db_get_schedule(schedule_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM schedules WHERE schedule_id=?", (schedule_id,)).fetchone()
    return dict(row) if row else None

def db_save_schedule(s: dict) -> str:
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sid = s.get('schedule_id') or str(uuid.uuid4())
    with get_db() as conn:
        conn.execute(
            "INSERT INTO schedules "
            "(schedule_id,section_key,subject_id,subject_name,course_code,"
            " teacher_username,teacher_name,day_of_week,start_time,end_time,"
            " grace_minutes,is_active,created_by,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,1,?,?,?) "
            "ON CONFLICT(schedule_id) DO UPDATE SET "
            "section_key=excluded.section_key, subject_id=excluded.subject_id, "
            "subject_name=excluded.subject_name, course_code=excluded.course_code, "
            "teacher_username=excluded.teacher_username, teacher_name=excluded.teacher_name, "
            "day_of_week=excluded.day_of_week, start_time=excluded.start_time, "
            "end_time=excluded.end_time, grace_minutes=excluded.grace_minutes, "
            "updated_at=excluded.updated_at",
            (sid,
             normalize_section_key(s.get('section_key', '')),
             s.get('subject_id', ''), s.get('subject_name', ''), s.get('course_code', ''),
             s.get('teacher_username', ''), s.get('teacher_name', ''),
             int(s.get('day_of_week', 0)),
             s.get('start_time', ''), s.get('end_time', ''),
             int(s.get('grace_minutes', 15)),
             s.get('created_by', ''), now, now)
        )
    return sid

def db_delete_schedule(schedule_id):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            "UPDATE schedules SET is_active=0, updated_at=? WHERE schedule_id=?",
            (now, schedule_id)
        )

def get_todays_schedules(username=None):
    """Return schedules that fall on today's weekday (0=Mon). If username provided, filter by it."""
    today_dow = datetime.now().weekday()
    if username:
        return [s for s in db_get_schedules_for_teacher(username)
                if int(s['day_of_week']) == today_dow]
    return [s for s in db_get_all_schedules()
            if int(s['day_of_week']) == today_dow]

def db_get_teacher_sessions(username):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM sessions WHERE teacher_username=? ORDER BY started_at DESC",
            (username,)
        ).fetchall()
    return [_session_row_with_logs(conn, r) for r in rows]

def _normalize_hhmm(value):
    """Normalize schedule time values to HH:MM (supports HH:MM, HH:MM:SS, and AM/PM formats)."""
    raw = str(value or '').strip()
    if not raw:
        return None
    # 24h HH:MM
    try:
        dt = datetime.strptime(raw, '%H:%M')
        return dt.strftime('%H:%M')
    except Exception:
        pass
    # 24h HH:MM:SS
    try:
        dt = datetime.strptime(raw, '%H:%M:%S')
        return dt.strftime('%H:%M')
    except Exception:
        pass
    # 12h formats, with and without space before AM/PM
    for fmt in ('%I:%M %p', '%I:%M%p'):
        try:
            dt = datetime.strptime(raw.upper(), fmt)
            return dt.strftime('%H:%M')
        except Exception:
            continue
    return None

def _time_mins(value):
    hhmm = _normalize_hhmm(value)
    if not hhmm:
        return None
    h, m = hhmm.split(':')
    return int(h) * 60 + int(m)

def check_and_start_scheduled_sessions():
    """
    Background-friendly function:
    1. Check for schedules that should have started but don't have an active session.
    2. Check for active sessions that should have ended.
    """
    now_dt = datetime.now()
    today_dow = now_dt.weekday()
    current_time_str = now_dt.strftime('%H:%M')
    
    with app.app_context():
        # Get all active schedules for today
        schedules = [s for s in db_get_all_schedules() if int(s['day_of_week']) == today_dow]
        active_sessions = get_active_sessions()
        
        for s in schedules:
            start_time = s['start_time']
            end_time   = s['end_time']
            start_hhmm = _normalize_hhmm(start_time)
            end_hhmm = _normalize_hhmm(end_time)
            if not start_hhmm or not end_hhmm:
                print(f"[AUTO WARN] Invalid schedule time format for schedule_id={s.get('schedule_id')} start={start_time} end={end_time}")
                continue
            
            # Start at schedule trigger time if this specific schedule has not run yet today.
            # Using schedule_id avoids false blocks from other sessions with same subject/section.
            today_prefix = now_dt.strftime('%Y-%m-%d') + '%'
            with get_db() as conn:
                schedule_id = s.get('schedule_id')
                if schedule_id:
                    already_ran = conn.execute(
                        "SELECT 1 FROM sessions WHERE schedule_id=? AND started_at LIKE ?",
                        (str(schedule_id), today_prefix)
                    ).fetchone()
                else:
                    # Legacy fallback for rows without schedule_id.
                    already_ran = conn.execute(
                        "SELECT 1 FROM sessions "
                        "WHERE teacher_username=? AND subject_id=? AND section_key=? AND started_at LIKE ?",
                        (
                            s.get('teacher_username', ''),
                            s['subject_id'],
                            normalize_section_key(s['section_key']),
                            today_prefix,
                        )
                    ).fetchone()
            
            start_dt = datetime.strptime(
                f"{now_dt.strftime('%Y-%m-%d')} {start_hhmm}:00",
                '%Y-%m-%d %H:%M:%S'
            )
            end_dt = datetime.strptime(
                f"{now_dt.strftime('%Y-%m-%d')} {end_hhmm}:00",
                '%Y-%m-%d %H:%M:%S'
            )
            if end_dt <= start_dt:
                print(f"[AUTO WARN] Invalid schedule window schedule_id={s.get('schedule_id')} start={start_hhmm} end={end_hhmm}")
                continue

            if not already_ran and start_dt <= now_dt < end_dt:
                # Automate session start
                sess_id = str(uuid.uuid4())[:13]
                subj = db_get_subject(s['subject_id'])
                late_cutoff_dt = datetime.strptime(start_hhmm, '%H:%M') + timedelta(minutes=s.get('grace_minutes', 15))
                late_cutoff = late_cutoff_dt.strftime('%H:%M')
                
                new_sess = {
                    'sess_id': sess_id,
                    'subject_id': s['subject_id'],
                    'subject_name': s['subject_name'],
                    'course_code': s['course_code'],
                    'units': subj.get('units', 3) if subj else 3,
                    'time_slot': f"{start_hhmm} - {end_hhmm}",
                    'section_key': s['section_key'],
                    'teacher_username': s['teacher_username'],
                    'teacher_name': s['teacher_name'],
                    'started_at': now_dt.strftime('%Y-%m-%d %H:%M:%S'),
                    'late_cutoff': f"{now_dt.strftime('%Y-%m-%d')} {late_cutoff}:00",
                    'auto_end_at': f"{now_dt.strftime('%Y-%m-%d')} {end_hhmm}:00",
                    'grace_period': s.get('grace_minutes', 15),
                    'schedule_id': s['schedule_id']
                }
                save_session(sess_id, new_sess)
                print(f"[AUTO] Started session {sess_id} for {s['subject_name']} ({s['teacher_username']})")
            elif already_ran and start_dt <= now_dt < end_dt:
                print(f"[AUTO] Skipped schedule_id={s.get('schedule_id')} (already ran today)")

        # 2. End sessions that passed their auto_end_at
        for sid, asess in active_sessions.items():
            auto_end = asess.get('auto_end_at')
            if auto_end:
                try:
                    end_dt = datetime.strptime(auto_end, '%Y-%m-%d %H:%M:%S')
                    if now_dt >= end_dt:
                        result = _finalize_session(
                            sid,
                            ended_time=now_dt.strftime('%Y-%m-%d %H:%M:%S'),
                            async_chain_and_email=True,
                        )
                        if result and not result.get('already_ended'):
                            print(f"[AUTO] Ended session {sid} automatically.")
                except:
                    pass

def check_and_end_expired_sessions():
    """System-wide safety check to end sessions after their own configured end time."""
    now_dt = datetime.now()
    now_str = now_dt.strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as conn:
        active = conn.execute(
            "SELECT sess_id, auto_end_at, schedule_id, subject_id, section_key "
            "FROM sessions WHERE ended_at IS NULL"
        ).fetchall()

    for s in active:
        sess_id = s['sess_id']

        # Primary source of truth: per-session auto_end_at stamped at session creation.
        auto_end_at = (s['auto_end_at'] or '').strip()
        if auto_end_at:
            try:
                end_dt = datetime.strptime(auto_end_at, '%Y-%m-%d %H:%M:%S')
                if now_dt >= end_dt:
                    result = _finalize_session(sess_id, ended_time=now_str, async_chain_and_email=True)
                    if result and not result.get('already_ended'):
                        print(f"[AUTO] Ended session {sess_id} (Auto End Reached)")
                continue
            except Exception:
                # If stored timestamp is malformed, fall back to schedule lookup below.
                pass

        # Fallback for legacy sessions missing auto_end_at: prefer exact schedule_id lookup.
        end_mins = None
        with get_db() as conn:
            schedule_id = (s['schedule_id'] or '').strip()
            sched = None
            if schedule_id:
                sched = conn.execute(
                    "SELECT end_time FROM schedules WHERE schedule_id=? AND is_active=1",
                    (schedule_id,)
                ).fetchone()
            if not sched:
                # Legacy fallback only.
                sched = conn.execute(
                    "SELECT end_time FROM schedules WHERE subject_id=? AND section_key=? AND is_active=1 "
                    "ORDER BY updated_at DESC LIMIT 1",
                    (s['subject_id'], s['section_key'])
                ).fetchone()

        if sched:
            end_mins = _time_mins(sched['end_time'])

        current_time_mins = _time_mins(now_dt.strftime('%H:%M'))
        if end_mins is None or current_time_mins is None:
            continue
        if current_time_mins >= end_mins:
            result = _finalize_session(sess_id, ended_time=now_str, async_chain_and_email=True)
            if result and not result.get('already_ended'):
                print(f"[AUTO] Ended session {sess_id} (Schedule End Time Reached)")

def automation_loop():
    """Single master loop for DAVS automation/synchronization.
    
    Runs every few seconds so scheduled sessions start almost immediately
    after the trigger time while still being lightweight.
    """
    poll_seconds = 5
    while True:
        try:
            check_and_start_scheduled_sessions()
            check_and_end_expired_sessions()
        except Exception as e:
            print(f"[AUTO ERROR] {e}")
        time.sleep(poll_seconds)

def ensure_automation_thread_running():
    """Start automation loop once per process even under flask/wsgi launch modes."""
    global AUTO_THREAD
    with AUTO_THREAD_LOCK:
        if AUTO_THREAD and AUTO_THREAD.is_alive():
            return
        AUTO_THREAD = Thread(target=automation_loop, daemon=True, name='davs-automation-loop')
        AUTO_THREAD.start()
        print('[AUTO] Automation loop started.')

@app.before_request
def _ensure_automation_thread_running():
    ensure_automation_thread_running()

from datetime import timedelta

DOW_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

# ── Excuse Request DB helpers ──────────────────────────────────────────────

def _excuse_pk_column(conn) -> str:
    cols = [r['name'] for r in conn.execute("PRAGMA table_info(excuse_requests)").fetchall()]
    for cand in ('id', 'excuse_id', 'request_id'):
        if cand in cols:
            return cand
    return 'rowid'

def _excuse_order_expr(conn) -> str:
    cols = [r['name'] for r in conn.execute("PRAGMA table_info(excuse_requests)").fetchall()]
    if 'created_at' in cols and 'submitted_at' in cols:
        return "COALESCE(er.created_at, er.submitted_at, '')"
    if 'created_at' in cols:
        return "er.created_at"
    if 'submitted_at' in cols:
        return "er.submitted_at"
    return "''"

def db_save_excuse_request(data: dict) -> int:
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        cols = [r['name'] for r in conn.execute("PRAGMA table_info(excuse_requests)").fetchall()]
        insert_cols = [
            'sess_id', 'nfc_id', 'student_name', 'student_id', 'student_email',
            'reason_type', 'reason_detail', 'attachment_file', 'status'
        ]
        values = [
            data.get('sess_id', ''), data.get('nfc_id', ''),
            data.get('student_name', ''), data.get('student_id', ''),
            data.get('student_email', ''), data.get('reason_type', ''),
            data.get('reason_detail', ''), data.get('attachment_file', ''),
            'pending'
        ]

        # Support both legacy and current schemas.
        if 'created_at' in cols:
            insert_cols.append('created_at')
            values.append(now)
        if 'submitted_at' in cols:
            insert_cols.append('submitted_at')
            values.append(now)

        placeholders = ','.join(['?'] * len(insert_cols))
        sql = f"INSERT INTO excuse_requests ({','.join(insert_cols)}) VALUES ({placeholders})"
        cur = conn.execute(sql, tuple(values))
        pk_col = _excuse_pk_column(conn)
        inserted_id = cur.lastrowid
        if pk_col != 'rowid':
            try:
                row = conn.execute(
                    f"SELECT {pk_col} AS pk FROM excuse_requests WHERE rowid=?",
                    (cur.lastrowid,)
                ).fetchone()
                pk_val = row['pk'] if row else None
                if pk_val is not None and str(pk_val).strip() != '':
                    inserted_id = pk_val
            except Exception:
                pass
        return inserted_id

def db_get_all_excuse_requests(status_filter=None):
    try:
        with get_db() as conn:
            pk_col = _excuse_pk_column(conn)
            order_expr = _excuse_order_expr(conn)
            pk_select = f"er.{pk_col}" if pk_col != 'rowid' else "er.rowid"
            # Explicitly alias er.id to avoid any column shadowing from JOIN
            base_sql = (
                f"SELECT {pk_select} AS id, er.rowid AS _rowid, er.sess_id, er.nfc_id, "
                "er.student_name, er.student_id, er.student_email, "
                "er.reason_type, er.reason_detail, er.attachment_file, "
                "er.status, er.reviewed_by, er.reviewed_at, "
                "COALESCE(er.created_at, er.submitted_at, '') AS created_at, "
                "s.subject_name AS session_subject, "
                "s.section_key AS session_section "
                "FROM excuse_requests er "
                "LEFT JOIN sessions s ON er.sess_id = s.sess_id "
            )
            if status_filter:
                rows = conn.execute(
                    base_sql + f"WHERE er.status=? ORDER BY {order_expr} DESC",
                    (status_filter,)
                ).fetchall()
            else:
                rows = conn.execute(
                    base_sql + f"ORDER BY {order_expr} DESC"
                ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('id') is None or str(d.get('id')).strip() == '':
                d['id'] = d.get('_rowid')
            d.pop('_rowid', None)
            out.append(d)
        return out
    except Exception as e:
        print(f'[DB] db_get_all_excuse_requests error: {e}')
        return []

def db_get_excuse_request(excuse_id):
    with get_db() as conn:
        pk_col = _excuse_pk_column(conn)
        pk_where = pk_col if pk_col != 'rowid' else 'rowid'
        row = conn.execute(
            f"SELECT *, {pk_where} AS id, rowid AS _rowid FROM excuse_requests WHERE {pk_where}=?",
            (excuse_id,)
        ).fetchone()
        if not row and pk_where != 'rowid':
            row = conn.execute(
                "SELECT *, rowid AS id, rowid AS _rowid FROM excuse_requests WHERE rowid=?",
                (excuse_id,)
            ).fetchone()
    if not row:
        return None
    d = dict(row)
    if d.get('id') is None or str(d.get('id')).strip() == '':
        d['id'] = d.get('_rowid')
    d.pop('_rowid', None)
    return d

def db_resolve_excuse(excuse_id: int, resolution: str, reviewed_by: str) -> dict | None:
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    session_sync = None
    with get_db() as conn:
        pk_col = _excuse_pk_column(conn)
        pk_where = pk_col if pk_col != 'rowid' else 'rowid'
        where_col = pk_where
        where_val = excuse_id
        row = conn.execute(
            f"SELECT *, {pk_where} AS id, rowid AS _rowid FROM excuse_requests WHERE {pk_where}=?",
            (excuse_id,)
        ).fetchone()
        if not row and pk_where != 'rowid':
            row = conn.execute(
                "SELECT *, rowid AS id, rowid AS _rowid FROM excuse_requests WHERE rowid=?",
                (excuse_id,)
            ).fetchone()
            if row:
                where_col = 'rowid'
                where_val = excuse_id
        if not row:
            return None
        row_dict = dict(row)
        if row_dict.get('id') is None or str(row_dict.get('id')).strip() == '':
            row_dict['id'] = row_dict.get('_rowid')
        conn.execute(
            f"UPDATE excuse_requests SET status=?, reviewed_by=?, reviewed_at=? WHERE {where_col}=?",
            (resolution, reviewed_by, now, where_val)
        )
        if resolution == 'approved':
            r_type = row_dict.get('reason_type', 'others')
            r_label = dict(EXCUSE_REASONS).get(r_type, r_type.title())
            r_detail = row_dict.get('reason_detail', '')
            note = f"{r_label}{' — ' + r_detail if r_detail else ''}"
            resolved_excuse_id = row_dict.get('id') or excuse_id
            
            # ✅ CRITICAL FIX: Use UPSERT to create record if student hasn't tapped yet
            # If student has no attendance_logs record, INSERT will create one
            # If record exists, UPDATE will mark it as excused
            conn.execute(
                "INSERT INTO attendance_logs "
                "(sess_id, nfc_id, student_name, student_id, status, tap_time, "
                " tx_hash, block_number, excuse_note, excuse_request_id, created_at) "
                "VALUES (?, ?, ?, ?, 'excused', ?, '', 0, ?, ?, ?) "
                "ON CONFLICT(sess_id, nfc_id) DO UPDATE SET "
                "status='excused', excuse_note=excluded.excuse_note, "
                "excuse_request_id=excluded.excuse_request_id",
                (row_dict['sess_id'], row_dict['nfc_id'], row_dict['student_name'], row_dict['student_id'],
                 now, note, resolved_excuse_id, now)
            )
            # Defer session save_session sync until after this DB transaction closes.
            session_sync = {
                'sess_id': row_dict['sess_id'],
                'nfc_id': row_dict['nfc_id'],
                'note': note,
            }
        row_dict.pop('_rowid', None)
    # Sync with live session if active (outside transaction to avoid sqlite write lock).
    if session_sync:
        sess = load_session(session_sync['sess_id'])
        if sess:
            excused = sess.setdefault('excused', [])
            if session_sync['nfc_id'] not in excused:
                excused.append(session_sync['nfc_id'])
            if session_sync['nfc_id'] in sess.get('absent', []):
                sess['absent'].remove(session_sync['nfc_id'])
            sess.setdefault('excuse_notes', {})[session_sync['nfc_id']] = session_sync['note']
            save_session(session_sync['sess_id'], sess)
            sessions_db[session_sync['sess_id']] = sess
    return row_dict

def _resolve_excuse_attachment_path(stored_value: str) -> str | None:
    """Resolve attachment paths across legacy/new storage styles."""
    raw = str(stored_value or '').strip()
    if not raw:
        return None
    raw_norm = raw.replace('\\', '/').lstrip('/')
    base = os.path.basename(raw_norm)
    candidates = [
        os.path.join(UPLOAD_FOLDER_EXCUSES, raw_norm),
        os.path.join(UPLOAD_FOLDER_EXCUSES, base),
        os.path.join(BASE_DIR, 'uploads', 'excuses', base),
        os.path.join(BASE_DIR, 'static', 'uploads', 'excuses', base),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

def _save_excuse_attachment(file_storage) -> str:
    """Save uploaded excuse file; returns stored filename. Raises ValueError on bad input."""
    orig = file_storage.filename or ''
    ext  = orig.rsplit('.', 1)[-1].lower() if '.' in orig else ''
    if ext not in ALLOWED_EXCUSE_EXTS:
        raise ValueError(f"Invalid file type '.{ext}'. Allowed: pdf, jpg, jpeg, png")
    data = file_storage.read()
    if len(data) > MAX_EXCUSE_FILE_MB * 1024 * 1024:
        raise ValueError(f"File too large (max {MAX_EXCUSE_FILE_MB} MB)")
    fname = f"{uuid.uuid4().hex}.{ext}"
    path  = os.path.join(UPLOAD_FOLDER_EXCUSES, fname)
    with open(path, 'wb') as fout:
        fout.write(data)
    return fname

def load_student_names():
    for s in db_get_all_students():
        student_name_map[s['nfcId']] = s['name']
    if student_name_map:
        print(f"[INFO] Loaded {len(student_name_map)} student names from SQLite cache.")
    if BLOCKCHAIN_ONLINE and contract:
        try:
            ef = contract.events.StudentRegistered.create_filter(
                from_block=0, to_block=web3.eth.block_number)
            for e in ef.get_all_entries():
                nid  = e['args']['nfcId']
                name = e['args']['name'].split(' | ')[0]
                student_name_map[nid] = name
            print(f"[INFO] Student names enriched from blockchain ({len(student_name_map)} total).")
        except Exception as ex:
            print(f"[WARNING] load_student_names blockchain error: {ex}")

init_db()
migrate_json_to_sqlite()
sessions_db = load_sessions()

student_name_map = {}
load_student_names()
recent_attendance = deque(maxlen=50)

DEPARTMENTS = {'DIT': {'label':'Department of Information Technology','courses':['BS Computer Science','BS Information Technology']}}
YEAR_LEVELS  = ['1st Year','2nd Year','3rd Year','4th Year']
SECTIONS     = ['A','B','C','D']

def fmt_time(dt_str):
    if not dt_str: return '—'
    try:
        dt = datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%b %d, %Y · %I:%M %p')
    except:
        return dt_str

def fmt_time_short(dt_str):
    if not dt_str: return '—'
    try:
        dt = datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%I:%M %p').lstrip('0')
    except:
        return dt_str

# ── Role constants ────────────────────────────────────────────────────────
SUPER_ADMIN_ROLE = 'super_admin'
ADMIN_ROLES  = {'admin', 'super_admin'}
STAFF_ROLES  = {'teacher', 'admin', 'super_admin'}

EXCUSE_REASONS = [
    ('sickness',    'Sickness / Illness'),
    ('lbm',         'LBM (Loose Bowel Movement)'),
    ('emergency',   'Family/Personal Emergency'),
    ('bereavement', 'Bereavement (Death in Family)'),
    ('medical',     'Medical Appointment / Check-up'),
    ('accident',    'Accident or Injury'),
    ('official',    'Official School Business / Event'),
    ('weather',     'Extreme Weather / Calamity'),
    ('transport',   'Transportation Problem'),
    ('others',      'Others (please specify)'),
]

ALLOWED_EXCUSE_EXTS = {'pdf', 'jpg', 'jpeg', 'png'}
MAX_EXCUSE_FILE_MB  = 5
UPLOAD_FOLDER_EXCUSES = os.path.join(BASE_DIR, 'static', 'uploads', 'excuses')
os.makedirs(UPLOAD_FOLDER_EXCUSES, exist_ok=True)

def login_required(f):
    @wraps(f)
    def dec(*a,**kw):
        if 'username' not in session:
            flash('Please log in first.')
            return redirect(url_for('login'))
        return f(*a,**kw)
    return dec

def admin_required(f):
    """Allows admin and super_admin (backwards-compatible)."""
    @wraps(f)
    def dec(*a,**kw):
        if 'username' not in session: return redirect(url_for('login'))
        if session.get('role') not in ADMIN_ROLES:
            flash('Admin access required.')
            return redirect(url_for('teacher_dashboard'))
        return f(*a,**kw)
    return dec

def super_admin_required(f):
    """Restricts access to super_admin role only."""
    @wraps(f)
    def dec(*a,**kw):
        if 'username' not in session: return redirect(url_for('login'))
        if session.get('role') != 'super_admin':
            flash('Super Admin access required.', 'danger')
            return redirect(url_for('teacher_dashboard'))
        return f(*a,**kw)
    return dec

def staff_required(f):
    """Any authenticated staff member (teacher/admin/super_admin)."""
    @wraps(f)
    def dec(*a,**kw):
        if 'username' not in session: return redirect(url_for('login'))
        if session.get('role') not in STAFF_ROLES:
            flash('Staff access required.')
            return redirect(url_for('login'))
        return f(*a,**kw)
    return dec

def get_current_user(): return db_get_user(session.get('username',''))

def parse_student(raw):
    parts = raw.split(' | ')
    r = {'name':parts[0],'student_id':'','course':'','year_level':'','section':'',
         'adviser':'','email':'','contact':'','semester':'','school_year':'',
         'date_registered':'','major':''}
    for p in parts[1:]:
        if   p.startswith('ID:'):       r['student_id']      = p[3:]
        elif p.startswith('Course:'):   r['course']          = p[7:]
        elif p.startswith('Year:'):     r['year_level']      = p[5:]
        elif p.startswith('Sec:'):      r['section']         = p[4:].upper()
        elif p.startswith('Adviser:'):  r['adviser']         = p[8:]
        elif p.startswith('Email:'):    r['email']           = p[6:]
        elif p.startswith('Tel:'):      r['contact']         = p[4:]
        elif p.startswith('Sem:'):      r['semester']        = p[4:]
        elif p.startswith('SY:'):       r['school_year']     = p[3:]
        elif p.startswith('RegDate:'): r['date_registered'] = p[8:]
        elif p.startswith('Major:'):    r['major']           = p[6:]
    return r

def get_all_students():
    global BLOCKCHAIN_ONLINE
    cached = db_get_all_students()

    if contract is None:
        # Contract not loaded (or reset state); do not attempt event sync.
        if cached:
            return cached
        print("[INFO] No blockchain contract; returning SQLite student cache (empty).")
        return []

    if not BLOCKCHAIN_ONLINE and contract:
        try:
            BLOCKCHAIN_ONLINE = web3.is_connected()
        except Exception:
            BLOCKCHAIN_ONLINE = False
    if BLOCKCHAIN_ONLINE and contract:
        try:
            ef = contract.events.StudentRegistered.create_filter(
                from_block=0, to_block=web3.eth.block_number)
            entries = ef.get_all_entries()
            students, seen = [], set()
            for e in entries:
                a = e['args']
                if a['nfcId'] in seen: continue
                seen.add(a['nfcId'])
                p = parse_student(a['name'])
                s = {**p, 'raw_name': a['name'], 'nfcId': a['nfcId'],
                     'address': a['studentAddr'],
                     'tx_hash': e['transactionHash'].hex()}
                ov = db_get_override(a['nfcId'])
                if ov.get('full_name'):       s['name']            = ov['full_name']
                if ov.get('student_id'):      s['student_id']      = ov['student_id']
                if ov.get('email'):           s['email']           = ov['email']
                if ov.get('contact'):         s['contact']         = ov['contact']
                if ov.get('adviser'):         s['adviser']         = ov['adviser']
                if ov.get('major'):           s['major']           = ov['major']
                if ov.get('semester'):        s['semester']        = ov['semester']
                if ov.get('school_year'):     s['school_year']     = ov['school_year']
                if ov.get('date_registered'): s['date_registered'] = ov['date_registered']
                if ov.get('course'):          s['course']          = ov['course']
                if ov.get('year_level'):      s['year_level']      = ov['year_level']
                if ov.get('section'):         s['section']         = ov['section'].upper()
                s['section'] = s['section'].strip().upper()
                students.append(s)
                db_save_student(s)
            # If blockchain is reachable but has no StudentRegistered events
            # (e.g., fresh contract deployment), keep using SQLite cache so
            # admin/teacher lists do not appear empty.
            if students:
                return students
            print("[INFO] Blockchain has no student events yet — using SQLite cache.")
            if cached:
                for s in cached:
                    ov = db_get_override(s['nfcId'])
                    if ov.get('full_name'):       s['name']            = ov['full_name']
                    if ov.get('student_id'):      s['student_id']      = ov['student_id']
                    if ov.get('email'):           s['email']           = ov['email']
                    if ov.get('contact'):         s['contact']         = ov['contact']
                    if ov.get('adviser'):         s['adviser']         = ov['adviser']
                    if ov.get('major'):           s['major']           = ov['major']
                    if ov.get('semester'):        s['semester']        = ov['semester']
                    if ov.get('school_year'):     s['school_year']     = ov['school_year']
                    if ov.get('date_registered'): s['date_registered'] = ov['date_registered']
                    if ov.get('course'):          s['course']          = ov['course']
                    if ov.get('year_level'):      s['year_level']      = ov['year_level']
                    if ov.get('section'):         s['section']         = ov['section'].upper()
                    s['section'] = (s.get('section') or '').strip().upper()
                return cached
            return []
        except Exception as _be:
            print(f"[WARNING] Blockchain unreachable: {_be} — falling back to SQLite cache.")
            BLOCKCHAIN_ONLINE = False
    if not cached:
        print("[WARNING] No students in SQLite cache and blockchain is offline.")
        return []
    for s in cached:
        ov = db_get_override(s['nfcId'])
        if ov.get('full_name'):       s['name']            = ov['full_name']
        if ov.get('student_id'):      s['student_id']      = ov['student_id']
        if ov.get('email'):           s['email']           = ov['email']
        if ov.get('contact'):         s['contact']         = ov['contact']
        if ov.get('adviser'):         s['adviser']         = ov['adviser']
        if ov.get('major'):           s['major']           = ov['major']
        if ov.get('semester'):        s['semester']        = ov['semester']
        if ov.get('school_year'):     s['school_year']     = ov['school_year']
        if ov.get('date_registered'): s['date_registered'] = ov['date_registered']
        if ov.get('course'):          s['course']          = ov['course']
        if ov.get('year_level'):      s['year_level']      = ov['year_level']
        if ov.get('section'):         s['section']         = ov['section'].upper()
        s['section'] = (s.get('section') or '').strip().upper()
    return cached

def get_attendance_records(nfc_id):
    try:
        ts, statuses = contract.functions.getAttendance(nfc_id).call()
        out = []
        for t, code in zip(ts, statuses):
            # Legacy contract returns bool; upgraded contract returns uint8.
            if isinstance(code, bool):
                status = 'present' if code else 'absent'
            else:
                status = {0: 'present', 1: 'late', 2: 'absent', 3: 'excused'}.get(int(code), 'absent')
            out.append((datetime.fromtimestamp(t).strftime('%Y-%m-%d %H:%M:%S'), status in ('present', 'late')))
        return out
    except:
        return []

def chain_status_code(status: str) -> int:
    return {
        'present': 0,
        'late': 1,
        'absent': 2,
        'excused': 3,
    }.get((status or '').lower(), 0)

def mark_attendance_on_chain(nfc_id: str, status: str):
    tx = contract.functions.markAttendanceWithStatus(
        nfc_id, chain_status_code(status)
    ).transact({'from': admin_account})
    receipt = web3.eth.wait_for_transaction_receipt(tx)
    return receipt['transactionHash'].hex(), receipt['blockNumber']

def get_student_by_nfc_cached(nfc_id: str):
    st = db_get_student(nfc_id)
    if st:
        return st
    return next((s for s in get_all_students() if s.get('nfcId') == nfc_id), None)

def get_student_attendance_stats(nfc_id):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT al.status FROM attendance_logs al "
            "JOIN sessions s ON al.sess_id = s.sess_id "
            "WHERE al.nfc_id=? AND s.ended_at IS NOT NULL",
            (nfc_id,)
        ).fetchall()
    total = late = excused = present = absent = 0
    for row in rows:
        status = row['status']
        total += 1
        if   status == 'excused': excused += 1
        elif status == 'late':    late    += 1
        elif status == 'present': present += 1
        else:                     absent  += 1
    attended = present + late
    rate = round(attended / total * 100, 1) if total else 0
    return {'total': total, 'present': present + late, 'late': late,
            'excused': excused, 'absent': absent, 'rate': rate}

# Course name alias map was moved to module load phase for initialize access


def student_matches_section(student, allowed_keys: set) -> bool:
    """
    Returns True if the student belongs to any section in allowed_keys.
    Handles course-name alias mismatches (e.g. BSIT vs BS Information Technology).
    """
    key = build_student_section_key(student)
    if key is None:
        return False
    if key in allowed_keys:
        return True
    # Try with canonical course name
    course = normalize_course_name(student.get('course') or '')
    year   = (student.get('year_level') or '').strip()
    sec    = (student.get('section') or '').strip().upper()
    if course and year and sec:
        normalized_key = normalize_section_key(f"{course}|{year}|{sec}")
        if normalized_key in allowed_keys:
            return True
    return False


def teacher_students(user):
    if not user: return []
    allowed = set()
    # Support legacy 'sections' assigned via direct user modal
    for s in user.get('sections', []):
        allowed.add(normalize_section_key(s))
    # Support 'my_subjects' assigned via Teacher Dashboard Add Subject
    for ms in user.get('my_subjects', []):
        allowed.add(normalize_section_key(ms.get('section_key', '')))
    # Support dynamic schedules created via Admin Schedules
    username = user.get('username', '')
    if username:
        schedules = db_get_schedules_for_teacher(username)
        for sched in schedules:
            allowed.add(normalize_section_key(sched.get('section_key', '')))
    
    allowed = {k for k in allowed if k}
    if not allowed:
        return []
    
    all_cached = db_get_all_students()
    if all_cached:
        return [s for s in all_cached if student_matches_section(s, allowed)]
    return [s for s in get_all_students() if student_matches_section(s, allowed)]

def get_active_sessions():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM sessions WHERE ended_at IS NULL").fetchall()
        result = {}
        for r in rows:
            s = _session_row_with_logs(conn, r)
            result[r['sess_id']] = s
    return result

def _finalize_session(sess_id, ended_time=None, async_chain_and_email=True):
    """Finalize a live session and keep DB/UI/blockchain/email in sync."""
    sess = load_session(sess_id)
    if not sess:
        return None
    if sess.get('ended_at'):
        return {'already_ended': True, 'ended_at': sess.get('ended_at', '')}

    all_students = get_all_students()
    section_key = normalize_section_key(sess.get('section_key', ''))
    section_students = [s for s in all_students if build_student_section_key(s) == section_key]

    present_set = set(sess.get('present', []))
    late_set = set(sess.get('late', []))
    excused_set = set(sess.get('excused', []))
    with get_db() as conn:
        excused_from_db = conn.execute(
            "SELECT DISTINCT nfc_id FROM attendance_logs WHERE sess_id=? AND status='excused'",
            (sess_id,)
        ).fetchall()
        approved_excuse_requests = conn.execute(
            "SELECT DISTINCT nfc_id FROM excuse_requests WHERE sess_id=? AND status='approved'",
            (sess_id,)
        ).fetchall()
    excused_set.update([row['nfc_id'] for row in excused_from_db])
    excused_set.update([row['nfc_id'] for row in approved_excuse_requests])

    ended_at = ended_time or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    absent_ids = []
    for st in section_students:
        nid = st['nfcId']
        if nid in present_set or nid in excused_set:
            continue
        db_save_attendance_log(
            sess_id=sess_id,
            nfc_id=nid,
            student_name=st.get('name', ''),
            student_id=st.get('student_id', ''),
            status='absent',
            tap_time=ended_at,
            tx_hash='',
            block_number=0,
        )
        absent_ids.append(nid)

    with get_db() as conn:
        conn.execute(
            "UPDATE sessions SET total_enrolled=?, ended_at=? WHERE sess_id=?",
            (len(section_students), ended_at, sess_id)
        )

    sess['ended_at'] = ended_at
    sess['absent'] = absent_ids
    save_session(sess_id, sess)
    sessions_db[sess_id] = sess

    def _post_finalize_worker():
        with app.app_context():
            if BLOCKCHAIN_ONLINE and contract and admin_account:
                for st in section_students:
                    nid = st['nfcId']
                    if nid in present_set or nid in excused_set:
                        continue
                    with get_db() as conn:
                        current = conn.execute(
                            "SELECT status FROM attendance_logs WHERE sess_id=? AND nfc_id=?",
                            (sess_id, nid)
                        ).fetchone()
                    if current and current['status'] == 'excused':
                        continue
                    try:
                        abs_tx, abs_block = mark_attendance_on_chain(nid, 'absent')
                        with get_db() as conn:
                            conn.execute(
                                "UPDATE attendance_logs SET tx_hash=?, block_number=? WHERE sess_id=? AND nfc_id=?",
                                (abs_tx, abs_block, sess_id, nid)
                            )
                    except Exception as e:
                        print(f"[WARN] Failed absent chain mark for {nid}: {e}")

            for st in section_students:
                nid = st['nfcId']
                if nid in present_set or nid in excused_set:
                    continue
                try:
                    with get_db() as conn:
                        lg = conn.execute(
                            "SELECT tx_hash, block_number FROM attendance_logs WHERE sess_id=? AND nfc_id=?",
                            (sess_id, nid)
                        ).fetchone()
                    send_student_attendance_receipt(
                        student_name=st.get('name', nid),
                        student_email=st.get('email', ''),
                        student_id=st.get('student_id', ''),
                        subject_name=sess.get('subject_name', ''),
                        section_key=sess.get('section_key', ''),
                        teacher_name=sess.get('teacher_name', ''),
                        tap_time=ended_at,
                        status='absent',
                        tx_hash=lg['tx_hash'] if lg else '',
                        block_num=lg['block_number'] if lg else '',
                        sess_id=sess_id,
                        nfc_id=nid,
                    )
                except Exception as e:
                    print(f"[EMAIL] Failed absence email for {nid}: {e}")

            try:
                users = db_get_all_users()
                teacher_username = sess.get('teacher', '')
                teacher_email = users.get(teacher_username, {}).get('email', '')
                if teacher_email:
                    logs = {lg['nfc_id']: lg for lg in db_get_session_attendance(sess_id)}
                    present_count = late_count = absent_count = excused_count = 0
                    rows = []
                    for st in section_students:
                        nid = st['nfcId']
                        lg = logs.get(nid, {})
                        st_status = (lg.get('status') or 'absent').lower() if lg else 'absent'
                        if st_status == 'late':
                            late_count += 1
                        elif st_status == 'present':
                            present_count += 1
                        elif st_status == 'excused':
                            excused_count += 1
                        else:
                            absent_count += 1
                        rows.append({
                            'name': st.get('name', '—'),
                            'student_id': st.get('student_id', ''),
                            'status': st_status,
                            'tap_time': lg.get('tap_time', '—') if lg else '—',
                            'tx_hash': lg.get('tx_hash', '') if lg else '',
                            'block_num': lg.get('block_number', '') if lg else '',
                        })
                    send_teacher_session_summary(
                        teacher_email=teacher_email,
                        teacher_name=sess.get('teacher_name', ''),
                        subject_name=sess.get('subject_name', ''),
                        section_key=sess.get('section_key', ''),
                        time_slot=sess.get('time_slot', ''),
                        started_at=sess.get('started_at', ''),
                        ended_at=ended_at,
                        present_count=present_count,
                        late_count=late_count,
                        absent_count=absent_count,
                        excused_count=excused_count,
                        student_rows=rows,
                    )
            except Exception as e:
                print(f"[EMAIL] Teacher summary error: {e}")

    if async_chain_and_email:
        Thread(target=_post_finalize_worker, daemon=True).start()
    else:
        _post_finalize_worker()

    return {
        'already_ended': False,
        'ended_at': ended_at,
        'present_count': len([n for n in present_set if n not in late_set]),
        'late_count': len(late_set),
        'absent_count': len(absent_ids),
        'excused_count': len(excused_set),
        'total_enrolled': len(section_students),
    }

def get_active_session_for_nfc(nfc_id):
    all_students = get_all_students()
    student = next((s for s in all_students if s['nfcId'] == nfc_id), None)
    if not student: return None, None
    student_key = build_student_section_key(student)
    if not student_key: return None, None
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM sessions WHERE ended_at IS NULL AND section_key=?",
            (student_key,)
        ).fetchone()
    if row:
        s = _session_row_with_logs(get_db(), row)
        return row['sess_id'], s
    return None, None

# ── AUTH ──────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET','POST'])
def login():
    if 'username' in session:
        return redirect(url_for('index') if session.get('role')=='admin' else url_for('teacher_dashboard'))
    if request.method == 'POST':
        u    = request.form['username'].strip().lower()
        p    = request.form['password']
        user = db_get_user(u)
        if not user or user['password'] != hash_password(p):
            flash('Invalid username or password.'); return redirect(url_for('login'))
        if user['status'] == 'pending':
            flash('Your account is pending admin approval.'); return redirect(url_for('login'))
        if user['status'] == 'rejected':
            flash('Your account was rejected.'); return redirect(url_for('login'))
        session.update({'username':u,'role':user['role'],'full_name':user['full_name']})
        return redirect(url_for('index') if user['role']=='admin' else url_for('teacher_dashboard'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/signup', methods=['GET','POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        confirm  = request.form['confirm_password']
        fullname = request.form['full_name'].strip()
        email    = request.form['email'].strip()
        # Role is always 'teacher' for self-registration.
        # Admins/Super Admins are created only by Super Admin via /superadmin/create-user.
        role     = 'teacher'
        raw_sections = request.form.getlist('sections')
        sections = [normalize_section_key(s) for s in raw_sections]
        if db_get_user(username):    flash('Username already taken.'); return redirect(url_for('signup'))
        if password != confirm:      flash('Passwords do not match.'); return redirect(url_for('signup'))
        if len(password) < 6:        flash('Password must be at least 6 characters.'); return redirect(url_for('signup'))
        db_save_user(username, {
            'username':username,'password':hash_password(password),'role':role,
            'full_name':fullname,'email':email,'status':'pending','sections':sections,
            'my_subjects':[],'created_at':datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        photo_file = request.files.get('profile_photo')
        if photo_file and photo_file.filename:
            ext = os.path.splitext(photo_file.filename)[1].lower()
            if ext in ('.jpg','.jpeg','.png','.gif','.webp'):
                fname = f"photo_{username}{ext}"
                photo_file.save(os.path.join(UPLOAD_FOLDER, fname))
                db_save_photo(username, fname)
        flash('Account created! Waiting for admin approval.'); return redirect(url_for('login'))
    return render_template('signup.html', departments=DEPARTMENTS, year_levels=YEAR_LEVELS, sections=SECTIONS)

# ── ADMIN MAIN ────────────────────────────────────────────────────────────────

@app.route('/')
@admin_required
def index():
    return render_template('index.html',
                           active_sessions=get_active_sessions(),
                           subjects_db=db_get_all_subjects(),
                           users_db=db_get_all_users())

def _register_save_pending_subjects(req, sess):
    import json as _json
    from datetime import datetime as _dt
    import uuid as _uuid
    pending_raw = req.form.get('pending_subjects_json', '[]').strip()
    try:
        pending = _json.loads(pending_raw) if pending_raw else []
    except Exception:
        pending = []
    if not pending:
        return 0
    all_existing = db_get_all_subjects()
    existing_codes = {v.get('course_code', '').upper(): k for k, v in all_existing.items()}
    saved_count = 0
    for subj in pending:
        code_upper = (subj.get('course_code') or '').upper().strip()
        name_val   = (subj.get('name') or '').strip()
        if not name_val:
            continue
        if code_upper and code_upper in existing_codes:
            continue
        new_id = str(_uuid.uuid4())[:8]
        db_save_subject(new_id, {
            'name':        name_val,
            'course_code': subj.get('course_code', ''),
            'units':       str(subj.get('units', '3')),
            'created_by':  sess.get('username', 'admin'),
            'created_at':  _dt.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        if code_upper:
            existing_codes[code_upper] = new_id
        saved_count += 1
    return saved_count

@app.route('/admin/settings', methods=['GET'])
@admin_required
def admin_settings():
    cfg = get_email_config()
    return render_template('admin_settings.html', cfg=cfg)
 
@app.route('/admin/settings/save', methods=['POST'])
@admin_required
def admin_settings_save():
    cfg = {
        'smtp_host':     request.form.get('smtp_host',     'smtp.gmail.com').strip(),
        'smtp_port':     request.form.get('smtp_port',     '587').strip(),
        'smtp_user':     request.form.get('smtp_user',     '').strip(),
        'smtp_password': request.form.get('smtp_password', '').strip(),
        'smtp_from':     request.form.get('smtp_from',     '').strip(),
        'enabled':       '1' if request.form.get('enabled') else '0',
    }
    save_email_config(cfg)
    flash('Email settings saved successfully.')
    return redirect(url_for('admin_settings'))
 
@app.route('/admin/settings/test', methods=['POST'])
@admin_required
def admin_settings_test():
    """Send a test email to verify SMTP config."""
    cfg = get_email_config()
    test_to = request.form.get('test_email', '').strip()
    if not test_to or '@' not in test_to:
        return jsonify({'ok': False, 'message': 'Invalid email address.'})
    if cfg.get('enabled') != '1':
        return jsonify({'ok': False, 'message': 'Email notifications are disabled. Enable them first.'})
    if not cfg.get('smtp_user') or not cfg.get('smtp_password'):
        return jsonify({'ok': False, 'message': 'SMTP credentials not configured.'})
    try:
        import smtplib, ssl
        from email.mime.multipart import MIMEMultipart
        from email.mime.text      import MIMEText
        msg            = MIMEMultipart('alternative')
        msg['Subject'] = '[DAVS] Test Email — SMTP Configuration Verified'
        msg['From']    = cfg.get('smtp_from') or cfg['smtp_user']
        msg['To']      = test_to
        msg.attach(MIMEText(f'''
        <div style="font-family:Arial,sans-serif;padding:24px;max-width:480px;">
          <div style="font-size:20px;font-weight:700;color:#1E4A1A;margin-bottom:8px;">
            ✓ DAVS Email Test Successful
          </div>
          <p style="color:#555;font-size:13px;">
            Your SMTP configuration is working correctly.<br>
            Email notifications will be sent from:
            <strong>{cfg.get("smtp_from") or cfg["smtp_user"]}</strong>
          </p>
          <p style="color:#94a3b8;font-size:11px;margin-top:16px;">
            Cavite State University — DAVS System
          </p>
        </div>''', 'html'))
        ctx  = ssl.create_default_context()
        port = int(cfg.get('smtp_port', 587))
        with smtplib.SMTP(cfg['smtp_host'], port, timeout=10) as srv:
            srv.ehlo(); srv.starttls(context=ctx)
            srv.login(cfg['smtp_user'], cfg['smtp_password'])
            srv.sendmail(msg['From'], [test_to], msg.as_string())
        return jsonify({'ok': True, 'message': f'Test email sent to {test_to}'})
    except Exception as e:
        return jsonify({'ok': False, 'message': str(e)})
 

@app.route('/register', methods=['GET','POST'])
@admin_required
def register():
    if request.method == 'POST':
        nfc_id = request.form['nfc_id'].strip().upper()
        name   = request.form['name'].strip()
        extras = []
        raw_date = request.form.get('date_registered','').strip()
        if raw_date:
            try:
                d = datetime.strptime(raw_date, '%Y-%m')
                raw_date = d.strftime('%B %Y')
            except:
                pass
        section_val = request.form.get('section','').strip().upper()
        
        # Generate CVSU email if not provided or empty
        email_val = request.form.get('email','').strip()
        if not email_val:
            # Generate CVSU pattern email: sc.firstname.lastname@cvsu.edu.ph
            clean = re.sub(r'\b[A-Za-z]\.\s*', '', name).strip()
            clean = re.sub(r'\b(JR|SR|II|III|IV)\.?\b', '', clean, flags=re.IGNORECASE).strip()
            clean = re.sub(r'\s+', ' ', clean)
            words = clean.split()
            if len(words) >= 2:
                first_slug = ''.join(re.sub(r'[^a-z]', '', w.lower()) for w in words[:-1])
                last_slug  = re.sub(r'[^a-z]', '', words[-1].lower())
                if first_slug and last_slug:
                    email_val = f'sc.{first_slug}.{last_slug}@cvsu.edu.ph'
        
        for k,prefix in [('student_id','ID'),('course','Course'),('year_level','Year'),
                          ('adviser','Adviser'),('contact','Tel'),
                          ('semester','Sem'),('school_year','SY')]:
            v = request.form.get(k,'').strip()
            if v: extras.append(f"{prefix}:{v}")
        
        # Add generated/provided email
        if email_val: extras.append(f"Email:{email_val}")
        if section_val: extras.append(f"Sec:{section_val}")
        if raw_date: extras.append(f"RegDate:{raw_date}")
        major = request.form.get('major','').strip() or 'N/A'
        extras.append(f"Major:{major}")
        on_chain = name + (' | ' + ' | '.join(extras) if extras else '')
        pk   = "0x" + secrets.token_hex(32)
        addr = web3.eth.account.from_key(pk).address
        p = parse_student(on_chain)
        student_name_map[nfc_id] = name
        db_save_student({**p, 'nfcId': nfc_id, 'raw_name': on_chain,
                         'address': addr, 'tx_hash': ''})
        photo_file = request.files.get('student_photo')
        if photo_file and photo_file.filename:
            ext = os.path.splitext(photo_file.filename)[1].lower()
            if ext in ('.jpg','.jpeg','.png','.gif','.webp'):
                fname = f"photo_{nfc_id.replace(' ','_')}{ext}"
                photo_file.save(os.path.join(UPLOAD_FOLDER, fname))
                db_save_photo(nfc_id, fname)
        saved_subj = _register_save_pending_subjects(request, session)
        if saved_subj:
            print(f"[INFO] {saved_subj} subject(s) added to catalogue from registration.")
        if BLOCKCHAIN_ONLINE and contract:
            try:
                tx = contract.functions.registerStudent(addr, nfc_id, on_chain).transact({'from': admin_account})
                receipt = web3.eth.wait_for_transaction_receipt(tx)
                tx_hash = receipt['transactionHash'].hex()
                with get_db() as _conn:
                    _conn.execute(
                        "UPDATE students SET reg_tx_hash=?, eth_address=?, updated_at=? WHERE nfc_id=?",
                        (tx_hash, addr, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), nfc_id))
                flash(f'Student {name} registered successfully (blockchain confirmed).')
            except Exception as e:
                msg = str(e)
                if 'already' in msg.lower():
                    flash(f'Student {name} saved. Note: NFC already on blockchain.')
                else:
                    flash(f'Student {name} saved locally. Blockchain error: {e}')
        else:
            flash(f'Student {name} registered successfully (offline mode).')
        return redirect(url_for('index'))
    return render_template('enroll.html', subjects_db=db_get_all_subjects(), users_db=db_get_all_users())


@app.route('/parse_registration_pdf', methods=['POST'])
@admin_required
def parse_registration_pdf():
    import traceback
 
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not (f.filename or '').lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files are supported'}), 400
 
    try:
        raw_bytes = f.read()
        if not raw_bytes:
            return jsonify({'error': 'Uploaded file is empty'}), 400
 
        full = _parse_cvsu_pdf_text(raw_bytes)
        if not full.strip():
            return jsonify({
                'error': 'Could not extract text. Install pypdf: pip install pypdf'
            }), 400
 
        print(f'[PDF] {len(full)} chars from "{f.filename}"')
        print(f'[PDF] PREVIEW:\n{full[:600]}\n[PDF] END')
 
        result = _extract_cvsu_fields(full)
 
        print(f'[PDF] name="{result["name"]}" id="{result["student_id"]}" '
              f'course="{result["course"]}" year="{result["year_level"]}" '
              f'sec="{result["section"]}" subjects={len(result["subjects"])} '
              f'email="{result["email"]}"')
 
        return jsonify({**result, 'added_to_catalogue': []})
 
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[parse_registration_pdf ERROR]\n{tb}')
        return jsonify({'error': f'{type(e).__name__}: {str(e)}'}), 500

@app.route('/batch_register', methods=['GET', 'POST'])
@admin_required
def batch_register():
    if request.method == 'GET':
        return render_template('batch_register.html',
                               subjects_db=db_get_all_subjects(),
                               users_db=db_get_all_users())
 
    # POST: receive JSON list of students with nfc_id already assigned
    students_json = request.form.get('students_data', '[]')
    try:
        students_in = json.loads(students_json)
    except Exception:
        flash('Invalid student data payload.')
        return redirect(url_for('batch_register'))
 
    # Filter to only students that have an nfc_id (skipped ones are excluded)
    students_in = [s for s in students_in if s.get('nfc_id')]
 
    success_count  = 0
    blockchain_ok  = 0
    errors         = []
    subjects_saved = 0
 
    # ── Save all subjects across all students to the catalogue ────────────
    existing_subj  = db_get_all_subjects()
    existing_codes = {
        v.get('course_code', '').upper(): k
        for k, v in existing_subj.items()
        if v.get('course_code')
    }
    for student in students_in:
        for subj in (student.get('subjects') or []):
            code_upper = (subj.get('course_code') or '').upper().strip()
            name_val   = (subj.get('name') or '').strip()
            if not name_val:
                continue
            if code_upper and code_upper in existing_codes:
                continue
            new_id = str(uuid.uuid4())[:8]
            db_save_subject(new_id, {
                'name':        name_val,
                'course_code': subj.get('course_code', ''),
                'units':       str(subj.get('units', '3')),
                'created_by':  session.get('username', 'admin'),
                'created_at':  datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            })
            if code_upper:
                existing_codes[code_upper] = new_id
            subjects_saved += 1
 
    # ── Register each student ─────────────────────────────────────────────
    for student in students_in:
        try:
            nfc_id = student.get('nfc_id', '').strip().upper()
            name   = (student.get('name') or '').strip()
 
            if not nfc_id:
                errors.append(f"Missing NFC ID for {name or 'Unknown'}")
                continue
            if not name:
                errors.append(f"Missing name for NFC {nfc_id}")
                continue
 
            # Build on-chain registration string
            extras    = []
            email_val = (student.get('email') or '').strip()
            if not email_val:
                email_val = _generate_cvsu_email(name)
 
            for field, prefix in [
                ('student_id',  'ID'),
                ('course',      'Course'),
                ('year_level',  'Year'),
                ('adviser',     'Adviser'),
                ('contact',     'Tel'),
                ('semester',    'Sem'),
                ('school_year', 'SY'),
            ]:
                v = (student.get(field) or '').strip()
                if v:
                    extras.append(f"{prefix}:{v}")
 
            if email_val:
                extras.append(f"Email:{email_val}")
 
            section_val = (student.get('section') or '').strip().upper()
            if section_val:
                extras.append(f"Sec:{section_val}")
 
            date_reg = (student.get('date_registered') or '').strip()
            if date_reg:
                extras.append(f"RegDate:{date_reg}")
 
            major = (student.get('major') or 'N/A').strip()
            extras.append(f"Major:{major}")
 
            on_chain = name + (' | ' + ' | '.join(extras) if extras else '')
 
            # Generate Ethereum address
            pk   = "0x" + secrets.token_hex(32)
            addr = web3.eth.account.from_key(pk).address
 
            # Parse and save to SQLite cache
            p = parse_student(on_chain)
            student_name_map[nfc_id] = name
            db_save_student({
                **p,
                'nfcId':      nfc_id,
                'raw_name':   on_chain,
                'address':    addr,
                'tx_hash':    '',
                'photo_file': student.get('photo_file', ''),
            })
            success_count += 1
 
            # Register on blockchain (non-fatal if offline)
            if BLOCKCHAIN_ONLINE and contract and admin_account:
                try:
                    tx = contract.functions.registerStudent(
                        addr, nfc_id, on_chain
                    ).transact({'from': admin_account})
                    receipt  = web3.eth.wait_for_transaction_receipt(tx)
                    tx_hash  = receipt['transactionHash'].hex()
                    blk      = receipt['blockNumber']
                    with get_db() as _conn:
                        _conn.execute(
                            "UPDATE students SET reg_tx_hash=?, eth_address=?, "
                            "reg_block=?, updated_at=? WHERE nfc_id=?",
                            (tx_hash, addr, blk,
                             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), nfc_id)
                        )
                    blockchain_ok += 1
                except Exception as be:
                    msg = str(be)
                    if 'already' in msg.lower():
                        blockchain_ok += 1  # already on chain — that's fine
                    else:
                        errors.append(f"Blockchain error for {name}: {be}")
 
        except Exception as e:
            errors.append(f"Error registering {student.get('name', 'Unknown')}: {e}")
 
    # Flash results
    if success_count:
        chain_note = (f" {blockchain_ok} confirmed on blockchain."
                      if BLOCKCHAIN_ONLINE else " (offline — saved to SQLite cache)")
        subj_note  = f" {subjects_saved} subject(s) added to catalogue." if subjects_saved else ""
        flash(f"Registered {success_count} student(s) successfully.{chain_note}{subj_note}")
 
    if errors:
        flash("Errors: " + " | ".join(errors[:5]) +
              (f" (+{len(errors)-5} more)" if len(errors) > 5 else ""))
 
    return redirect(url_for('dashboard'))

@app.route('/parse_batch_pdfs', methods=['POST'])
@admin_required
def parse_batch_pdfs():
    import traceback
 
    if 'files' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
 
    files     = request.files.getlist('files')
    pdf_files = [f for f in files if f.filename and f.filename.lower().endswith('.pdf')]
    if not pdf_files:
        return jsonify({'error': 'No PDF files found in the upload'}), 400
 
    students = []
    errors   = []
 
    for f in pdf_files:
        filename = f.filename or 'unknown.pdf'
        try:
            raw_bytes = f.read()
            if not raw_bytes:
                errors.append(f'Empty file: {filename}')
                continue
 
            full = _parse_cvsu_pdf_text(raw_bytes)
            if not full.strip():
                errors.append(f'Could not extract text from: {filename}')
                continue
 
            result = _extract_cvsu_fields(full)
 
            # Fallback: derive name from filename if PDF gave nothing
            # Handles:  LASTNAME_FIRSTNAME MIDDLE.pdf
            #           AMBATA_JHAY VIC_GUILLERMO.pdf
            if not result['name']:
                base = re.sub(r'\.pdf$', '', filename, flags=re.IGNORECASE)
                if '_' in base:
                    parts  = base.split('_', 1)
                    last   = parts[0].replace('-', ' ').strip().title()
                    first  = parts[1].replace('_', ' ').replace('-', ' ').strip().title()
                    result['name'] = f"{first} {last}"
                else:
                    result['name'] = base.replace('-', ' ').replace('_', ' ').strip().title()
 
                if result['name']:
                    print(f'[BATCH PDF] Name from filename: "{result["name"]}"')
                    if not result['email']:
                        result['email'] = _generate_cvsu_email(result['name'])
 
            if not result['name'] and not result['student_id']:
                errors.append(f'No usable data extracted from: {filename}')
                continue
 
            result['filename'] = filename
            students.append(result)
 
            print(f'[BATCH PDF] OK  "{filename}" -> "{result["name"]}" '
                  f'id="{result["student_id"]}" yr="{result["year_level"]}" '
                  f'sec="{result["section"]}" subj={len(result["subjects"])}')
 
        except Exception as e:
            tb = traceback.format_exc()
            print(f'[BATCH PDF ERROR] {filename}\n{tb}')
            errors.append(f'{filename}: {type(e).__name__} — {str(e)}')
 
    if not students:
        return jsonify({
            'error':   'No valid student data could be extracted.',
            'details': errors,
        }), 400
 
    # Sort alphabetically by surname (e.g. Ambata before Berongoy)
    students.sort(key=_surname_sort_key)
    for i, s in enumerate(students):
        s['sort_order'] = i + 1
 
    return jsonify({
        'students':     students,
        'total_parsed': len(students),
        'errors':       errors,
        'sorted_by':    'surname_first',
    })

@app.route('/mark', methods=['POST'])
@login_required
def mark():
    nfc_id = request.form['nfc_id'].strip().upper()
    try:
        tx = contract.functions.markAttendanceWithStatus(
            nfc_id, chain_status_code('present')
        ).transact({'from':admin_account})
        web3.eth.wait_for_transaction_receipt(tx)
        name = student_name_map.get(nfc_id,"Unknown")
        recent_attendance.append({'nfc_id':nfc_id,'name':name,'timestamp':time.time()})
        flash(f'Attendance marked for {name}')
    except Exception as e:
        flash(f'Error: {e}')
    return redirect(url_for('index'))

@app.route('/dashboard')
@admin_required
def dashboard():
    students  = get_all_students()
    all_users = db_get_all_users()
    teachers  = {u:d for u,d in all_users.items() if d.get('role') in ('admin','teacher')}
    with get_db() as _conn:
        _photo_rows = _conn.execute("SELECT person_id, filename FROM photos").fetchall()
    photos_db = {r['person_id']: r['filename'] for r in _photo_rows}
    return render_template('dashboard.html',
        students=students, teachers=teachers,
        photos_db=photos_db, fmt_time=fmt_time)

@app.route('/upload_photo', methods=['POST'])
@login_required
def upload_photo():
    person_id = request.form.get('person_id','').strip()
    if not person_id or 'photo' not in request.files:
        return jsonify({'error':'Missing data'}), 400
    f = request.files['photo']
    if not f or f.filename == '':
        return jsonify({'error':'No file selected'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.jpg','.jpeg','.png','.gif','.webp'):
        return jsonify({'error':'Only image files allowed'}), 400
    filename = f"photo_{person_id.replace(' ','_')}{ext}"
    f.save(os.path.join(UPLOAD_FOLDER, filename))
    db_save_photo(person_id, filename)
    return jsonify({'ok':True,'filename':filename,'url':f'/static/uploads/{filename}'})

@app.route('/get_my_photo')
@login_required
def get_my_photo():
    photo = db_get_photo(session.get('username',''))
    if photo: return jsonify({'url':f'/static/uploads/{photo}'})
    return jsonify({'url':None})

@app.route('/api/my_profile')
@login_required
def api_my_profile():
    user = db_get_user(session.get('username',''))
    if not user: return jsonify({'error':'Not found'}), 404
    return jsonify({
        'username':  user.get('username',''),
        'full_name': user.get('full_name',''),
        'email':     user.get('email',''),
        'role':      user.get('role',''),
        'status':    user.get('status',''),
        'created':   user.get('created_at',''),
        'sections':  user.get('sections',[]),
        'photo':     db_get_photo(session.get('username','')) or '',
    })

@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    username = session.get('username')
    user = db_get_user(username)
    if not user: return jsonify({'error':'Not logged in'}), 401
    data = request.get_json()
    if data.get('full_name'):
        user['full_name'] = data['full_name'].strip()
        session['full_name'] = user['full_name']
    if data.get('email'):
        user['email'] = data['email'].strip()
    if data.get('password') and len(data['password']) >= 6:
        user['password'] = hash_password(data['password'])
    new_username = (data.get('new_username') or '').strip().lower()
    if new_username and new_username != username:
        if db_get_user(new_username):
            return jsonify({'error': 'Username already taken'}), 409
        db_save_user(new_username, user)
        db_delete_user(username)
        try:
            photo = db_get_photo(username)
            if photo:
                db_save_photo(new_username, photo)
                db_delete_photo(username)
        except Exception:
            pass
        session['username'] = new_username
        username = new_username
    else:
        db_save_user(username, user)
    return jsonify({'ok': True, 'full_name': user['full_name'], 'username': session.get('username', username)})

@app.route('/delete_photo', methods=['POST'])
@login_required
def delete_photo():
    person_id = request.form.get('person_id','').strip()
    existing  = db_get_photo(person_id)
    if existing:
        old_file = os.path.join(UPLOAD_FOLDER, existing)
        if os.path.exists(old_file):
            try: os.remove(old_file)
            except: pass
        db_delete_photo(person_id)
    return jsonify({'ok':True})

@app.route('/reports')
@admin_required
def attendance_report():
    report = []
    for s in get_all_students():
        stats = get_student_attendance_stats(s['nfcId'])
        report.append({**s, **stats})
    return render_template('attendance_report.html',
                           students=sorted(report, key=lambda x: -x['rate']),
                           subjects_db=db_get_all_subjects(), fmt_time=fmt_time)

@app.route('/view/<nfc_id>')
@login_required
def view_attendance(nfc_id):
    if session.get('role') == 'teacher':
        user = get_current_user()
        if not any(s['nfcId']==nfc_id for s in teacher_students(user)):
            flash('Access denied.'); return redirect(url_for('teacher_dashboard'))
    records      = get_attendance_records(nfc_id)
    student_info = next((s for s in get_all_students() if s['nfcId']==nfc_id), {})
    return render_template('attendance.html', nfc_id=nfc_id, records=records,
                           student=student_info, fmt_time=fmt_time)

@app.route('/api/student_sessions/<nfc_id>')
@login_required
def student_sessions_api(nfc_id):
    with get_db() as conn:
        log_rows = conn.execute(
            "SELECT al.nfc_id, al.status, al.tx_hash, al.block_number, "
            "al.tap_time, al.excuse_note, "
            "s.sess_id, s.subject_name, s.course_code, s.section_key, "
            "s.teacher_name, s.time_slot, s.started_at "
            "FROM attendance_logs al "
            "JOIN sessions s ON al.sess_id = s.sess_id "
            "WHERE al.nfc_id = ? "
            "ORDER BY s.started_at DESC",
            (nfc_id,)
        ).fetchall()
    result = []
    seen_sessions = set()
    for row in log_rows:
        seen_sessions.add(row['sess_id'])
        result.append({
            'subject_name': row['subject_name'] or '',
            'course_code':  row['course_code']  or '',
            'teacher_name': row['teacher_name'] or '',
            'section_key':  row['section_key']  or '',
            'time_slot':    row['time_slot']     or '',
            'date':         (row['started_at'] or '')[:10],
            'started_at':   row['started_at']   or '',
            'status':       (row['status'] or 'absent').lower(),
            'tap_time':     row['tap_time']      or '',
            'tx_hash':      row['tx_hash']       or '',
            'block':        str(row['block_number']) if row['block_number'] else '',
            'excuse_note':  row['excuse_note']   or '',
        })
    if not result:
        all_students    = get_all_students()
        student         = next((x for x in all_students if x['nfcId']==nfc_id), None)
        student_section = build_student_section_key(student) if student else ''
        with get_db() as conn:
            sess_rows = conn.execute(
                "SELECT * FROM sessions WHERE ended_at IS NOT NULL ORDER BY started_at DESC"
            ).fetchall()
        for row in sess_rows:
            s   = _row_to_dict(row)
            sec = normalize_section_key(s.get('section_key',''))
            if student_section and sec != student_section: continue
            if s.get('sess_id') in seen_sessions: continue
            if   nfc_id in s.get('excused',[]): status = 'excused'
            elif nfc_id in s.get('late',   []): status = 'late'
            elif nfc_id in s.get('present',[]): status = 'present'
            elif student_section == sec:        status = 'absent'
            else: continue
            tx_info = s.get('tx_hashes',{}).get(nfc_id,{})
            result.append({
                'subject_name': s.get('subject_name',''),
                'course_code':  s.get('course_code',''),
                'teacher_name': s.get('teacher_name',''),
                'section_key':  s.get('section_key',''),
                'time_slot':    s.get('time_slot',''),
                'date':         (s.get('started_at','') or '')[:10],
                'started_at':   s.get('started_at',''),
                'status':       status,
                'tap_time':     '',
                'tx_hash':      tx_info.get('tx_hash',''),
                'block':        str(tx_info.get('block','')),
                'excuse_note':  '',
            })
    result.sort(key=lambda x: x['date'], reverse=True)
    return jsonify(result)

@app.route('/update_student', methods=['POST'])
@admin_required
def update_student():
    data   = request.get_json()
    nfc_id = data.get('nfc_id','').strip()
    if not nfc_id: return jsonify({'error':'Missing nfc_id'}), 400
    fields = ['full_name','student_id','email','contact','adviser','major',
              'semester','school_year','date_registered','course','year_level','section']
    update_data = {f: (data.get(f) or '').strip() for f in fields}
    if update_data.get('section'):
        update_data['section'] = update_data['section'].strip().upper()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    set_parts = []
    params    = []
    db_col_map = {
        'full_name':'full_name','student_id':'student_id','email':'email',
        'contact':'contact','adviser':'adviser','major':'major',
        'semester':'semester','school_year':'school_year',
        'date_registered':'date_registered','course':'program',
        'year_level':'year_level','section':'section',
    }
    for field, db_col in db_col_map.items():
        val = update_data.get(field,'')
        if val:
            set_parts.append(f"{db_col}=?")
            params.append(val)
    if set_parts:
        set_parts.append("updated_at=?")
        params.append(now)
        params.append(nfc_id)
        with get_db() as conn:
            conn.execute(
                f"UPDATE students SET {', '.join(set_parts)} WHERE nfc_id=?",
                params
            )
    override_data = {f: update_data[f] for f in fields if update_data.get(f)}
    if override_data:
        db_save_override(nfc_id, override_data)
    return jsonify({'ok': True})

@app.route('/update_faculty', methods=['POST'])
@admin_required
def update_faculty():
    data     = request.get_json()
    username = data.get('username','').strip()
    user     = db_get_user(username)
    if not user: return jsonify({'error':'User not found'}), 404
    if data.get('full_name'):  user['full_name'] = data['full_name'].strip()
    if data.get('email') is not None: user['email'] = data['email'].strip()
    # Role update: normal admin can only set 'teacher'; super_admin can set any role
    new_role = data.get('role', '')
    requester_role = session.get('role', '')
    if new_role:
        if requester_role == 'super_admin' and new_role in ('teacher', 'admin', 'super_admin'):
            user['role'] = new_role
        elif requester_role == 'admin' and new_role == 'teacher':
            user['role'] = new_role
        elif new_role not in ('teacher', 'admin', 'super_admin'):
            return jsonify({'error': 'Invalid role.'}), 400
        else:
            return jsonify({'error': 'You do not have permission to assign this role.'}), 403
    if data.get('status') in ('approved','pending','rejected'): user['status'] = data['status']
    if 'sections' in data and isinstance(data['sections'], list):
        user['sections'] = [normalize_section_key(s) for s in data['sections']]
    new_pw = (data.get('new_password') or '').strip()
    if new_pw:
        if len(new_pw) < 6:
            return jsonify({'error':'Password must be at least 6 characters'}), 400
        user['password'] = hash_password(new_pw)
    new_username = (data.get('new_username') or '').strip().lower()
    if new_username and new_username != username:
        if db_get_user(new_username):
            return jsonify({'error':f'Username "{new_username}" is already taken'}), 409
        user['username'] = new_username
        db_save_user(new_username, user)
        db_delete_user(username)
        db_rename_photo_key(username, new_username)
    else:
        db_save_user(username, user)
    return jsonify({'ok':True})

@app.route('/export')
@login_required
def export_page():
    return render_template('export.html',
        students=get_all_students(), users_db=db_get_all_users(),
        subjects_db=db_get_all_subjects(), active_sessions=get_active_sessions(),
        fmt_time=fmt_time)

@app.route('/export/all.csv')
@admin_required
def export_csv_all():
    out = io.StringIO(); w = csv.writer(out)
    w.writerow(['Name','NFC ID','Student ID','Course','Year','Section','Adviser','Email','Contact','Date & Time','Status'])
    for s in get_all_students():
        for ts,p in (get_attendance_records(s['nfcId']) or [('No records','')]):
            w.writerow([s['name'],s['nfcId'],s['student_id'],s['course'],s['year_level'],s['section'],
                        s['adviser'],s['email'],s['contact'],ts,'Present' if p is True else ('Absent' if p is False else '')])
    out.seek(0)
    return Response(out.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition':f'attachment; filename=attendance_all_{datetime.now().strftime("%Y%m%d")}.csv'})

@app.route('/export/<nfc_id>.csv')
@login_required
def export_csv_single(nfc_id):
    name = student_name_map.get(nfc_id,'Unknown'); out = io.StringIO(); w = csv.writer(out)
    w.writerow(['Student Name','NFC ID','Date & Time','Status'])
    for ts,p in get_attendance_records(nfc_id):
        w.writerow([name,nfc_id,ts,'Present' if p else 'Absent'])
    out.seek(0)
    return Response(out.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition':f'attachment; filename=attendance_{nfc_id}.csv'})

@app.route('/admin/users')
@admin_required
def manage_users():
    all_u    = db_get_all_users()
    pending  = {u:d for u,d in all_u.items() if d['status']=='pending'}
    approved = {u:d for u,d in all_u.items() if d['status']=='approved' and u!='admin'}
    rejected = {u:d for u,d in all_u.items() if d['status']=='rejected'}
    photos_db = db_get_all_photos()
    return render_template('admin_users.html', pending=pending, approved=approved,
                           rejected=rejected, fmt_time=fmt_time, photos_db=photos_db)


@app.route('/admin/approve/<username>', methods=['POST'])
@app.route('/admin/users/<username>/approve', methods=['POST'])
@admin_required
def approve_user(username):
    user = db_get_user(username)
    if user: user['status']='approved'; db_save_user(username,user); flash(f'{user["full_name"]} approved.')
    return redirect(url_for('manage_users'))

@app.route('/admin/reject/<username>', methods=['POST'])
@app.route('/admin/users/<username>/reject', methods=['POST'])
@admin_required
def reject_user(username):
    user = db_get_user(username)
    if user: user['status']='rejected'; db_save_user(username,user); flash(f'{user["full_name"]} rejected.')
    return redirect(url_for('manage_users'))

@app.route('/admin/delete/<username>', methods=['POST'])
@app.route('/admin/users/<username>/delete', methods=['POST'])
@admin_required
def delete_user(username):
    user = db_get_user(username)
    if user and username != 'admin':
        db_delete_user(username); flash(f'{user["full_name"]} deleted.')
    return redirect(url_for('manage_users'))

@app.route('/admin/subjects')
@admin_required
def manage_subjects():
    return render_template('admin_subjects.html', subjects=db_get_all_subjects(), fmt_time=fmt_time)

@app.route('/admin/subjects/add', methods=['POST'])
@admin_required
def add_subject():
    name = request.form.get('name','').strip()
    if not name: flash('Subject name cannot be empty.'); return redirect(url_for('manage_subjects'))
    for s in db_get_all_subjects().values():
        if s['name'].lower() == name.lower():
            flash(f'Subject "{name}" already exists.'); return redirect(url_for('manage_subjects'))
    course_code = request.form.get('course_code','').strip().upper()
    units = request.form.get('units','3').strip()
    if units not in ('2','3'): units = '3'
    sid = str(uuid.uuid4())[:8]
    db_save_subject(sid, {'name':name,'course_code':course_code,'units':units,
                          'created_by':session.get('username',''),
                          'created_at':datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    flash(f'Subject "{name}" added.')
    return redirect(url_for('manage_subjects'))

@app.route('/admin/subjects/<sid>/rename', methods=['POST'])
@app.route('/admin/subjects/rename/<sid>', methods=['POST'])
@admin_required
def rename_subject(sid):
    if request.is_json:
        data     = request.get_json()
        new_name = (data.get('name') or '').strip()
        if not new_name: return jsonify({'error':'Name cannot be empty'}), 400
        subj = db_get_subject(sid)
        if not subj: return jsonify({'error':'Subject not found'}), 404
        subj['name'] = new_name
        if 'course_code' in data: subj['course_code'] = data['course_code'].strip().upper()
        if 'units' in data and str(data['units']) in ('2','3'): subj['units'] = str(data['units'])
        db_save_subject(sid, subj)
        return jsonify({'ok':True})
    new_name = request.form.get('name','').strip()
    if not new_name: flash('Name cannot be empty.'); return redirect(url_for('manage_subjects'))
    subj = db_get_subject(sid)
    if subj:
        old = subj['name']; subj['name'] = new_name
        db_save_subject(sid, subj); flash(f'"{old}" renamed to "{new_name}".')
    return redirect(url_for('manage_subjects'))

@app.route('/admin/subjects/delete/<sid>', methods=['POST'])
@app.route('/admin/subjects/<sid>/delete', methods=['POST'])
@admin_required
def delete_subject(sid):
    for s in get_active_sessions().values():
        if s.get('subject_id') == sid:
            flash('Cannot delete — a live session is using this subject.'); return redirect(url_for('manage_subjects'))
    subj = db_get_subject(sid)
    if subj:
        db_delete_subject(sid)
        with get_db() as conn:
            rows = conn.execute("SELECT username, my_subjects_json FROM users").fetchall()
        for row in rows:
            my_s  = json.loads(row['my_subjects_json'] or '[]')
            new_s = [ms for ms in my_s if ms.get('subject_id') != sid]
            if len(new_s) != len(my_s):
                with get_db() as conn:
                    conn.execute("UPDATE users SET my_subjects_json=? WHERE username=?",
                                 (json.dumps(new_s), row['username']))
        flash(f'Subject "{subj["name"]}" deleted.')
    return redirect(url_for('manage_subjects'))

@app.route('/admin/sessions')
@admin_required
def admin_sessions():
    with get_db() as _conn:
        _active_rows = _conn.execute("SELECT * FROM sessions WHERE ended_at IS NULL").fetchall()
        _ended_rows  = _conn.execute("SELECT * FROM sessions WHERE ended_at IS NOT NULL ORDER BY ended_at DESC").fetchall()
        active = {r['sess_id']: _session_row_with_logs(_conn, r) for r in _active_rows}
        ended  = {r['sess_id']: _session_row_with_logs(_conn, r) for r in _ended_rows}
    return render_template('admin_sessions.html', active=active, ended=ended,
                           subjects_db=db_get_all_subjects(), fmt_time=fmt_time)

def _build_teacher_context(user):
    if not user: return {}, [], []
    students = teacher_students(user)
    sections = {}
    for key in user.get('sections', []):
        norm_key = normalize_section_key(key)
        parts    = norm_key.split('|')
        sec_stud = [s for s in students if build_student_section_key(s) == norm_key]
        sections[norm_key] = {
            'label':   f"{parts[2]} — {parts[1]}" if len(parts)==3 else norm_key,
            'course':  parts[0] if parts else '',
            'year':    parts[1] if len(parts)>1 else '',
            'section': parts[2] if len(parts)>2 else '',
            'students': sec_stud, 'count': len(sec_stud)
        }
    all_subj = db_get_all_subjects()
    my_subjects = []
    for ms in user.get('my_subjects', []):
        sid  = ms.get('subject_id')
        skey = normalize_section_key(ms.get('section_key',''))
        if not sid or sid not in all_subj: continue
        active_sid = None
        for sess_id, sess_obj in get_active_sessions().items():
            if (sess_obj.get('subject_id')==sid
                    and normalize_section_key(sess_obj.get('section_key',''))==skey
                    and sess_obj.get('teacher')==session['username']):
                active_sid = sess_id; break
        parts = skey.split('|')
        subj_info = all_subj[sid]
        # Get student count from ALL students matching this section key
        # (not just from teacher.sections dict which may be empty)
        all_stu_for_section = [s for s in db_get_all_students() if build_student_section_key(s) == skey]
        sec_count = len(all_stu_for_section) if all_stu_for_section else (sections[skey]['count'] if skey in sections else 0)
        if skey in sections:
            sec_label = sections[skey]['label']
        else:
            sec_label = f"{parts[2]} — {parts[1]}" if len(parts)==3 else skey
        my_subjects.append({
            'subject_id':   sid,
            'subject_name': subj_info['name'],
            'course_code':  subj_info.get('course_code',''),
            'units':        subj_info.get('units','3'),
            'section_key':  skey,
            'section_label': sec_label,
            'course':       parts[0] if parts else '',
            'year':         parts[1] if len(parts)>1 else '',
            'section':      parts[2] if len(parts)>2 else '',
            'active_session_id': active_sid,
            'student_count':     sec_count
        })
    return sections, my_subjects, all_subj

@app.route('/teacher')
@login_required
def teacher_dashboard():
    if session.get('role') == 'admin':
        return redirect(url_for('index'))
    user = get_current_user()
    if not user:
        session.clear()
        return redirect(url_for('login'))
    sections, my_subjects, _ = _build_teacher_context(user)
    live_sessions = {sid: s for sid, s in get_active_sessions().items()
                     if s.get('teacher') == session['username']
                     or s.get('teacher_name') == session.get('full_name')}
    with get_db() as conn:
        total_sessions = conn.execute(
            "SELECT COUNT(*) FROM sessions "
            "WHERE (teacher_username=? OR teacher_name=?) AND ended_at IS NOT NULL",
            (session['username'], session.get('full_name', ''))
        ).fetchone()[0]
    all_teacher_students = teacher_students(user)
    total_students = len(all_teacher_students)
    return render_template('teacher_dashboard.html',
        user=user, sections=sections, my_subjects=my_subjects,
        live_sessions=live_sessions, total_sessions=total_sessions,
        total_students=total_students, fmt_time=fmt_time, fmt_time_short=fmt_time_short)

@app.route('/teacher/sessions-students')
@login_required
def teacher_sessions_students():
    if session.get('role') == 'admin':
        return redirect(url_for('index'))
    user = get_current_user()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM sessions "
            "WHERE teacher_username=? OR teacher_name=? "
            "ORDER BY started_at DESC",
            (session['username'], session.get('full_name', ''))
        ).fetchall()
        sessions_data = {r['sess_id']: _session_row_with_logs(conn, r) for r in rows}
    subjects = sorted(set(
        s.get('subject_name', '') for s in sessions_data.values()
        if s.get('subject_name')
    ))
    report = []
    for s in teacher_students(user):
        stats = get_student_attendance_stats(s['nfcId'])
        report.append({**s, **stats})
    students = sorted(report, key=lambda x: -x['rate'])
    return render_template('teacher_sessions_students.html',
        user=user, sessions_data=sessions_data,
        sessions_json={sid: {
            'subject_name': s.get('subject_name', ''),
            'course_code':  s.get('course_code', ''),
            'section_key':  s.get('section_key', ''),
            'teacher_name': s.get('teacher_name', ''),
            'started_at':   s.get('started_at', ''),
            'ended_at':     s.get('ended_at', ''),
            'time_slot':    s.get('time_slot', ''),
            'present':      s.get('present', []),
            'late':         s.get('late', []),
            'excused':      s.get('excused', []),
            'tx_hashes':    s.get('tx_hashes', {}),
        } for sid, s in sessions_data.items()},
        subjects=subjects, students=students,
        now=str(datetime.now().year), fmt_time=fmt_time, fmt_time_short=fmt_time_short)

@app.route('/teacher/create-session')
@login_required
def teacher_create_session():
    if session.get('role') == 'admin': return redirect(url_for('index'))
    user = get_current_user()
    if not user: session.clear(); return redirect(url_for('login'))
    sections, my_subjects, all_subj = _build_teacher_context(user)
    return render_template('teacher_create_session.html', user=user, sections=sections,
                           my_subjects=my_subjects, subjects_db=all_subj)

@app.route('/teacher/records')
@login_required
def teacher_records():
    if session.get('role') == 'admin': return redirect(url_for('attendance_report'))
    user = get_current_user()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM sessions WHERE teacher_username=? AND ended_at IS NOT NULL ORDER BY started_at DESC",
            (session['username'],)
        ).fetchall()
        teacher_sessions_data = {r['sess_id']: _session_row_with_logs(conn, r) for r in rows}
    subjects = sorted(set(s.get('subject_name','') for s in teacher_sessions_data.values() if s.get('subject_name')))
    all_students = get_all_students()
    return render_template('teacher_records.html', user=user,
                           sessions_data=teacher_sessions_data, subjects=subjects,
                           all_students=all_students, fmt_time=fmt_time)

@app.route('/api/session_attendance/<sess_id>')
@login_required
def api_session_attendance(sess_id):
    try:
        sess = load_session(sess_id)
        if sess is None:
            return jsonify({'error': 'Session not found'}), 404
        if not _is_my_session(sess):
            return jsonify({'error': 'Access denied'}), 403
        logs = db_get_session_attendance(sess_id)
        logs_by_nfc = {lg['nfc_id']: lg for lg in logs}
        section_key = normalize_section_key(sess.get('section_key', ''))
        sk_parts    = section_key.split('|')
        program     = sk_parts[0] if len(sk_parts) > 0 else ''
        year_level  = sk_parts[1] if len(sk_parts) > 1 else ''
        section_val = sk_parts[2] if len(sk_parts) > 2 else ''

        # Get excuse request details
        excuse_details = {}
        with get_db() as _conn:
            excuses = _conn.execute(
                "SELECT nfc_id, reason_type, reason_detail, attachment_file FROM excuse_requests WHERE sess_id=? AND status='approved'",
                (sess_id,)
            ).fetchall()
            for exc in excuses:
                excuse_details[exc['nfc_id']] = {
                    'reason': exc['reason_type'],
                    'reason_detail': exc['reason_detail'],
                    'attachment_file': exc['attachment_file'],
                }

        # Historical-first view:
        # - always include students who have logs in this session (even if moved section later)
        # - also include currently enrolled students with no logs as "absent"
        students_map = {}

        for lg in logs:
            nid = lg['nfc_id']
            st  = get_student_by_nfc_cached(nid) or {}
            excuse_info = excuse_details.get(nid, {})
            students_map[nid] = {
                'nfc_id':     nid,
                'name':       lg.get('student_name') or st.get('name') or nid,
                'student_id': lg.get('student_id')   or st.get('student_id', ''),
                'status':     (lg.get('status') or 'absent').lower(),
                'tx_hash':    lg.get('tx_hash') or '',
                'block':      str(lg.get('block_number') or ''),
                'time':       lg.get('tap_time') or '',
                'reason':     excuse_info.get('reason') or lg.get('excuse_note') or '',
                'reason_detail': excuse_info.get('reason_detail') or '',
                'attachment_url': url_for('admin_excuse_attachment', excuse_id=lg.get('excuse_request_id')) if lg.get('excuse_request_id') else '',
            }

        # Primary approach: use get_all_students() and match by section_key
        # This is more reliable as it handles both blockchain-sourced and database students
        all_students = get_all_students()
        enrolled = [s for s in all_students if build_student_section_key(s) == section_key]

        # Fallback: if no students found via section_key, try exact database query
        if not enrolled and program and year_level and section_val:
            with get_db() as _conn:
                _rows = _conn.execute(
                    "SELECT * FROM students WHERE program=? AND year_level=? AND section=?",
                    (program, year_level, section_val)
                ).fetchall()
            enrolled = [_student_row(r) for r in _rows]

        for s in enrolled:
            nid = s['nfcId']
            if nid in students_map:
                continue
            students_map[nid] = {
                'nfc_id':     nid,
                'name':       s.get('name', nid),
                'student_id': s.get('student_id', ''),
                'status':     'absent',
                'tx_hash':    '',
                'block':      '',
                'time':       '',
                'reason':     '',
                'reason_detail': '',
                'attachment_url': '',
            }

        students_out = sorted(students_map.values(), key=lambda x: (x.get('name') or 'Unknown').lower())
        return jsonify({
            'students':     students_out,
            'subject_name': sess.get('subject_name', ''),
            'course_code':  sess.get('course_code', ''),
            'section_key':  section_key,
            'time_slot':    sess.get('time_slot', ''),
            'started_at':   sess.get('started_at', ''),
            'ended_at':     sess.get('ended_at', ''),
        })
    except Exception as e:
        import traceback
        print(f"[ERROR api_session_attendance] {e}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/teacher/subjects/add', methods=['POST'])
@login_required
def teacher_add_subject():
    if session.get('role') == 'admin': return redirect(url_for('index'))
    user       = get_current_user()
    subject_id = request.form.get('subject_id','').strip()
    section_key= normalize_section_key(request.form.get('section_key','').strip())
    if not subject_id or not section_key:
        flash('Please select both a subject and a section.'); return redirect(url_for('teacher_create_session'))
    subj = db_get_subject(subject_id)
    if not subj: flash('Subject not found.'); return redirect(url_for('teacher_create_session'))
    for ms in user.get('my_subjects',[]):
        if ms['subject_id']==subject_id and normalize_section_key(ms['section_key'])==section_key:
            flash('Already assigned.'); return redirect(url_for('teacher_create_session'))
    user.setdefault('my_subjects',[]).append({'subject_id':subject_id,'section_key':section_key})
    db_save_user(session['username'], user)
    flash(f'Subject "{subj["name"]}" added to your schedule.')
    return redirect(url_for('teacher_create_session'))

@app.route('/teacher/subjects/<subject_id>/<path:section_key>/remove', methods=['POST'])
@login_required
def teacher_remove_subject(subject_id, section_key):
    if session.get('role') == 'admin': return redirect(url_for('index'))
    user = get_current_user()
    norm_key = normalize_section_key(section_key)
    user['my_subjects'] = [ms for ms in user.get('my_subjects',[])
                           if not (ms['subject_id']==subject_id
                                   and normalize_section_key(ms['section_key'])==norm_key)]
    db_save_user(session['username'], user)
    flash('Subject removed from your schedule.')
    return redirect(url_for('teacher_create_session'))

def _parse_time_slot_end(time_slot_str, reference_dt):
    """
    Parse a time slot like '9:00 AM – 11:00 AM' and return
    a datetime for the END time on the same date as reference_dt.
    Returns None if parsing fails.
    """
    if not time_slot_str:
        return None
    try:
        # Handle both '–' (en-dash) and '-' (hyphen) separators
        sep = '–' if '–' in time_slot_str else '-'
        parts = time_slot_str.split(sep)
        if len(parts) < 2:
            return None
        end_str = parts[1].strip()
        # Parse end time — e.g. "11:00 AM"
        end_t   = datetime.strptime(end_str, '%I:%M %p').time()
        return reference_dt.replace(
            hour=end_t.hour, minute=end_t.minute,
            second=0, microsecond=0
        )
    except Exception as _pe:
        print(f"[WARN] Could not parse time slot end: {_pe}")
        return None
 
 
def _parse_time_slot_start(time_slot_str, reference_dt):
    """Parse the START time of a time slot string."""
    if not time_slot_str:
        return None
    try:
        sep     = '–' if '–' in time_slot_str else '-'
        parts   = time_slot_str.split(sep)
        start_s = parts[0].strip()
        start_t = datetime.strptime(start_s, '%I:%M %p').time()
        return reference_dt.replace(
            hour=start_t.hour, minute=start_t.minute,
            second=0, microsecond=0
        )
    except Exception:
        return None
 
 
def _auto_end_session_thread(sess_id, auto_end_at_str, app_context):
    """
    Background thread: waits until auto_end_at then ends the session
    automatically (same logic as end_session route).
    Teacher can still end early by clicking End Session.
    """
    import threading as _th
    try:
        auto_end_dt = datetime.strptime(auto_end_at_str, '%Y-%m-%d %H:%M:%S')
        wait_secs   = (auto_end_dt - datetime.now()).total_seconds()
        if wait_secs > 0:
            print(f"[AUTO-END] Session {sess_id} will auto-end in {int(wait_secs)}s at {auto_end_at_str}")
            _th.Event().wait(timeout=wait_secs)
 
        # Check if already ended manually by teacher
        with app_context:
            current = load_session(sess_id)
            if not current or current.get('ended_at'):
                print(f"[AUTO-END] Session {sess_id} already ended — skipping auto-end.")
                return
 
            print(f"[AUTO-END] Auto-ending session {sess_id}...")
            result = _finalize_session(
                sess_id,
                ended_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                async_chain_and_email=True,
            )
            if result and not result.get('already_ended'):
                print(f"[AUTO-END] Session {sess_id} ended automatically at {result.get('ended_at')}")
 
    except Exception as _ate:
        print(f"[AUTO-END] Error in auto-end thread for {sess_id}: {_ate}")
 
 
@app.route('/teacher/session/start', methods=['POST'])
@login_required
def start_session():
    if session.get('role') == 'admin': return redirect(url_for('index'))
    subject_id  = request.form.get('subject_id','').strip()
    section_key = normalize_section_key(request.form.get('section_key','').strip())
    if not subject_id or not section_key:
        flash('Missing subject or section.'); return redirect(url_for('teacher_create_session'))
    for s in get_active_sessions().values():
        if (s.get('teacher')==session['username']
                and normalize_section_key(s.get('section_key',''))==section_key):
            flash('You already have an active session for that section.')
            return redirect(url_for('teacher_create_session'))
 
    time_slot    = request.form.get('time_slot','').strip()
    grace_period = int(request.form.get('grace_period', 15) or 15)
    grace_period = max(1, min(grace_period, 120))  # clamp 1–120 min
 
    subj_data = db_get_subject(subject_id) or {}
    units     = int(subj_data.get('units', 3))
    now       = datetime.now()
    from datetime import timedelta
 
    # ── Late cutoff: based on actual session start + grace period ─────────────
    # Teacher sets grace period (default 15 min). Students tapping after
    # this cutoff are marked LATE instead of PRESENT.
    late_cutoff_dt = now.replace(second=0, microsecond=0) + timedelta(minutes=grace_period)
 
    # ── Auto-end: derived from time slot end time ─────────────────────────────
    # If teacher selects "9:00 AM – 11:00 AM", session auto-ends at 11:00 AM.
    # Teacher can still end early by clicking End Session.
    auto_end_dt  = _parse_time_slot_end(time_slot, now)
    auto_end_str = auto_end_dt.strftime('%Y-%m-%d %H:%M:%S') if auto_end_dt else None
 
    sess_id  = str(uuid.uuid4())[:12]
    new_sess = {
        'subject_id':    subject_id,
        'subject_name':  subj_data.get('name',''),
        'course_code':   subj_data.get('course_code',''),
        'units':         units,
        'time_slot':     time_slot,
        'section_key':   section_key,
        'teacher':       session['username'],
        'teacher_name':  session['full_name'],
        'started_at':    now.strftime('%Y-%m-%d %H:%M:%S'),
        'late_cutoff':   late_cutoff_dt.strftime('%Y-%m-%d %H:%M:%S'),
        'auto_end_at':   auto_end_str,
        'grace_period':  grace_period,
        'ended_at':      None,
        'present':[],'late':[],'excused':[],'warned':[],'absent':[],
        'tap_log':[],'warn_log':[],'invalid_log':[],'excuse_notes':{},'tx_hashes':{}
    }
    save_session(sess_id, new_sess)
    sessions_db[sess_id] = new_sess
 
    # ── Launch auto-end background thread ─────────────────────────────────────
    if auto_end_str:
        import threading as _th
        ctx = app.app_context()
        t   = _th.Thread(
            target=_auto_end_session_thread,
            args=(sess_id, auto_end_str, ctx),
            daemon=True
        )
        t.start()
        flash(f'Session started. Auto-ends at {auto_end_dt.strftime("%I:%M %p")}. Grace period: {grace_period} min.')
    else:
        flash(f'Classroom session started for {subj_data.get("name","")}.')
 
    return redirect(url_for('live_session', sess_id=sess_id))

@app.route('/api/blockchain_status')
def blockchain_status():
    global BLOCKCHAIN_ONLINE
    try:
        BLOCKCHAIN_ONLINE = web3.is_connected()
    except:
        BLOCKCHAIN_ONLINE = False
    student_count = len(db_get_all_students())
    return jsonify({
        'online': BLOCKCHAIN_ONLINE,
        'student_cache_count': student_count,
        'message': 'Blockchain online' if BLOCKCHAIN_ONLINE else f'Offline — {student_count} students loaded from cache'
    })

def _is_my_session(sess):
    """
    Return True if the current logged-in user owns this session.
    Checks both teacher_username (new) and teacher_name (fallback for old records),
    so sessions created before the schema migration still work.
    """
    if session.get('role') in ('admin', 'super_admin'):
        return True
    username   = session.get('username', '')
    full_name  = session.get('full_name', '')
    # sess['teacher'] is set by _session_row_with_logs from teacher_username column
    # sess['teacher_name'] is the display name column
    teacher_u  = sess.get('teacher', '') or ''
    teacher_n  = sess.get('teacher_name', '') or ''
    return (teacher_u == username) or (teacher_n and teacher_n == full_name)


@app.route('/teacher/session/<sess_id>')
@login_required
def live_session(sess_id):
    sess = load_session(sess_id)
    if sess is None: flash('Session not found.'); return redirect(url_for('teacher_dashboard'))
    if not _is_my_session(sess):
        flash('Access denied.'); return redirect(url_for('teacher_dashboard'))
    all_students     = get_all_students()
    section_key      = sess.get('section_key','')
    section_students = [s for s in all_students
                        if build_student_section_key(s) == normalize_section_key(section_key)]
    present_set  = set(sess.get('present',[]))
    late_set     = set(sess.get('late',[]))
    excused_set  = set(sess.get('excused',[]))
    student_statuses = []
    excuse_notes = sess.get('excuse_notes', {})
    for s in section_students:
        nid = s['nfcId']
        if   nid in excused_set: status = 'excused'
        elif nid in late_set:    status = 'late'
        elif nid in present_set: status = 'present'
        else:                    status = 'absent'
        student_statuses.append({**s, 'status': status, 'reason': excuse_notes.get(nid, '')})
    return render_template('session_live.html', sess=sess, sess_id=sess_id,
                           section_students=section_students,
                           student_statuses=student_statuses,
                           present_list=[s for s in section_students if s['nfcId'] in present_set],
                           absent_list=[s for s in section_students if s['nfcId'] not in present_set and s['nfcId'] not in excused_set],
                           tap_log=sess.get('tap_log',[]),
                           is_active=not sess.get('ended_at'),
                           fmt_time=fmt_time, fmt_time_short=fmt_time_short)

@app.route('/teacher/session/<sess_id>/end', methods=['POST'])
@login_required
def end_session(sess_id):
    sess = load_session(sess_id)
    if sess is None:
        flash('Session not found.'); return redirect(url_for('teacher_dashboard'))
    
    if not _is_my_session(sess):
        flash('Access denied.'); return redirect(url_for('teacher_dashboard'))
    result = _finalize_session(
        sess_id,
        ended_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        async_chain_and_email=True,
    )
    if not result:
        flash('Session not found.'); return redirect(url_for('teacher_dashboard'))

    flash(
        f"Session ended. {result.get('present_count', 0)} present, "
        f"{result.get('late_count', 0)} late, {result.get('absent_count', 0)} absent.")
    return redirect(url_for('teacher_dashboard'))

@app.route('/teacher/session/<sess_id>/excuse', methods=['POST'])
@login_required
def excuse_student(sess_id):
    nfc_id = request.form.get('nfc_id', '').strip().upper()
    reason_type = request.form.get('reason_type', '').strip()
    reason_detail = request.form.get('reason_detail', '').strip()
    sess = load_session(sess_id)
    if sess is None:
        return jsonify({'error': 'Session not found'}), 404
    if not _is_my_session(sess):
        return jsonify({'error': 'Access denied'}), 403
    if not nfc_id:
        return jsonify({'error': 'Missing NFC ID'}), 400
    if reason_type not in dict(EXCUSE_REASONS):
        return jsonify({'error': f'Invalid reason type: {reason_type}'}), 400
    if reason_type == 'others' and not reason_detail.strip():
        return jsonify({'error': 'Details required for "Others"'}), 400
    
    # Handle multiple file uploads — save the first valid one for DB record
    attachment_file = ''
    uploaded_files = request.files.getlist('attachments')
    for uploaded in uploaded_files:
        if uploaded and uploaded.filename:
            try:
                attachment_file = _save_excuse_attachment(uploaded)
                break  # save first valid attachment
            except ValueError as ve:
                return jsonify({'error': str(ve)}), 400
    
    student = db_get_student(nfc_id) or get_student_by_nfc_cached(nfc_id)
    if not student:
        return jsonify({'error': 'Student not found'}), 404
    
    reason_label = dict(EXCUSE_REASONS).get(reason_type, reason_type.title())

    excuse_id = db_save_excuse_request({
        'sess_id':        sess_id,
        'nfc_id':         nfc_id,
        'student_name':   student.get('name', nfc_id),
        'student_id':     student.get('student_id', ''),
        'student_email':  student.get('email', ''),
        'reason_type':    reason_type,
        'reason_detail':  reason_detail,
        'attachment_file': attachment_file,
    })
    db_resolve_excuse(excuse_id, 'approved', session.get('username', ''))
    sess = load_session(sess_id) or sess
    
    # Blockchain with auto-registration
    exc_tx = ''; exc_block = 0
    if BLOCKCHAIN_ONLINE and contract and admin_account:
        try:
            on_chain = contract.functions.studentsByNfc(nfc_id).call()
            is_registered_on_chain = on_chain[2]
            if not is_registered_on_chain:
                try:
                    import hashlib as _hl
                    pk_seed = _hl.sha256(f"davs-student-{nfc_id}".encode()).hexdigest()
                    pk = "0x" + pk_seed
                    addr = web3.eth.account.from_key(pk).address
                    raw_name = student.get('raw_name', '') or student.get('name', '')
                    reg_tx = contract.functions.registerStudent(
                        addr, nfc_id, raw_name
                    ).transact({'from': admin_account})
                    reg_receipt = web3.eth.wait_for_transaction_receipt(reg_tx)
                    with get_db() as _conn:
                        _conn.execute(
                            "UPDATE students SET reg_tx_hash=?, eth_address=?, reg_block=?, updated_at=? WHERE nfc_id=?",
                            (reg_receipt['transactionHash'].hex(), addr,
                            reg_receipt['blockNumber'],
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            nfc_id)
                        )
                except Exception as reg_err:
                    print(f"[WARNING] Auto-registration failed: {reg_err}")
            exc_tx, exc_block = mark_attendance_on_chain(nfc_id, 'excused')
        except Exception as _e:
            print(f"[WARN] excused blockchain tx failed: {_e}")
    
    db_save_attendance_log(
        sess_id=sess_id, nfc_id=nfc_id,
        student_name=student.get('name', nfc_id),
        student_id=student.get('student_id', ''),
        status='excused',
        tap_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        tx_hash=exc_tx, block_number=exc_block,
        excuse_note=f"{reason_label}{' — ' + reason_detail if reason_detail else ''}"
    )
    with get_db() as conn:
        conn.execute(
            "UPDATE attendance_logs SET excuse_request_id=? WHERE sess_id=? AND nfc_id=?",
            (excuse_id, sess_id, nfc_id)
        )
    
    send_student_attendance_receipt(
        student_name=student.get('name', ''),
        student_email=student.get('email', ''),
        student_id=student.get('student_id', ''),
        subject_name=sess.get('subject_name', ''),
        section_key=sess.get('section_key', ''),
        teacher_name=sess.get('teacher_name', ''),
        tap_time=datetime.now().strftime('%B %d, %Y  %I:%M %p'),
        status='excused',
        tx_hash=exc_tx,
        block_num=exc_block,
    )
    
    name = student.get('name', nfc_id)
    return jsonify({'status': 'ok', 'name': name, 'nfc_id': nfc_id, 'reason': reason_label})

@app.route('/teacher/sessions')
@login_required
def teacher_sessions():
    if session.get('role') == 'admin': return redirect(url_for('index'))
    with get_db() as _conn:
        _rows = _conn.execute("SELECT * FROM sessions WHERE teacher_username=? ORDER BY started_at DESC",
                              (session['username'],)).fetchall()
    my_sessions = {r['sess_id']: _row_to_dict(r) for r in _rows}
    return render_template('teacher_sessions.html', my_sessions=my_sessions, fmt_time=fmt_time)

@app.route('/teacher/reports')
@login_required
def teacher_reports():
    if session.get('role') == 'admin': return redirect(url_for('attendance_report'))
    user   = get_current_user()
    report = []
    for s in teacher_students(user):
        stats = get_student_attendance_stats(s['nfcId'])
        report.append({**s, **stats})
    return render_template('teacher_reports.html', user=user,
                           students=sorted(report, key=lambda x: -x['rate']))

@app.route('/teacher/export/section.csv')
@login_required
def teacher_export():
    user     = get_current_user()
    sec_key  = request.args.get('section','')
    students = teacher_students(user)
    if sec_key:
        norm_key = normalize_section_key(sec_key)
        students = [s for s in students if build_student_section_key(s) == norm_key]
    out = io.StringIO(); w = csv.writer(out)
    w.writerow(['Name','NFC ID','Student ID','Course','Year','Section','Date & Time','Status'])
    for s in students:
        for ts,p in (get_attendance_records(s['nfcId']) or [('No records','')]):
            w.writerow([s['name'],s['nfcId'],s['student_id'],s['course'],s['year_level'],s['section'],
                        ts,'Present' if p is True else ('Absent' if p is False else '')])
    out.seek(0)
    return Response(out.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition':f'attachment; filename=section_{datetime.now().strftime("%Y%m%d")}.csv'})

@app.route('/api/attendance/recent')
@login_required
def recent_attendance_api():
    since = request.args.get('since', type=float, default=0)
    user  = get_current_user()
    evts  = [e for e in recent_attendance if e['timestamp'] > since]
    if user and user.get('role') == 'teacher':
        my_nfcs = {s['nfcId'] for s in teacher_students(user)}
        evts    = [e for e in evts if e['nfc_id'] in my_nfcs]
    return jsonify(evts)

@app.route('/api/session/<sess_id>/poll')
@login_required
def poll_session(sess_id):
    """
    Poll for new taps.

    KEY FIX: Uses server-side time (returned as 'server_time') so the frontend
    never relies on the browser clock. Browser clocks can be ahead of the server
    which caused taps to be permanently missed (since_dt > created_at → never found).

    Tap detection uses attendance_logs.created_at (real DB timestamp) instead of
    tap_log[].timestamp which is always 0 when loaded from DB.
    """
    since = request.args.get('since', type=float, default=0)
    now_ts = time.time()  # server time — returned to client so they stay in sync

    # Reload session from DB so present/late/excused lists are always fresh
    sess = load_session(sess_id)
    if sess is None:
        return jsonify({'error': 'not found', 'active': False}), 404
    sessions_db[sess_id] = sess

    new_taps     = []
    new_warnings = []
    new_invalids = []

    if since > 0:
        # Subtract 2 seconds as a buffer to handle sub-second timing edge cases
        # and any minor clock drift between DB writes and poll timing
        since_buffered = since - 2
        since_dt = datetime.fromtimestamp(since_buffered).strftime('%Y-%m-%d %H:%M:%S')

        with get_db() as conn:
            tap_rows = conn.execute(
                "SELECT al.nfc_id, al.student_name, al.student_id, al.status, "
                "al.tap_time, al.tx_hash, al.block_number, al.created_at "
                "FROM attendance_logs al "
                "WHERE al.sess_id = ? "
                "  AND al.status IN ('present','late') "
                "  AND al.created_at > ? "
                "ORDER BY al.created_at ASC",
                (sess_id, since_dt)
            ).fetchall()

        for row in tap_rows:
            try:
                ts = datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S').timestamp()
            except Exception:
                ts = since + 1
            new_taps.append({
                'nfc_id':     row['nfc_id'],
                'name':       row['student_name'],
                'student_id': row['student_id'],
                'time':       row['tap_time'],
                'timestamp':  ts,
                'tx_hash':    row['tx_hash'],
                'block':      row['block_number'],
                'is_late':    row['status'] == 'late',
            })

        # Warnings/invalids from in-memory (not in DB) — use in-memory session
        mem_sess = sessions_db.get(sess_id, sess)
        new_warnings = [t for t in mem_sess.get('warn_log', [])
                        if t.get('timestamp', 0) > since_buffered]
        new_invalids = [t for t in mem_sess.get('invalid_log', [])
                        if t.get('timestamp', 0) > since_buffered]

    return jsonify({
        'present_count': len(sess.get('present', [])),
        'late_count':    len(sess.get('late', [])),
        'excused_count': len(sess.get('excused', [])),
        'warned_count':  len(sess.get('warned', [])),
        'new_taps':      new_taps,
        'new_warnings':  new_warnings,
        'new_invalids':  new_invalids,
        'active':        not sess.get('ended_at'),
        'late_ids':      sess.get('late', []),
        'excused_ids':   sess.get('excused', []),
        'present_ids':   sess.get('present', []),
        'server_time':   now_ts,
        'auto_end_at':   sess.get('auto_end_at'),
        'grace_period':  sess.get('grace_period', 15),
    })

@app.route('/api/attendance/stats')
@login_required
def attendance_stats():
    period     = request.args.get('period',      'today')
    f_month    = request.args.get('month',       '').strip()
    f_year_num = request.args.get('year_num',    '').strip()
    f_subject  = request.args.get('subject',     '').strip()
    f_section  = request.args.get('section_key', request.args.get('section', '')).strip()
    f_year_lvl = request.args.get('year_level',  '').strip()
    f_program  = request.args.get('program',     '').strip()
    f_sec_ltr  = request.args.get('section_letter','').strip()
    f_instr    = request.args.get('instructor',  '').strip()
    f_tod      = request.args.get('time_of_day', '').strip()
    role       = session.get('role')
    username   = session.get('username')
    now        = datetime.now()
    if not f_year_num:
        # Backward compatibility for older clients still using `year`.
        f_year_num = request.args.get('year', '').strip() if period in ('month', 'year') else ''
    if period == 'today':
        start_dt = now.replace(hour=0,minute=0,second=0,microsecond=0); end_dt = None
    elif period == 'month':
        yr = int(f_year_num) if f_year_num and f_year_num.isdigit() else now.year
        mo = int(f_month)    if f_month    and f_month.isdigit()    else now.month
        start_dt = datetime(yr,mo,1)
        end_dt   = datetime(yr,mo,_cal.monthrange(yr,mo)[1],23,59,59)
    elif period == 'year':
        yr = int(f_year_num) if f_year_num and f_year_num.isdigit() else now.year
        start_dt = datetime(yr,1,1); end_dt = datetime(yr,12,31,23,59,59)
    else:
        start_dt = datetime(2000,1,1); end_dt = None
    where  = ["s.started_at >= ?"]
    params = [start_dt.strftime('%Y-%m-%d %H:%M:%S')]
    if end_dt:
        where.append("s.started_at <= ?"); params.append(end_dt.strftime('%Y-%m-%d %H:%M:%S'))
    if role == 'teacher':
        where.append("s.teacher_username = ?"); params.append(username)
    if f_section:
        where.append("s.section_key = ?"); params.append(normalize_section_key(f_section))
    if f_program:
        where.append("s.section_key LIKE ?"); params.append(f_program + '%')
    if f_year_lvl:
        where.append("s.section_key LIKE ?"); params.append('%|' + f_year_lvl + '|%')
    if f_sec_ltr:
        where.append("s.section_key LIKE ?"); params.append('%|' + f_sec_ltr)
    if f_subject:
        where.append("s.subject_name = ?"); params.append(f_subject)
    if f_instr:
        where.append("s.teacher_name = ?"); params.append(f_instr)
    if f_tod == 'morning':
        where.append("CAST(strftime('%H',s.started_at) AS INTEGER) < 12")
    elif f_tod == 'afternoon':
        where.append("CAST(strftime('%H',s.started_at) AS INTEGER) >= 12")
    elif f_tod and ':' in f_tod:
        where.append("s.time_slot = ?"); params.append(f_tod)
    wsql = " AND ".join(where)
    if period == 'today':   tkey_expr = "strftime('%H:00',s.started_at)"
    elif period == 'month': tkey_expr = "strftime('%m/%d',s.started_at)"
    elif period == 'year':  tkey_expr = "strftime('%m',s.started_at)"
    else:                   tkey_expr = "strftime('%Y',s.started_at)"
    with get_db() as conn:
        donut_rows = conn.execute(
            "SELECT al.status, COUNT(*) as cnt "
            "FROM attendance_logs al "
            "JOIN sessions s ON al.sess_id = s.sess_id "
            "WHERE " + wsql + " GROUP BY al.status", params
        ).fetchall()
        trend_rows = conn.execute(
            "SELECT " + tkey_expr + " as tkey, al.status, COUNT(*) as cnt "
            "FROM attendance_logs al "
            "JOIN sessions s ON al.sess_id = s.sess_id "
            "WHERE " + wsql + " GROUP BY tkey, al.status ORDER BY tkey", params
        ).fetchall()
        subj_rows = conn.execute(
            "SELECT s.subject_name, s.course_code, al.status, COUNT(*) as cnt "
            "FROM attendance_logs al "
            "JOIN sessions s ON al.sess_id = s.sess_id "
            "WHERE " + wsql + " GROUP BY s.subject_name, s.course_code, al.status", params
        ).fetchall()
        sess_count = conn.execute(
            "SELECT COUNT(DISTINCT s.sess_id) as cnt FROM sessions s WHERE " + wsql, params
        ).fetchone()['cnt']
        subj_labels_rows = conn.execute(
            "SELECT DISTINCT s.subject_name, s.course_code FROM sessions s "
            "WHERE " + wsql + " ORDER BY s.subject_name", params
        ).fetchall()
    donut = {'present':0,'late':0,'absent':0,'excused':0}
    for r in donut_rows:
        if r['status'] in donut: donut[r['status']] = r['cnt']
    trend_buckets = {}
    for r in trend_rows:
        k = r['tkey']
        if k not in trend_buckets:
            trend_buckets[k] = {'present':0,'late':0,'absent':0,'excused':0}
        if r['status'] in trend_buckets[k]:
            trend_buckets[k][r['status']] = r['cnt']
    subjects_breakdown = {}
    for r in subj_rows:
        code  = r['course_code']
        label = f"[{code}] {r['subject_name']}" if code else r['subject_name']
        if label not in subjects_breakdown:
            subjects_breakdown[label] = {'present':0,'late':0,'absent':0,'excused':0,'sessions':0}
        if r['status'] in subjects_breakdown[label]:
            subjects_breakdown[label][r['status']] = r['cnt']
    subj_labels_out = [
        (f"[{r['course_code']}] {r['subject_name']}" if r['course_code'] else r['subject_name'])
        for r in subj_labels_rows
    ]
    resp = jsonify({
        'role': role, 'period': period, 'donut': donut,
        'trend': trend_buckets, 'subjects': subjects_breakdown,
        'all_subjects': subj_labels_out, 'session_count': sess_count,
    })
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    resp.headers['Pragma']        = 'no-cache'
    resp.headers['Expires']       = '0'
    return resp

@app.route('/api/block_number')
def api_block_number():
    try: return jsonify({'block':web3.eth.block_number})
    except: return jsonify({'block':None})

# ── MARK PICO (NFC tap handler) ───────────────────────────────────────────────

@app.route('/mark_pico', methods=['POST'])
def mark_pico():
    data=request.get_json()
    if not data or 'nfc_id' not in data: return jsonify({'status':'error'}), 400
    nfc_id=data['nfc_id'].strip().upper()
    print(f"[NFC TAP] {nfc_id}")

    if nfc_is_waiting():
        nfc_set_uid(nfc_id)
        return jsonify({'status':'registration','uid':nfc_id})

    # ✅ FIX: Check DB for excused students instead of in-memory sessions_db
    # This ensures that after app restart, excused students are still blocked
    with get_db() as conn:
        excused_logs = conn.execute(
            "SELECT DISTINCT sess_id FROM attendance_logs "
            "WHERE nfc_id=? AND status='excused'",
            (nfc_id,)
        ).fetchall()
    
    # Check if any of those sessions are still active
    if excused_logs:
        active_sessions = get_active_sessions()
        for log in excused_logs:
            if log['sess_id'] in active_sessions:
                return jsonify({'status': 'excused',
                                'message': 'This student is marked as Excused and cannot tap in.',
                                'nfc_id': nfc_id})

    sess_id, sess = get_active_session_for_nfc(nfc_id)
    if not sess:
        all_s   = get_all_students()
        student = next((s for s in all_s if s['nfcId']==nfc_id), None)
        active  = get_active_sessions()
        invalid_entry={'nfc_id':nfc_id,'timestamp':time.time(),
                       'reason':'Student not registered' if not student else 'No active session for this section'}
        for sid, asess in active.items():
            asess.setdefault('invalid_log',[]).append(invalid_entry)
            save_session(sid, asess)
        return jsonify({'status':'no_session',
                        'message':"No active session for this student's section.",
                        'debug_student':student,
                        'debug_active_sessions':list(active.keys())})

    name        = student_name_map.get(nfc_id,'Unknown')
    all_s       = get_all_students()
    student_info= next((s for s in all_s if s['nfcId']==nfc_id),{})
    student_id  = student_info.get('student_id','')

    # Check duplicate tap
    if nfc_id in sess.get('present',[]):
        warn_entry={'nfc_id':nfc_id,'name':name,'student_id':student_id,'timestamp':time.time()}
        if nfc_id not in sess.get('warned',[]): sess.setdefault('warned',[]).append(nfc_id)
        sess.setdefault('warn_log',[]).append(warn_entry)
        save_session(sess_id, sess)
        sessions_db[sess_id] = sess
        return jsonify({'status':'already_marked','name':name,'student_id':student_id,
                        'message':f'{name} is already marked present.'})

    # Determine late status
    now_dt       = datetime.now()
    late_cutoff  = sess.get('late_cutoff','')
    is_late      = False
    if late_cutoff:
        try:
            cutoff_dt = datetime.strptime(late_cutoff, '%Y-%m-%d %H:%M:%S')
            is_late   = now_dt > cutoff_dt
        except:
            is_late = False

    # Record on blockchain (skip gracefully if offline)
    tx_hash=None; block_num=None
    if BLOCKCHAIN_ONLINE and contract and admin_account:
            try:
                # Step 1: Check if student is registered on-chain
                on_chain = contract.functions.studentsByNfc(nfc_id).call()
                is_registered_on_chain = on_chain[2]  # isRegistered bool
    
                # Step 2: Auto-register on-chain if not registered
                # This handles students enrolled via SQLite (seed data, offline mode, etc.)
                if not is_registered_on_chain:
                    print(f"[BLOCKCHAIN] Student {nfc_id} not on-chain — auto-registering...")
                    try:
                        # Generate a deterministic address from nfc_id for offline-enrolled students
                        import hashlib as _hl
                        pk_seed   = _hl.sha256(f"davs-student-{nfc_id}".encode()).hexdigest()
                        pk        = "0x" + pk_seed
                        addr      = web3.eth.account.from_key(pk).address
                        raw_name  = student_info.get('raw_name', '') or name
                        reg_tx    = contract.functions.registerStudent(
                            addr, nfc_id, raw_name
                        ).transact({'from': admin_account})
                        reg_receipt = web3.eth.wait_for_transaction_receipt(reg_tx)
                        reg_tx_hash = reg_receipt['transactionHash'].hex()
                        reg_block   = reg_receipt['blockNumber']
                        # Update SQLite with the on-chain registration proof
                        with get_db() as _conn:
                            _conn.execute(
                                "UPDATE students SET reg_tx_hash=?, eth_address=?, "
                                "reg_block=?, updated_at=? WHERE nfc_id=?",
                                (reg_tx_hash, addr,
                                reg_block,
                                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                nfc_id)
                            )
                        print(f"[BLOCKCHAIN] Auto-registered {name} — TX: {reg_tx_hash} Block: {reg_block}")
                    except Exception as reg_err:
                        print(f"[WARNING] Auto-registration failed for {nfc_id}: {reg_err}")
                        # Continue anyway — markAttendance will fail but we still save to SQLite
    
                # Step 3: Mark attendance on blockchain → get TX hash + block number
                tx_hash, block_num = mark_attendance_on_chain(nfc_id, 'late' if is_late else 'present')
                print(f"[BLOCKCHAIN] Attendance marked — TX: {tx_hash} Block: {block_num}")
    
            except Exception as e:
                print(f"[WARNING] Blockchain mark failed: {e} — attendance saved to SQLite only.")

    tap_time      = datetime.now().strftime('%H:%M:%S')
    tap_timestamp = time.time()
    status_label  = 'late' if is_late else 'present'

    # Save to attendance_logs table
    db_save_attendance_log(
        sess_id=sess_id, nfc_id=nfc_id,
        student_name=name, student_id=student_id,
        status=status_label, tap_time=tap_time,
        tx_hash=tx_hash or '', block_number=block_num or 0
    )

    # Keep in-memory session dict in sync for live polling
    sess.setdefault('present',[]).append(nfc_id)
    if is_late and nfc_id not in sess.get('late',[]):
        sess.setdefault('late',[]).append(nfc_id)

    # FIX: always set tap_timestamp so poll_session can detect new taps
    sess.setdefault('tap_log',[]).append({
        'nfc_id':     nfc_id,
        'name':       name,
        'time':       tap_time,
        'timestamp':  tap_timestamp,
        'tx_hash':    tx_hash,
        'block':      block_num,
        'student_id': student_id,
        'is_late':    is_late,
    })
    sess.setdefault('tx_hashes',{})[nfc_id] = {
        'tx_hash': tx_hash, 'block': block_num, 'time': tap_time
    }
    save_session(sess_id, sess)
    sessions_db[sess_id] = sess

    recent_attendance.append({
        'nfc_id':    nfc_id,
        'name':      name,
        'timestamp': tap_timestamp,
        'subject':   sess.get('subject_name',''),
        'is_late':   is_late,
    })
    
    # ── Email: send attendance receipt to student ─────────────────────────
    student_email = student_info.get('email', '')
    send_student_attendance_receipt(
        student_name   = name,
        student_email  = student_email,
        student_id     = student_id,
        subject_name   = sess.get('subject_name', ''),
        section_key    = sess.get('section_key', ''),
        teacher_name   = sess.get('teacher_name', ''),
        tap_time       = datetime.now().strftime('%B %d, %Y  %I:%M %p'),
        status         = status_label,
        tx_hash        = tx_hash or '',
        block_num      = block_num or '',
    )

    return jsonify({
        'status':'ok','name':name,'student_id':student_id,
        'time':tap_time,'subject':sess.get('subject_name',''),
        'tx_hash':tx_hash,'block':block_num,
        'attendance_status': status_label,
        'is_late': is_late
    })

@app.route('/debug/tap/<nfc_id>')
def debug_tap(nfc_id):
    nfc_id=nfc_id.strip().upper()
    all_students=get_all_students()
    student=next((s for s in all_students if s['nfcId']==nfc_id),None)
    active=get_active_sessions()
    student_key=build_student_section_key(student) if student else None
    matching_session=None
    for sid,s in active.items():
        if normalize_section_key(s.get('section_key',''))==student_key:
            matching_session={**s,'session_id':sid}; break
    result={
        '1_nfc_id_received':nfc_id,
        '2_student_found':student is not None,
        '3_student_info':student,
        '4_student_section_key':student_key,
        '5_active_sessions':[{
            'session_id':sid,'subject':s.get('subject_name'),
            'section_key':s.get('section_key'),'teacher':s.get('teacher_name')
        } for sid,s in active.items()],
        '6_matching_session':matching_session,
        '7_verdict':('STUDENT NOT REGISTERED' if not student
                     else 'NO ACTIVE SESSION' if not matching_session
                     else 'SHOULD WORK')
    }
    from flask import current_app
    resp=current_app.response_class(json.dumps(result,indent=2),mimetype='application/json')
    return resp

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT HELPERS — shared colour palette + formatters
# ══════════════════════════════════════════════════════════════════════════════
def _xl_helpers():
    """Return a dict of reusable Excel style helpers."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    C = {
        'bg':      '1E4A1A', 'header':  '1E4A1A', 'accent':  '2D6A27',
        'gold':    'F5C518', 'surface': 'FFFFFF',  'border':  'D4DDD4',
        'present': '2D6A27', 'late':    'D4A017',  'absent':  'C0392B', 'excused': '2980B9',
        'present_bg': 'E8F5E9', 'late_bg': 'FFF8E1', 'absent_bg': 'FFEBEE', 'excused_bg': 'E3F2FD',
        'white':   'FFFFFF', 'row_alt': 'F0F2F0', 'row_def': 'FFFFFF',
        'muted':   '5A6B5A', 'sub_hdr': '2D6A27',
    }
    def fill(h):  return PatternFill('solid', fgColor=h)
    def thin_border():
        s = Side(style='thin', color='CBD5E1')
        return Border(left=s, right=s, top=s, bottom=s)
    def ctr(wrap=True):  return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    def lft():  return Alignment(horizontal='left', vertical='center', wrap_text=True)
    def rgt():  return Alignment(horizontal='right', vertical='center', wrap_text=True)
    def header_font(size=10):
        return Font(name='Calibri', bold=True, color='FFFFFF', size=size)
    def normal_font(size=10, color='111111', bold=False):
        return Font(name='Calibri', size=size, color=color, bold=bold)
    def title_font(size=16, color=None):
        return Font(name='Calibri', bold=True, size=size, color=color or C['accent'])
    def make_header_row(ws, row_num, headers, widths, bg=None):
        bg = bg or C['header']
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(ci)].width = w
            c = ws.cell(row=row_num, column=ci, value=h)
            c.font = header_font(); c.fill = fill(bg)
            c.alignment = ctr(); c.border = thin_border()
        ws.row_dimensions[row_num].height = 22
    def data_row(ws, row_num, values, alt=False, col_formats=None):
        rf = fill(C['row_alt'] if alt else C['row_def'])
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.border = thin_border()
            cf = (col_formats or {}).get(ci)
            if cf and cf[0] == 'status':
                status = val
                status_colors = {'Present': C['present'], 'Late': C['late'], 'Absent': C['absent'], 'Excused': C['excused']}
                fg = status_colors.get(status, '111111')
                c.font = Font(name='Calibri', size=10, bold=True, color=fg)
                c.fill = rf; c.alignment = ctr()
            elif cf and cf[0] == 'tx':
                c.font = Font(name='Courier New', size=8, color=C['muted'])
                c.fill = rf; c.alignment = lft()
            elif cf and cf[0] == 'num':
                c.font = normal_font(); c.fill = rf; c.alignment = ctr()
            else:
                c.font = normal_font()
                c.fill = rf
                c.alignment = lft() if ci <= 3 else ctr()
        ws.row_dimensions[row_num].height = 17
    def title_block(ws, title, subtitle_lines, n_cols):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=title)
        c.font = title_font(16, C['gold'])
        c.fill = fill(C['bg']); c.alignment = ctr()
        ws.row_dimensions[1].height = 36
        for col in range(2, n_cols+1): ws.cell(row=1, column=col).fill = fill(C['bg'])
        for i, sub in enumerate(subtitle_lines, 2):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=n_cols)
            c = ws.cell(row=i, column=1, value=sub)
            c.font = Font(name='Calibri', size=9, color='94A3B8', italic=True)
            c.fill = fill(C['bg']); c.alignment = ctr()
            ws.row_dimensions[i].height = 16
            for col in range(2, n_cols+1): ws.cell(row=i, column=col).fill = fill(C['bg'])
        next_row = len(subtitle_lines) + 2
        for col in range(1, n_cols+1):
            ws.cell(row=next_row, column=col).fill = fill('F8FAFC')
        ws.row_dimensions[next_row].height = 7
        return next_row + 1
    def stat_block(ws, start_row, donut_data, n_cols=8):
        total = sum(donut_data.values())
        boxes = [
            ('✓  PRESENT',  donut_data.get('present',  donut_data.get('Present',  0)), C['present'],  C['present_bg']),
            ('⏱  LATE',     donut_data.get('late',     donut_data.get('Late',     0)), C['late'],     C['late_bg']),
            ('✕  ABSENT',   donut_data.get('absent',   donut_data.get('Absent',   0)), C['absent'],   C['absent_bg']),
            ('◎  EXCUSED',  donut_data.get('excused',  donut_data.get('Excused',  0)), C['excused'],  C['excused_bg']),
        ]
        cols_per = max(1, n_cols // 4)
        for bi, (label, val, fg, bg2) in enumerate(boxes):
            sc = bi * cols_per + 1; ec = sc + cols_per - 1
            pct = f"{round(val/total*100,1)}%" if total else "0%"
            for row_offset, (text, sz, bold, height) in enumerate([
                (label, 9, True, 18), (val, 26, True, 40), (pct, 9, False, 18),
            ]):
                r = start_row + row_offset
                ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=ec)
                c = ws.cell(row=r, column=sc, value=text)
                c.font = Font(name='Calibri', size=sz, bold=bold, color=fg)
                c.fill = fill(bg2); c.alignment = ctr(); ws.row_dimensions[r].height = height
                for col in range(sc+1, ec+1): ws.cell(row=r, column=col).fill = fill(bg2)
        spacer_row = start_row + 3
        for col in range(1, n_cols+1):
            ws.cell(row=spacer_row, column=col).fill = fill('F8FAFC')
        ws.row_dimensions[spacer_row].height = 8
        return spacer_row + 1
    def totals_row(ws, row_num, values, n_cols):
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            c.fill = fill(C['sub_hdr']); c.border = thin_border()
            c.alignment = lft() if ci == 1 else ctr()
        ws.row_dimensions[row_num].height = 20
    def add_bar_chart(wb, ws_name, data_ws_name, title,
                      first_data_row, last_data_row, n_series,
                      cat_col, series_cols_start,
                      series_titles, series_colors,
                      chart_anchor, width=18, height=12):
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import SeriesLabel
        chart_ws = wb[ws_name]; data_ws = wb[data_ws_name]
        chart = BarChart()
        chart.type = 'col'; chart.grouping = 'stacked'; chart.overlap = 100
        chart.title = title; chart.style = 10
        chart.y_axis.title = 'Count'; chart.x_axis.title = ''
        chart.width = width; chart.height = height; chart.legend.position = 'b'
        cats = Reference(data_ws, min_col=cat_col, min_row=first_data_row, max_row=last_data_row)
        for i in range(n_series):
            col = series_cols_start + i
            data_ref = Reference(data_ws, min_col=col, min_row=first_data_row-1, max_row=last_data_row)
            chart.series.append(data_ref)
        chart.set_categories(cats)
        for i, (title_s, color) in enumerate(zip(series_titles, series_colors)):
            if i < len(chart.series):
                chart.series[i].title = SeriesLabel(v=title_s)
                chart.series[i].graphicalProperties.solidFill = color
                chart.series[i].graphicalProperties.line.solidFill = color
        chart_ws.add_chart(chart, chart_anchor)
    def add_pie_chart(wb, ws_name, data_ws_name, title,
                      first_data_row, last_data_row,
                      label_col, val_col,
                      series_colors, chart_anchor, width=12, height=10):
        from openpyxl.chart import PieChart, Reference
        chart_ws = wb[ws_name]; data_ws = wb[data_ws_name]
        chart = PieChart()
        chart.title = title; chart.style = 10; chart.width = width; chart.height = height
        labels = Reference(data_ws, min_col=label_col, min_row=first_data_row, max_row=last_data_row)
        data   = Reference(data_ws, min_col=val_col,   min_row=first_data_row, max_row=last_data_row)
        chart.add_data(data); chart.set_categories(labels)
        chart_ws.add_chart(chart, chart_anchor)
    return dict(
        C=C, fill=fill, thin_border=thin_border, ctr=ctr, lft=lft, rgt=rgt,
        header_font=header_font, normal_font=normal_font, title_font=title_font,
        make_header_row=make_header_row, data_row=data_row,
        title_block=title_block, stat_block=stat_block, totals_row=totals_row,
        add_bar_chart=add_bar_chart, add_pie_chart=add_pie_chart,
    )

# ── EXPORT ROUTES ─────────────────────────────────────────────────────────────

@app.route('/export/student_sessions/<nfc_id>')
@login_required
def export_student_sessions(nfc_id):
    """Export one student's full attendance history with blockchain proof."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import SeriesLabel

        f_status  = request.args.get('status','').strip()
        f_subject = request.args.get('subject','').strip()
        stud_name = request.args.get('name','Student').strip()
        now       = datetime.now()

        all_students = get_all_students()
        student      = next((x for x in all_students if x['nfcId']==nfc_id), None)

        with get_db() as conn:
            log_rows = conn.execute(
                "SELECT al.*, s.subject_name, s.course_code, s.section_key, "
                "s.teacher_name, s.time_slot, s.started_at, s.ended_at "
                "FROM attendance_logs al "
                "JOIN sessions s ON al.sess_id = s.sess_id "
                "WHERE al.nfc_id=? ORDER BY s.started_at DESC",
                (nfc_id,)
            ).fetchall()

        rows = []
        status_counts = {'Present':0,'Late':0,'Absent':0,'Excused':0}
        for lg in log_rows:
            status = lg['status'].capitalize()
            if f_status and status.lower() != f_status.lower(): continue
            if f_subject and lg['subject_name'] != f_subject: continue
            status_counts[status] = status_counts.get(status, 0) + 1
            rows.append({
                'subject':   lg['subject_name'],
                'code':      lg['course_code'] or '',
                'section':   (lg['section_key'] or '').replace('|',' · '),
                'teacher':   lg['teacher_name'] or '',
                'date':      (lg['started_at'] or '')[:10],
                'time_slot': lg['time_slot'] or '',
                'status':    status,
                'tap_time':  lg['tap_time'] or '—',
                'tx_hash':   lg['tx_hash']  or '—',
                'block':     str(lg['block_number']) if lg['block_number'] else '—',
                'excuse':    lg['excuse_note'] or '',
            })

        H = _xl_helpers(); C = H['C']
        wb = Workbook()
        ws = wb.active; ws.title = 'Attendance Log'
        prog  = student.get('course','') if student else ''
        yr    = student.get('year_level','') if student else ''
        sec   = student.get('section','') if student else ''
        sid_  = student.get('student_id','') if student else ''
        headers = ['#','Subject','Course Code','Section','Instructor',
                   'Date','Time Slot','Status','Tap Time','TX Hash','Block #','Excuse Note']
        widths  = [4, 32, 12, 24, 22, 14, 16, 10, 12, 52, 10, 22]
        subtitles = [
            'Cavite State University — DAVS Attendance Record',
            f'Student: {stud_name}  |  ID: {sid_}  |  NFC: {nfc_id}',
            f'Program: {prog}  |  Year: {yr}  |  Section: {sec}',
            f'Exported: {now.strftime("%B %d, %Y %I:%M %p")}',
        ]
        first_data = H['title_block'](ws, f'Student Attendance Report — {stud_name}', subtitles, len(headers))
        first_data = H['stat_block'](ws, first_data, status_counts, len(headers))
        H['make_header_row'](ws, first_data, headers, widths)
        first_data += 1
        col_fmt = {8: ('status',), 10: ('tx',), 11: ('num',)}
        for ri, row in enumerate(rows, first_data):
            H['data_row'](ws, ri, [
                ri - first_data + 1,
                row['subject'], row['code'], row['section'], row['teacher'],
                row['date'], row['time_slot'], row['status'],
                row['tap_time'], row['tx_hash'], row['block'], row['excuse'],
            ], alt=(ri % 2 == 0), col_formats=col_fmt)
        last_data = first_data + len(rows) - 1
        total_vals = ['TOTAL', '', '', '', '', '', '',
                      f"{status_counts['Present']}P / {status_counts['Late']}L / {status_counts['Absent']}A / {status_counts['Excused']}E",
                      '', '', '', '']
        H['totals_row'](ws, last_data + 1, total_vals, len(headers))
        ws.cell(row=last_data+3, column=1,
                value=f'Generated by DAVS on {now.strftime("%B %d, %Y %I:%M %p")}').font \
            = __import__('openpyxl').styles.Font(name='Calibri', size=9, italic=True, color='94A3B8')

        wc = wb.create_sheet('Charts')
        wc.sheet_view.showGridLines = False
        from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
        wc.merge_cells('A1:N1')
        wc['A1'] = f'Attendance Summary — {stud_name}'
        wc['A1'].font = XFont(name='Calibri', bold=True, size=14, color=C['gold'])
        wc['A1'].fill = XFill('solid', fgColor=C['bg'])
        wc['A1'].alignment = XAlign(horizontal='center', vertical='center')
        wc.row_dimensions[1].height = 32
        for col in range(2, 15): wc.cell(row=1, column=col).fill = XFill('solid', fgColor=C['bg'])
        wc.cell(row=3, column=1, value='Status').font = XFont(bold=True, size=9)
        wc.cell(row=3, column=2, value='Count').font  = XFont(bold=True, size=9)
        status_order = ['Present','Late','Absent','Excused']
        for ri, st in enumerate(status_order, 4):
            wc.cell(row=ri, column=1, value=st)
            wc.cell(row=ri, column=2, value=status_counts.get(st, 0))
        pie = PieChart()
        pie.title = 'Attendance Status Breakdown'
        pie.style = 10; pie.width = 14; pie.height = 10
        pie.add_data(Reference(wc, min_col=2, min_row=4, max_row=7))
        pie.set_categories(Reference(wc, min_col=1, min_row=4, max_row=7))
        wc.add_chart(pie, 'D3')
        subj_counts = {}
        for r in rows: subj_counts[r['subject']] = subj_counts.get(r['subject'], 0) + 1
        wc.cell(row=3, column=9,  value='Subject').font  = XFont(bold=True, size=9)
        wc.cell(row=3, column=10, value='Count').font    = XFont(bold=True, size=9)
        for ri2, (sn, cnt) in enumerate(sorted(subj_counts.items()), 4):
            wc.cell(row=ri2, column=9,  value=sn[:30])
            wc.cell(row=ri2, column=10, value=cnt)
        if subj_counts:
            bar = BarChart()
            bar.type = 'bar'; bar.grouping = 'clustered'
            bar.title = 'Sessions by Subject'; bar.style = 10; bar.width = 18; bar.height = 10
            bar.y_axis.title = 'Count'
            cats2 = Reference(wc, min_col=9,  min_row=4, max_row=3+len(subj_counts))
            data2 = Reference(wc, min_col=10, min_row=3, max_row=3+len(subj_counts))
            bar.add_data(data2, titles_from_data=True); bar.set_categories(cats2)
            if bar.series: bar.series[0].graphicalProperties.solidFill = C['accent']
            wc.add_chart(bar, 'D21')

        name_slug = stud_name.replace(' ','_')
        fname = (request.args.get('filename') or
                 f"{name_slug}_Attendance_Record_{now.strftime('%Y-%m-%d')}.xlsx")
        output = io.BytesIO()
        wb.save(output); output.seek(0)
        return Response(output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename="{fname}"'})
    except Exception:
        import traceback
        return Response(f'Export error: {traceback.format_exc()}', status=500, mimetype='text/plain')


@app.route('/export/session/<sess_id>')
@login_required
def export_session_attendance(sess_id):
    """Export one classroom session — attendance list with blockchain proof + charts."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
        import io as _io

        sess = load_session(sess_id)
        if not sess:
            flash('Session not found.'); return redirect(url_for('admin_sessions'))
        if not _is_my_session(sess):
            flash('Access denied.'); return redirect(url_for('teacher_sessions'))

        now          = datetime.now()
        section_key  = normalize_section_key(sess.get('section_key',''))
        all_students = get_all_students()
        enrolled     = sorted([s for s in all_students if build_student_section_key(s)==section_key],
                               key=lambda x: x['name'])
        present_ids  = set(sess.get('present', []))
        late_ids     = set(sess.get('late',    []))
        excused_ids  = set(sess.get('excused', []))
        att_logs     = {lg['nfc_id']: lg for lg in db_get_session_attendance(sess_id)}

        # Get excuse request details
        excuse_details = {}
        with get_db() as _conn:
            excuses = _conn.execute(
                "SELECT nfc_id, reason_type, reason_detail, attachment_file FROM excuse_requests WHERE sess_id=? AND status='approved'",
                (sess_id,)
            ).fetchall()
            for exc in excuses:
                excuse_details[exc['nfc_id']] = {
                    'reason': exc['reason_type'],
                    'reason_detail': exc['reason_detail'],
                    'attachment_file': exc['attachment_file'],
                }

        counts = {'Present':0,'Late':0,'Absent':0,'Excused':0}
        rows = []
        REASON_LABELS = {
            'sickness': 'Sickness / Illness', 'lbm': 'LBM', 'emergency': 'Family Emergency',
            'bereavement': 'Bereavement', 'medical': 'Medical Appointment', 'accident': 'Accident / Injury',
            'official': 'Official School Business', 'weather': 'Extreme Weather / Calamity',
            'transport': 'Transportation Problem', 'others': 'Others'
        }
        for st in enrolled:
            nid = st['nfcId']
            if   nid in excused_ids: status = 'Excused'
            elif nid in late_ids:    status = 'Late'
            elif nid in present_ids: status = 'Present'
            else:                    status = 'Absent'
            counts[status] += 1
            lg = att_logs.get(nid, {})
            excuse_info = excuse_details.get(nid, {})
            excuse_reason = ''
            if excuse_info.get('reason'):
                excuse_reason = REASON_LABELS.get(excuse_info['reason'], excuse_info['reason'])
                if excuse_info.get('reason_detail'):
                    excuse_reason += f" ({excuse_info['reason_detail']})"
            rows.append({
                'name':       st['name'],
                'student_id': st.get('student_id','—'),
                'nfc_id':     nid,
                'program':    st.get('course',''),
                'year':       st.get('year_level',''),
                'section':    st.get('section',''),
                'status':     status,
                'tap_time':   lg.get('tap_time','') or '—',
                'tx_hash':    lg.get('tx_hash','')  or '—',
                'block':      str(lg.get('block_number','')) if lg.get('block_number') else '—',
                'excuse_reason': excuse_reason,
                'excuse_document': excuse_info.get('attachment_file', '') or '',
            })

        H = _xl_helpers(); C = H['C']
        wb = Workbook()
        ws = wb.active; ws.title = 'Attendance'
        subj  = sess.get('subject_name','')
        code  = sess.get('course_code','')
        sec   = section_key.replace('|',' · ')
        instr = sess.get('teacher_name','')
        slot  = sess.get('time_slot','—')
        started = sess.get('started_at','—')
        ended   = sess.get('ended_at','Still running')
        n_cols  = 13
        subtitles = [
            'Cavite State University — DAVS Session Attendance Report',
            f'Subject: {subj}  {"["+code+"]" if code else ""}',
            f'Section: {sec}  |  Instructor: {instr}',
            f'Time Slot: {slot}  |  Started: {started}  |  Ended: {ended}',
            f'Exported: {now.strftime("%B %d, %Y %I:%M %p")}',
        ]
        first_data = H['title_block'](ws, 'Session Attendance Report', subtitles, n_cols)
        first_data = H['stat_block'](ws, first_data, counts, n_cols)
        headers = ['#','Student Name','Student ID','NFC Card UID',
                   'Program','Year','Sec','Status','Tap Time','TX Hash','Block #','Excuse Reason','Document']
        widths  = [4, 28, 14, 14, 28, 10, 6, 10, 12, 52, 10, 30, 20]
        H['make_header_row'](ws, first_data, headers, widths)
        first_data += 1
        col_fmt = {8: ('status',), 10: ('tx',), 11: ('num',)}
        for ri, row in enumerate(rows, first_data):
            H['data_row'](ws, ri, [
                ri-first_data+1, row['name'], row['student_id'], row['nfc_id'],
                row['program'], row['year'], row['section'],
                row['status'], row['tap_time'], row['tx_hash'], row['block'], row['excuse_reason'], row['excuse_document'],
            ], alt=(ri%2==0), col_formats=col_fmt)
        last_data = first_data + len(rows) - 1
        total_row_vals = ['TOTAL',f'{len(enrolled)} enrolled','','',
                          '','','',
                          f"{counts['Present']}P/{counts['Late']}L/{counts['Absent']}A/{counts['Excused']}E",
                          '','','','','']
        H['totals_row'](ws, last_data+1, total_row_vals, len(headers))
        ws.cell(row=last_data+3, column=1,
                value=f'Generated by DAVS on {now.strftime("%B %d, %Y %I:%M %p")}') \
            .font = XFont(name='Calibri', size=9, italic=True, color='94A3B8')

        wc = wb.create_sheet('Charts')
        wc.sheet_view.showGridLines = False
        wc.merge_cells('A1:N1')
        wc['A1'] = f'Attendance Charts — {subj} {"["+code+"]" if code else ""}'
        wc['A1'].font = XFont(name='Calibri', bold=True, size=14, color=C['gold'])
        wc['A1'].fill = XFill('solid', fgColor=C['bg'])
        wc['A1'].alignment = XAlign(horizontal='center', vertical='center')
        wc.row_dimensions[1].height = 32
        for col in range(2,15): wc.cell(row=1,column=col).fill = XFill('solid', fgColor=C['bg'])
        status_order = ['Present','Late','Absent','Excused']
        wc.cell(row=3,column=1,value='Status').font = XFont(bold=True, size=9, color=C['muted'])
        wc.cell(row=3,column=2,value='Count').font  = XFont(bold=True, size=9, color=C['muted'])
        for ri,st in enumerate(status_order,4):
            wc.cell(row=ri,column=1,value=st)
            wc.cell(row=ri,column=2,value=counts[st])
        pie = PieChart()
        pie.title = 'Attendance Status Breakdown'
        pie.style = 10; pie.width = 14; pie.height = 12
        pie.add_data(Reference(wc,min_col=2,min_row=4,max_row=7))
        pie.set_categories(Reference(wc,min_col=1,min_row=4,max_row=7))
        wc.add_chart(pie, 'D3')
        prog_counts = {}
        for r in rows:
            key = f"{r['year']}"; prog_counts[key] = prog_counts.get(key,{'Present':0,'Late':0,'Absent':0,'Excused':0})
            prog_counts[key][r['status']] += 1
        if len(prog_counts) > 1:
            r3c = 9
            wc.cell(row=3,column=r3c,  value='Year Level').font = XFont(bold=True,size=9,color=C['muted'])
            wc.cell(row=3,column=r3c+1,value='Present').font    = XFont(bold=True,size=9,color=C['muted'])
            wc.cell(row=3,column=r3c+2,value='Late').font       = XFont(bold=True,size=9,color=C['muted'])
            wc.cell(row=3,column=r3c+3,value='Absent').font     = XFont(bold=True,size=9,color=C['muted'])
            for ri3,(yr_k,yc) in enumerate(sorted(prog_counts.items()),4):
                wc.cell(row=ri3,column=r3c,  value=yr_k)
                wc.cell(row=ri3,column=r3c+1,value=yc['Present'])
                wc.cell(row=ri3,column=r3c+2,value=yc['Late'])
                wc.cell(row=ri3,column=r3c+3,value=yc['Absent'])
            bar = BarChart(); bar.type='col'; bar.grouping='stacked'; bar.overlap=100
            bar.title='Attendance by Year Level'; bar.style=10; bar.width=16; bar.height=12
            n_yl = len(prog_counts)
            bar.add_data(Reference(wc,min_col=r3c+1,min_row=3,max_row=3+n_yl),titles_from_data=True)
            bar.set_categories(Reference(wc,min_col=r3c,min_row=4,max_row=3+n_yl))
            for i,clr in enumerate([C['present'],C['late'],C['absent']]):
                if i < len(bar.series): bar.series[i].graphicalProperties.solidFill = clr
            wc.add_chart(bar,'D21')

        sec_last = section_key.split('|')[-1] if section_key else 'Sec'
        date_str = (started or '')[:10]
        code_part = f'_{code}' if code else ''
        fname = (request.args.get('filename') or
                 f"Session_Attendance{code_part}_{sec_last}_{date_str}.xlsx")
        output = _io.BytesIO()
        wb.save(output); output.seek(0)
        return Response(output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename="{fname}"'})
    except Exception:
        import traceback
        return Response(f'Export error: {traceback.format_exc()}', status=500, mimetype='text/plain')


@app.route('/export/stats.xlsx')
@app.route('/export/stats/xlsx', methods=['POST'])
@login_required
def export_stats_xlsx():
    """Unified analytics export — GET or POST. Produces rich multi-sheet workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
        import io as _io

        if request.method == 'POST':
            import base64
            from urllib.parse import parse_qs
            body       = request.get_json() or {}
            qs         = parse_qs(body.get('params',''))
            def qp(k): return qs.get(k,[''])[0]
            period     = qp('period') or 'all'
            f_section  = qp('section_key') or qp('section')
            f_year_lvl = qp('year_level')
            f_subject  = qp('subject')
            f_instr    = qp('instructor')
            f_month    = qp('month')
            f_year_num = qp('year_num')
            f_program  = qp('program')
            f_sec_ltr  = qp('section_letter')
            f_tod      = qp('time_of_day')
        else:
            period     = request.args.get('period',     'all')
            f_section  = request.args.get('section_key', request.args.get('section','')).strip()
            f_year_lvl = request.args.get('year_level', '').strip()
            f_subject  = request.args.get('subject',    '').strip()
            f_instr    = request.args.get('instructor', '').strip()
            f_month    = request.args.get('month',      '').strip()
            f_year_num = request.args.get('year_num',   '').strip()
            f_program  = request.args.get('program',    '').strip()
            f_sec_ltr  = request.args.get('section_letter','').strip()
            f_tod      = request.args.get('time_of_day','').strip()

        role     = session.get('role')
        username = session.get('username')
        now      = datetime.now()
        if not f_year_num:
            f_year_num = (qp('year') if request.method == 'POST' else request.args.get('year', '').strip()) if period in ('month', 'year') else ''

        if period == 'today':
            start_dt = now.replace(hour=0,minute=0,second=0,microsecond=0); end_dt = None
            period_label = now.strftime('Today — %B %d, %Y')
        elif period == 'month':
            yr = int(f_year_num) if f_year_num and f_year_num.isdigit() else now.year
            mo = int(f_month)    if f_month    and f_month.isdigit()    else now.month
            start_dt = datetime(yr,mo,1)
            end_dt   = datetime(yr,mo,_cal.monthrange(yr,mo)[1],23,59,59)
            period_label = start_dt.strftime('%B %Y')
        elif period == 'year':
            yr = int(f_year_num) if f_year_num and f_year_num.isdigit() else now.year
            start_dt = datetime(yr,1,1); end_dt = datetime(yr,12,31,23,59,59)
            period_label = str(yr)
        else:
            start_dt = datetime(2000,1,1); end_dt = None; period_label = 'All Time'

        all_sess = load_sessions()
        # Always use SQLite cache for export — get_all_students() may return
        # empty if blockchain restarted and events are gone. SQLite is the
        # reliable source of truth for enrolled students.
        all_stud = db_get_all_students()
        for _st in all_stud:
            _ov = db_get_override(_st['nfcId'])
            if _ov.get('course'):      _st['course']      = _ov['course']
            if _ov.get('year_level'):  _st['year_level']  = _ov['year_level']
            if _ov.get('section'):     _st['section']     = _ov['section'].upper()
            if _ov.get('full_name'):   _st['name']        = _ov['full_name']
            if _ov.get('student_id'):  _st['student_id']  = _ov['student_id']
            _st['section'] = (_st.get('section') or '').strip().upper()
        filtered = {}
        f_section_norm = normalize_section_key(f_section) if f_section else ''
        for sid, s in all_sess.items():
            if not s.get('started_at'): continue
            try:    sess_dt = datetime.strptime(s['started_at'],'%Y-%m-%d %H:%M:%S')
            except: continue
            if sess_dt < start_dt: continue
            if end_dt and sess_dt > end_dt: continue
            if role == 'teacher' and s.get('teacher') != username: continue
            sk = s.get('section_key',''); sk_parts = sk.split('|')
            if f_section_norm and normalize_section_key(sk) != f_section_norm: continue
            if f_program and (len(sk_parts)<1 or sk_parts[0] != f_program): continue
            if f_year_lvl and (len(sk_parts)<2 or sk_parts[1] != f_year_lvl): continue
            if f_sec_ltr and (len(sk_parts)<3 or sk_parts[2] != f_sec_ltr): continue
            if f_subject and s.get('subject_name','') != f_subject: continue
            if f_instr   and s.get('teacher_name','') != f_instr:   continue
            if f_tod:
                if ':' in f_tod:
                    if s.get('time_slot','') != f_tod: continue
                else:
                    try:
                        h = sess_dt.hour
                        if f_tod == 'morning'   and h >= 12: continue
                        if f_tod == 'afternoon' and h <  12: continue
                    except: pass
            filtered[sid] = s

        af = []
        if f_program:  af.append('Program: '+f_program)
        if f_year_lvl: af.append('Year: '+f_year_lvl)
        if f_sec_ltr:  af.append('Section: '+f_sec_ltr)
        if f_section:  af.append('Section: '+f_section.replace('|',' · '))
        if f_subject:  af.append('Subject: '+f_subject)
        if f_instr:    af.append('Instructor: '+f_instr)
        if f_tod:      af.append('Time: '+f_tod)
        filter_label = ' | '.join(af) if af else 'All data'

        donut  = {'Present':0,'Late':0,'Absent':0,'Excused':0}
        trend  = {}; subj_d = {}; sess_rows = []; detail_rows = []; by_section = {}
        REASON_LABELS = {
            'sickness': 'Sickness / Illness', 'lbm': 'LBM', 'emergency': 'Family Emergency',
            'bereavement': 'Bereavement', 'medical': 'Medical Appointment', 'accident': 'Accident / Injury',
            'official': 'Official School Business', 'weather': 'Extreme Weather / Calamity',
            'transport': 'Transportation Problem', 'others': 'Others'
        }

        for sid, s in sorted(filtered.items(), key=lambda x: x[1].get('started_at','')):
            sk       = normalize_section_key(s.get('section_key',''))
            enrolled = [st for st in all_stud if build_student_section_key(st)==sk]
            en_ids   = {st['nfcId'] for st in enrolled}
            pre      = set(s.get('present',[])); late = set(s.get('late',[]))
            exc      = set(s.get('excused',[])); abs_ = en_ids - pre - late - exc
            cnt      = {'enrolled':len(en_ids),
                        'present': len(pre-late), 'late': len(late),
                        'absent':  len(abs_),      'excused': len(exc)}
            for k in ('present','late','absent','excused'): donut[k.capitalize()] += cnt[k]
            code      = s.get('course_code','')
            subj_lbl  = f"[{code}] {s.get('subject_name','')}" if code else s.get('subject_name','')
            rate      = round((cnt['present']+cnt['late'])/cnt['enrolled']*100,1) if cnt['enrolled'] else 0
            date_key  = s['started_at'][:10]
            if date_key not in trend: trend[date_key] = {'present':0,'late':0,'absent':0,'excused':0}
            for k in ('present','late','absent','excused'): trend[date_key][k] += cnt[k]
            sn = s.get('subject_name','Unknown')
            if sn not in subj_d: subj_d[sn] = {'code':code,'present':0,'late':0,'absent':0,'excused':0}
            for k in ('present','late','absent','excused'): subj_d[sn][k] += cnt[k]
            if sk not in by_section: by_section[sk] = {'present':0,'late':0,'absent':0,'excused':0,'enrolled':0}
            for k in ('present','late','absent','excused','enrolled'): by_section[sk][k] += cnt[k]
            sess_rows.append([subj_lbl, fmt_time(s['started_at']),
                              sk.replace('|',' · '), s.get('teacher_name',''),
                              s.get('time_slot',''),
                              cnt['enrolled'], cnt['present'], cnt['late'],
                              cnt['absent'],   cnt['excused'], rate])
            att_logs = {lg['nfc_id']: lg for lg in db_get_session_attendance(sid)}
            # Get excuse request details for this session
            excuse_details = {}
            with get_db() as _conn:
                excuses = _conn.execute(
                    "SELECT nfc_id, reason_type, reason_detail, attachment_file FROM excuse_requests WHERE sess_id=? AND status='approved'",
                    (sid,)
                ).fetchall()
                for exc_row in excuses:
                    excuse_details[exc_row['nfc_id']] = {
                        'reason': exc_row['reason_type'],
                        'reason_detail': exc_row['reason_detail'],
                        'attachment_file': exc_row['attachment_file'],
                    }
            for st in enrolled:
                nid = st['nfcId']
                if   nid in exc:  status = 'Excused'
                elif nid in late: status = 'Late'
                elif nid in pre:  status = 'Present'
                else:             status = 'Absent'
                lg = att_logs.get(nid,{})
                excuse_info = excuse_details.get(nid, {})
                excuse_reason = ''
                if excuse_info.get('reason'):
                    excuse_reason = REASON_LABELS.get(excuse_info['reason'], excuse_info['reason'])
                    if excuse_info.get('reason_detail'):
                        excuse_reason += f" ({excuse_info['reason_detail']})"
                detail_rows.append([
                    st['name'], st.get('student_id',''), nid,
                    st.get('course',''), st.get('year_level',''), st.get('section',''),
                    subj_lbl, fmt_time(s['started_at']), s.get('time_slot',''),
                    s.get('teacher_name',''), status,
                    lg.get('tx_hash','') or s.get('tx_hashes',{}).get(nid,{}).get('tx_hash',''),
                    str(lg.get('block_number','') or s.get('tx_hashes',{}).get(nid,{}).get('block','')),
                    excuse_reason,
                    excuse_info.get('attachment_file', '') or '',
                ])

        total_all = sum(donut.values())
        H = _xl_helpers(); C = H['C']
        wb = Workbook()

        # Sheet 1: Summary
        ws1 = wb.active; ws1.title = 'Summary'
        n_cols = 11
        subtitles = [
            'Cavite State University — Decentralized Attendance Verification System',
            f'Period: {period_label}  |  Filters: {filter_label}',
            f'Generated: {now.strftime("%B %d, %Y  %I:%M %p")}  |  Role: {role.upper()}',
        ]
        first_row = H['title_block'](ws1, 'DAVS — Attendance Analytics Report', subtitles, n_cols)
        first_row = H['stat_block'](ws1, first_row, donut, n_cols)
        hdrs = ['Subject','Session Date & Time','Section','Instructor','Time Slot',
                'Enrolled','Present','Late','Absent','Excused','Rate %']
        wids = [36,22,26,22,16,10,10,9,9,10,10]
        H['make_header_row'](ws1, first_row, hdrs, wids)
        first_row += 1
        for ri, row in enumerate(sess_rows, first_row):
            vals = list(row); vals[10] = f"{vals[10]}%"
            col_fmt = {7:('num',),8:('num',),9:('num',),10:('num',)}
            H['data_row'](ws1, ri, vals, alt=(ri%2==0), col_formats=col_fmt)
        last_data = first_row + len(sess_rows) - 1
        tr = last_data + 1
        H['totals_row'](ws1, tr,
            ['TOTAL','','','','',
             f"=SUM(F{first_row}:F{last_data})",
             f"=SUM(G{first_row}:G{last_data})",
             f"=SUM(H{first_row}:H{last_data})",
             f"=SUM(I{first_row}:I{last_data})",
             f"=SUM(J{first_row}:J{last_data})",
             f'=IFERROR(TEXT((G{tr}+H{tr})/F{tr},"0.0%"),"-")'],
            n_cols)
        ws1.cell(row=tr+2, column=1,
                 value=f'Generated by DAVS on {now.strftime("%B %d, %Y %I:%M %p")}') \
            .font = XFont(name='Calibri', size=9, italic=True, color='94A3B8')
        ws1.freeze_panes = ws1.cell(row=first_row, column=1)

        # Sheet 2: Student Detail
        ws2 = wb.create_sheet('Student Detail'); n2 = 16
        subtitles2 = [
            'Cavite State University — DAVS',
            f'Period: {period_label}  |  Filters: {filter_label}',
            'Each row represents one student in one session including full blockchain proof.',
        ]
        dr = H['title_block'](ws2, 'Student Attendance Detail', subtitles2, n2)
        det_hdrs = ['Student Name','Student ID','NFC Card UID','Program','Year','Sec',
                    'Subject','Session Date','Time Slot','Instructor',
                    'Status','TX Hash','Block #','Excuse Reason','Document']
        det_wids = [28,14,14,28,10,6,32,20,16,22,10,52,10,26,20]
        H['make_header_row'](ws2, dr, det_hdrs, det_wids)
        dr += 1
        for ri, row in enumerate(detail_rows, dr):
            col_fmt = {11:('status',),12:('tx',),13:('num',)}
            H['data_row'](ws2, ri, row, alt=(ri%2==0), col_formats=col_fmt)
        ws2.freeze_panes = ws2.cell(row=dr, column=1)

        # Sheet 3: By Date
        ws3 = wb.create_sheet('By Date'); n3 = 5
        subtitles3 = [f'Period: {period_label}  |  Attendance counts per session date']
        tr3 = H['title_block'](ws3, 'Attendance Trend by Date', subtitles3, n3)
        H['make_header_row'](ws3, tr3, ['Date','Present','Late','Absent','Excused'], [18,12,12,12,12])
        tr3 += 1
        for ri, (date, td) in enumerate(sorted(trend.items()), tr3):
            col_fmt = {2:('num',),3:('num',),4:('num',),5:('num',)}
            H['data_row'](ws3, ri, [date,td['present'],td['late'],td['absent'],td['excused']],
                          alt=(ri%2==0), col_formats=col_fmt)
        last_tr3 = tr3 + len(trend) - 1
        if trend:
            bar3 = BarChart(); bar3.type='col'; bar3.grouping='stacked'; bar3.overlap=100
            bar3.title='Daily Attendance Trend'; bar3.style=10; bar3.width=22; bar3.height=14
            bar3.y_axis.title='Count'; bar3.legend.position='b'
            cats3 = Reference(ws3, min_col=1, min_row=tr3, max_row=last_tr3)
            data3 = Reference(ws3, min_col=2, min_row=tr3-1, max_row=last_tr3, max_col=5)
            bar3.add_data(data3, titles_from_data=True); bar3.set_categories(cats3)
            colors3 = [C['present'],C['late'],C['absent'],C['excused']]
            for i,clr in enumerate(colors3):
                if i < len(bar3.series):
                    bar3.series[i].graphicalProperties.solidFill = clr
                    bar3.series[i].graphicalProperties.line.solidFill = clr
            ws3.add_chart(bar3, f'G{tr3}')

        # Sheet 4: By Subject
        ws4 = wb.create_sheet('By Subject'); n4 = 7
        subtitles4 = [f'Period: {period_label}  |  Aggregate attendance per subject']
        ts4 = H['title_block'](ws4, 'Attendance by Subject', subtitles4, n4)
        H['make_header_row'](ws4, ts4,
                             ['Subject','Code','Present','Late','Absent','Excused','Rate %'],
                             [36,12,12,12,12,12,12])
        ts4 += 1
        for ri, (sn, sd) in enumerate(sorted(subj_d.items()), ts4):
            total2 = sd['present']+sd['late']+sd['absent']+sd['excused']
            rate2  = f"{round((sd['present']+sd['late'])/total2*100,1)}%" if total2 else '—'
            col_fmt4 = {3:('num',),4:('num',),5:('num',),6:('num',)}
            H['data_row'](ws4, ri, [sn, sd['code'], sd['present'],sd['late'],sd['absent'],sd['excused'],rate2],
                          alt=(ri%2==0), col_formats=col_fmt4)
        last_ts4 = ts4 + len(subj_d) - 1
        if subj_d:
            bar4 = BarChart(); bar4.type='bar'; bar4.grouping='stacked'; bar4.overlap=100
            bar4.title='Attendance by Subject'; bar4.style=10; bar4.width=22; bar4.height=14
            bar4.x_axis.title='Count'; bar4.legend.position='b'
            cats4 = Reference(ws4, min_col=1, min_row=ts4, max_row=last_ts4)
            data4 = Reference(ws4, min_col=3, min_row=ts4-1, max_row=last_ts4, max_col=6)
            bar4.add_data(data4, titles_from_data=True); bar4.set_categories(cats4)
            colors4 = [C['present'],C['late'],C['absent'],C['excused']]
            for i,clr in enumerate(colors4):
                if i < len(bar4.series): bar4.series[i].graphicalProperties.solidFill = clr
            ws4.add_chart(bar4, f'I{ts4}')

# Sheet 5: By Section
        ws5 = wb.create_sheet('By Section'); n5 = 7
        subtitles5 = [f'Period: {period_label}  |  Aggregate attendance per section']
        ts5 = H['title_block'](ws5, 'Attendance by Section', subtitles5, n5)
        H['make_header_row'](ws5, ts5,
                             ['Section','Enrolled','Present','Late','Absent','Excused','Rate %'],
                             [32,10,12,12,12,12,12])
        ts5 += 1
        for ri, (sec_k, sc) in enumerate(sorted(by_section.items()), ts5):
            total5 = sc['present']+sc['late']+sc['absent']+sc['excused']
            rate5  = f"{round((sc['present']+sc['late'])/sc['enrolled']*100,1)}%" if sc['enrolled'] else '-'
            col_fmt5 = {2:('num',),3:('num',),4:('num',),5:('num',),6:('num',)}
            H['data_row'](ws5, ri, [sec_k.replace('|',' · '), sc['enrolled'],
                                    sc['present'],sc['late'],sc['absent'],sc['excused'],rate5],
                          alt=(ri%2==0), col_formats=col_fmt5)
        last_ts5 = ts5 + len(by_section) - 1
        # Add bar chart for By Section sheet
        if by_section:
            from openpyxl.chart import BarChart, Reference
            bar5 = BarChart(); bar5.type='bar'; bar5.grouping='stacked'; bar5.overlap=100
            bar5.title='Attendance by Section'; bar5.style=10; bar5.width=22; bar5.height=14
            bar5.x_axis.title='Count'; bar5.legend.position='b'
            cats5 = Reference(ws5, min_col=1, min_row=ts5, max_row=last_ts5)
            data5 = Reference(ws5, min_col=3, min_row=ts5-1, max_row=last_ts5, max_col=6)
            bar5.add_data(data5, titles_from_data=True); bar5.set_categories(cats5)
            colors5 = [C['present'],C['late'],C['absent'],C['excused']]
            for i,clr in enumerate(colors5):
                if i < len(bar5.series):
                    bar5.series[i].graphicalProperties.solidFill = clr
                    bar5.series[i].graphicalProperties.line.solidFill = clr
            ws5.add_chart(bar5, f'I{ts5}')
 
        # Sheet 6: Charts Dashboard
        # ── Layout strategy ────────────────────────────────────────────────
        # Row 1-2:  Title + subtitle (dark header)
        # Row 3:    Spacer
        # Row 4-7:  PIE chart data (cols A-B), Pie chart anchored at C4
        # Row 10+:  Subject bar data (cols A-E), Subject bar anchored at G10
        # Row 25+:  Section bar data (cols A-E), Section bar anchored at G25
        # All data columns are hidden so only charts show.
        # ──────────────────────────────────────────────────────────────────
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
 
        wc = wb.create_sheet('Charts')
        wc.sheet_view.showGridLines = False
 
        # Title rows
        n_chart_cols = 20
        wc.merge_cells(f'A1:{chr(64+n_chart_cols)}1')
        wc['A1'] = 'DAVS — Attendance Analytics Charts'
        wc['A1'].font      = XFont(name='Calibri', bold=True, size=16, color=C['gold'])
        wc['A1'].fill      = XFill('solid', fgColor=C['bg'])
        wc['A1'].alignment = XAlign(horizontal='center', vertical='center')
        wc.row_dimensions[1].height = 36
        for col in range(2, n_chart_cols+1):
            wc.cell(row=1, column=col).fill = XFill('solid', fgColor=C['bg'])
 
        wc.merge_cells(f'A2:{chr(64+n_chart_cols)}2')
        wc['A2'] = f'Period: {period_label}  |  Filters: {filter_label}  |  Generated: {now.strftime("%B %d, %Y")}'
        wc['A2'].font      = XFont(name='Calibri', size=9, italic=True, color='94A3B8')
        wc['A2'].fill      = XFill('solid', fgColor=C['bg'])
        wc['A2'].alignment = XAlign(horizontal='center', vertical='center')
        for col in range(2, n_chart_cols+1):
            wc.cell(row=2, column=col).fill = XFill('solid', fgColor=C['bg'])
 
        # ── 1. Overall Status Pie chart ───────────────────────────────────
        # Data in cols A-B, rows 4-7 (hidden)
        pie_labels = ['Present','Late','Absent','Excused']
        pie_vals   = [donut[k] for k in pie_labels]
        for ri_p, (lbl, val) in enumerate(zip(pie_labels, pie_vals), 4):
            wc.cell(row=ri_p, column=30, value=lbl)
            wc.cell(row=ri_p, column=31, value=val)
 
        pie_c = PieChart()
        pie_c.title  = f'Overall Attendance Status — {period_label}'
        pie_c.style  = 10; pie_c.width = 16; pie_c.height = 14
        pie_c.add_data(Reference(wc, min_col=31, min_row=4, max_row=7))
        pie_c.set_categories(Reference(wc, min_col=30, min_row=4, max_row=7))
        pie_colors = [C['present'], C['late'], C['absent'], C['excused']]
        for i, clr in enumerate(pie_colors):
            if i < len(pie_c.series) and pie_c.series[i].dPt:
                pass  # openpyxl PieChart slice colors require DataPoint — skip per-slice color
        wc.add_chart(pie_c, 'B4')
 
        # ── 2. Attendance by Subject bar chart ────────────────────────────
        # Data in cols A-E, rows 10+ (cols A-B already hidden, C-F hidden below)
        SUBJ_DATA_ROW = 10
        if subj_d:
            wc.cell(row=SUBJ_DATA_ROW-1, column=33, value='Subject')
            wc.cell(row=SUBJ_DATA_ROW-1, column=34, value='Present')
            wc.cell(row=SUBJ_DATA_ROW-1, column=35, value='Late')
            wc.cell(row=SUBJ_DATA_ROW-1, column=36, value='Absent')
            wc.cell(row=SUBJ_DATA_ROW-1, column=37, value='Excused')
            for ri_s, (sn, sd) in enumerate(sorted(subj_d.items()), SUBJ_DATA_ROW):
                wc.cell(row=ri_s, column=33, value=sn[:30])
                wc.cell(row=ri_s, column=34, value=sd['present'])
                wc.cell(row=ri_s, column=35, value=sd['late'])
                wc.cell(row=ri_s, column=36, value=sd['absent'])
                wc.cell(row=ri_s, column=37, value=sd['excused'])
            n_subj      = len(subj_d)
            subj_last   = SUBJ_DATA_ROW + n_subj - 1
            bar_s = BarChart(); bar_s.type='bar'; bar_s.grouping='stacked'; bar_s.overlap=100
            bar_s.title  = 'Attendance by Subject'
            bar_s.style  = 10; bar_s.width = 20; bar_s.height = max(10, n_subj * 1.2)
            bar_s.legend.position = 'b'
            bar_s.add_data(
                Reference(wc, min_col=34, min_row=SUBJ_DATA_ROW-1,
                            max_row=subj_last, max_col=37),
                titles_from_data=True)
            bar_s.set_categories(
                Reference(wc, min_col=33, min_row=SUBJ_DATA_ROW, max_row=subj_last))
            for i, clr in enumerate([C['present'],C['late'],C['absent'],C['excused']]):
                if i < len(bar_s.series):
                    bar_s.series[i].graphicalProperties.solidFill = clr
                    bar_s.series[i].graphicalProperties.line.solidFill = clr
            # Anchor subject chart below the pie chart — row 4 + ~20 rows down
            wc.add_chart(bar_s, 'B36')

        # ── 3. Attendance by Section bar chart ───────────────────────────
        # Data in cols H-L, rows 10+
        SEC_DATA_ROW = 10
        if by_section:
            wc.cell(row=SEC_DATA_ROW-1, column=38, value='Section')
            wc.cell(row=SEC_DATA_ROW-1, column=39, value='Present')
            wc.cell(row=SEC_DATA_ROW-1, column=40, value='Late')
            wc.cell(row=SEC_DATA_ROW-1, column=41, value='Absent')
            wc.cell(row=SEC_DATA_ROW-1, column=42, value='Excused')
            for ri_sc, (sec_k, sc) in enumerate(sorted(by_section.items()), SEC_DATA_ROW):
                wc.cell(row=ri_sc, column=38, value=sec_k.replace('|',' · ')[:30])
                wc.cell(row=ri_sc, column=39, value=sc['present'])
                wc.cell(row=ri_sc, column=40, value=sc['late'])
                wc.cell(row=ri_sc, column=41, value=sc['absent'])
                wc.cell(row=ri_sc, column=42, value=sc['excused'])
            n_sec     = len(by_section)
            sec_last  = SEC_DATA_ROW + n_sec - 1
            bar_sc = BarChart(); bar_sc.type='bar'; bar_sc.grouping='stacked'; bar_sc.overlap=100
            bar_sc.title  = 'Attendance by Section'
            bar_sc.style  = 10; bar_sc.width = 20; bar_sc.height = max(10, n_sec * 1.4)
            bar_sc.legend.position = 'b'
            bar_sc.add_data(
                Reference(wc, min_col=39, min_row=SEC_DATA_ROW-1,
                            max_row=sec_last, max_col=42),
                titles_from_data=True)
            bar_sc.set_categories(
                Reference(wc, min_col=38, min_row=SEC_DATA_ROW, max_row=sec_last))
            for i, clr in enumerate([C['present'],C['late'],C['absent'],C['excused']]):
                if i < len(bar_sc.series):
                    bar_sc.series[i].graphicalProperties.solidFill = clr
                    bar_sc.series[i].graphicalProperties.line.solidFill = clr
            # Anchor section chart to the right of the pie chart
            wc.add_chart(bar_sc, 'N4')

        parts = ['DAVS_Attendance_Report', period_label.replace(' ','_').replace(',','')]
        if f_program:  parts.append(f_program.replace('BS ','BS').replace(' ','_'))
        if f_year_lvl: parts.append(f_year_lvl.replace(' ','_'))
        if f_sec_ltr:  parts.append(f'Section_{f_sec_ltr}')
        if f_subject:  parts.append(re.sub(r'[^A-Za-z0-9]','_',f_subject)[:20])
        if f_instr:    parts.append(f_instr.split()[0])
        fname = request.args.get('filename') or ('_'.join(parts)+f'_{now.strftime("%Y-%m-%d")}.xlsx')
        fname = re.sub(r'_+','_', fname)
        # Strip characters outside latin-1 range — the em dash (—) in
        # period_label e.g. "Today — March 21, 2026" causes UnicodeEncodeError
        # when Werkzeug encodes the Content-Disposition HTTP header.
        fname = fname.replace('\u2014','-').replace('\u2013','-')  # em/en dash
        fname = fname.encode('ascii','ignore').decode('ascii')      # drop rest
        fname = re.sub(r'[^\w\-.]','_', fname)
        fname = re.sub(r'_+','_', fname).strip('_')
        if not fname.endswith('.xlsx'):
            fname += '.xlsx'
        output = _io.BytesIO()
        wb.save(output); output.seek(0)
        return Response(output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'})
    except Exception:
        import traceback
        return Response(f'Export error: {traceback.format_exc()}', status=500, mimetype='text/plain')


@app.route('/export/stats.csv')
@login_required
def export_stats_csv():
    period     = request.args.get('period', 'all')
    f_month    = request.args.get('month', '').strip()
    f_year_num = request.args.get('year_num', '').strip()
    f_subject  = request.args.get('subject', '').strip()
    f_section  = request.args.get('section_key', request.args.get('section', '')).strip()
    f_year_lvl = request.args.get('year_level', '').strip()
    f_program  = request.args.get('program', '').strip()
    f_sec_ltr  = request.args.get('section_letter', '').strip()
    f_instr    = request.args.get('instructor', '').strip()
    f_tod      = request.args.get('time_of_day', '').strip()
    role       = session.get('role')
    username   = session.get('username')
    now        = datetime.now()

    if not f_year_num:
        f_year_num = request.args.get('year', '').strip() if period in ('month', 'year') else ''

    if period == 'today':
        start_dt = now.replace(hour=0, minute=0, second=0, microsecond=0); end_dt = None
    elif period == 'month':
        yr = int(f_year_num) if f_year_num and f_year_num.isdigit() else now.year
        mo = int(f_month) if f_month and f_month.isdigit() else now.month
        start_dt = datetime(yr, mo, 1); end_dt = datetime(yr, mo, _cal.monthrange(yr, mo)[1], 23, 59, 59)
    elif period == 'year':
        yr = int(f_year_num) if f_year_num and f_year_num.isdigit() else now.year
        start_dt = datetime(yr, 1, 1); end_dt = datetime(yr, 12, 31, 23, 59, 59)
    else:
        start_dt = datetime(2000, 1, 1); end_dt = None

    where = ["s.started_at >= ?"]
    params = [start_dt.strftime('%Y-%m-%d %H:%M:%S')]
    if end_dt:
        where.append("s.started_at <= ?")
        params.append(end_dt.strftime('%Y-%m-%d %H:%M:%S'))
    if role == 'teacher':
        where.append("s.teacher_username = ?")
        params.append(username)
    if f_section:
        where.append("s.section_key = ?")
        params.append(normalize_section_key(f_section))
    if f_program:
        where.append("s.section_key LIKE ?")
        params.append(f_program + '%')
    if f_year_lvl:
        where.append("s.section_key LIKE ?")
        params.append('%|' + f_year_lvl + '|%')
    if f_sec_ltr:
        where.append("s.section_key LIKE ?")
        params.append('%|' + f_sec_ltr)
    if f_subject:
        where.append("s.subject_name = ?")
        params.append(f_subject)
    if f_instr:
        where.append("s.teacher_name = ?")
        params.append(f_instr)
    if f_tod == 'morning':
        where.append("CAST(strftime('%H',s.started_at) AS INTEGER) < 12")
    elif f_tod == 'afternoon':
        where.append("CAST(strftime('%H',s.started_at) AS INTEGER) >= 12")
    elif f_tod and ':' in f_tod:
        where.append("s.time_slot = ?")
        params.append(f_tod)
    wsql = " AND ".join(where)

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Session ID', 'Subject', 'Section', 'Instructor', 'Date', 'Time Slot', 'Total Records', 'Present', 'Late', 'Absent', 'Excused', 'Rate%'])
    with get_db() as conn:
        rows = conn.execute(
            "SELECT s.sess_id, s.subject_name, s.section_key, s.teacher_name, s.started_at, s.time_slot, "
            "SUM(CASE WHEN al.status='present' THEN 1 ELSE 0 END) AS present_count, "
            "SUM(CASE WHEN al.status='late' THEN 1 ELSE 0 END) AS late_count, "
            "SUM(CASE WHEN al.status='absent' THEN 1 ELSE 0 END) AS absent_count, "
            "SUM(CASE WHEN al.status='excused' THEN 1 ELSE 0 END) AS excused_count, "
            "COUNT(*) AS total_count "
            "FROM attendance_logs al "
            "JOIN sessions s ON al.sess_id = s.sess_id "
            "WHERE " + wsql + " "
            "GROUP BY s.sess_id, s.subject_name, s.section_key, s.teacher_name, s.started_at, s.time_slot "
            "ORDER BY s.started_at",
            params
        ).fetchall()
    for r in rows:
        total = int(r['total_count'] or 0)
        present = int(r['present_count'] or 0)
        late = int(r['late_count'] or 0)
        absent = int(r['absent_count'] or 0)
        excused = int(r['excused_count'] or 0)
        rate = round((present + late) / total * 100, 1) if total else 0
        w.writerow([
            (r['sess_id'] or '')[:8],
            r['subject_name'] or '',
            normalize_section_key(r['section_key'] or '').replace('|', ' · '),
            r['teacher_name'] or '',
            (r['started_at'] or '')[:10],
            r['time_slot'] or '',
            total, present, late, absent, excused, rate
        ])
    out.seek(0)
    fname = f"attendance_{period}_{now.strftime('%Y%m%d')}.csv"
    return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment;filename={fname}'})

@app.route('/request_registration_scan', methods=['POST'])
@admin_required
def request_registration_scan():
    try:
        nfc_set_waiting(session.get('username','admin'))
        return jsonify({'status':'ready'})
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}), 500

@app.route('/get_scanned_uid')
@login_required
def get_scanned_uid():
    uid = nfc_get_uid()
    if uid:
        return jsonify({'uid': uid})
    return jsonify({'uid': None})

@app.route('/registration_status')
def registration_status():
    return jsonify({'waiting': nfc_is_waiting()})

@app.route('/receive_pico_uid', methods=['POST'])
def receive_pico_uid():
    data=request.get_json()
    if not data or 'uid' not in data: return jsonify({'status':'error'}), 400
    uid=data['uid'].strip().upper()
    nfc_set_uid(uid)
    return jsonify({'status':'ok','uid':uid})


# ══════════════════════════════════════════════════════════════════════════════
# SUPER ADMIN ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/superadmin/users')
@super_admin_required
def superadmin_users():
    users = db_get_all_users()
    super_admin_count = sum(1 for u in users.values() if u.get('role') == 'super_admin')
    return render_template('superadmin_users.html',
                           users=users,
                           ADMIN_ROLES=ADMIN_ROLES,
                           super_admin_count=super_admin_count)

@app.route('/superadmin/create-user', methods=['GET', 'POST'])
@super_admin_required
def superadmin_create_user():
    if request.method == 'POST':
        username  = request.form.get('username', '').strip().lower()
        fullname  = request.form.get('full_name', '').strip()
        email     = request.form.get('email', '').strip()
        role      = request.form.get('role', 'teacher')
        password  = request.form.get('password', '').strip()
        if not username or not fullname or not password:
            flash('Username, full name, and password are required.', 'danger')
            return redirect(url_for('superadmin_create_user'))
        if role not in {'teacher', 'admin', 'super_admin'}:
            flash('Invalid role.', 'danger')
            return redirect(url_for('superadmin_create_user'))
        # Prevent duplicate Super Admin
        if role == 'super_admin':
            with get_db() as conn:
                existing = conn.execute(
                    "SELECT * FROM users WHERE role='super_admin'"
                ).fetchone()
            if existing:
                flash('A Super Admin already exists. Only one Super Admin is allowed.', 'danger')
                return redirect(url_for('superadmin_create_user'))
        if db_get_user(username):
            flash(f'Username "{username}" is already taken.', 'danger')
            return redirect(url_for('superadmin_create_user'))
        if len(password) < 8:
            flash('Password must be at least 8 characters.', 'danger')
            return redirect(url_for('superadmin_create_user'))
        db_save_user(username, {
            'username': username, 'password': hash_password(password),
            'role': role, 'full_name': fullname, 'email': email,
            'status': 'approved', 'sections': [], 'my_subjects': [],
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        flash(f'Account "{username}" ({role}) created successfully.', 'success')
        return redirect(url_for('superadmin_users'))
    # Check if Super Admin exists for frontend
    with get_db() as conn:
        super_admin_exists = conn.execute(
            "SELECT COUNT(*) as cnt FROM users WHERE role='super_admin'"
        ).fetchone()['cnt'] > 0
    return render_template('superadmin_create_user.html', show_super_admin=not super_admin_exists)

@app.route('/superadmin/promote/<username>', methods=['POST'])
@super_admin_required
def superadmin_promote(username):
    u = db_get_user(username)
    if not u:
        flash('User not found.', 'danger')
        return redirect(url_for('superadmin_users'))
    new_role = request.form.get('role', 'teacher')
    if new_role not in {'teacher', 'admin', 'super_admin'}:
        flash('Invalid role.', 'danger')
        return redirect(url_for('superadmin_users'))
    # Prevent creating a second super_admin via promote
    if new_role == 'super_admin' and u.get('role') != 'super_admin':
        with get_db() as conn:
            existing = conn.execute(
                "SELECT COUNT(*) as cnt FROM users WHERE role='super_admin'"
            ).fetchone()['cnt']
        if existing >= 1:
            flash('A Super Admin already exists. Only one Super Admin is allowed.', 'danger')
            return redirect(url_for('superadmin_users'))
    u['role'] = new_role
    u['password'] = u.get('password', '')
    db_save_user(username, u)
    flash(f'"{username}" role changed to {new_role}.', 'success')
    return redirect(url_for('superadmin_users'))

# ── Admin: create instructor (admin can create teachers, not other admins) ──

@app.route('/admin/create-instructor', methods=['GET', 'POST'])
@admin_required
def admin_create_instructor():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        fullname = request.form.get('full_name', '').strip()
        email    = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not fullname or not password:
            flash('Username, full name, and password are required.', 'danger')
            return redirect(url_for('admin_create_instructor'))
        if db_get_user(username):
            flash(f'Username "{username}" is already taken.', 'danger')
            return redirect(url_for('admin_create_instructor'))
        if len(password) < 8:
            flash('Password must be at least 8 characters.', 'danger')
            return redirect(url_for('admin_create_instructor'))
        db_save_user(username, {
            'username': username, 'password': hash_password(password),
            'role': 'teacher', 'full_name': fullname, 'email': email,
            'status': 'approved', 'sections': [], 'my_subjects': [],
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        flash(f'Instructor account "{username}" created successfully.', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin_create_instructor.html')

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN SCHEDULING ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/schedules')
@admin_required
def admin_schedules():
    all_schedules = db_get_all_schedules()
    subjects   = db_get_all_subjects()
    users      = db_get_all_users()
    # Only show teacher-role accounts in the instructor dropdown
    teachers   = {u: v for u, v in users.items() if v.get('role') == 'teacher' and v.get('status') == 'approved'}
    sections   = _get_all_section_keys()
    # Access control: normal admin only sees schedules they are assigned to
    current_role = session.get('role', '')
    current_username = session.get('username', '')
    if current_role == 'super_admin':
        schedules = all_schedules
    else:
        schedules = [s for s in all_schedules if s.get('teacher_username') == current_username]
    return render_template('admin_schedules.html',
                           schedules=schedules,
                           subjects=subjects,
                           teachers=teachers,
                           sections=sections,
                           dow_names=DOW_NAMES,
                           is_super_admin=(current_role == 'super_admin'))

@app.route('/admin/schedules/create', methods=['POST'])
@admin_required
def admin_schedule_create():
    try:
        subject_id  = request.form.get('subject_id', '')
        subj        = db_get_subject(subject_id)
        if not subj:
            flash('Subject not found.', 'danger')
            return redirect(url_for('admin_schedules'))
        teacher_u = request.form.get('teacher_username', '')
        teacher   = db_get_user(teacher_u)
        if not teacher:
            flash('Instructor not found.', 'danger')
            return redirect(url_for('admin_schedules'))
        program      = request.form.get('program', '').strip()
        year_level   = request.form.get('year_level', '').strip()
        # Support both 'section' (new form) and 'section_letter' (old form)
        sec_letter   = (request.form.get('section', '') or request.form.get('section_letter', '')).strip()
        if program and year_level and sec_letter:
            section_key = normalize_section_key(f"{program}|{year_level}|{sec_letter}")
        else:
            section_key  = normalize_section_key(request.form.get('section_key', ''))
            
        day_of_week  = int(request.form.get('day_of_week', 0))
        start_time   = request.form.get('start_time', '')
        end_time     = request.form.get('end_time', '')
        grace_min    = int(request.form.get('grace_minutes', 15))
        if not section_key or not start_time or not end_time:
            flash('Section, start time, and end time are required.', 'danger')
            return redirect(url_for('admin_schedules'))
        start_mins = _time_mins(start_time)
        end_mins = _time_mins(end_time)
        if start_mins is None or end_mins is None or end_mins <= start_mins:
            flash('End time must be later than start time.', 'danger')
            return redirect(url_for('admin_schedules'))
        db_save_schedule({
            'section_key': section_key,
            'subject_id':  subject_id,
            'subject_name': subj['name'],
            'course_code': subj.get('course_code', ''),
            'teacher_username': teacher_u,
            'teacher_name': teacher.get('full_name', teacher_u),
            'day_of_week': day_of_week,
            'start_time':  start_time,
            'end_time':    end_time,
            'grace_minutes': grace_min,
            'created_by':  session.get('username', ''),
        })
        flash('Schedule created successfully.', 'success')
    except Exception as e:
        flash(f'Error creating schedule: {e}', 'danger')
    return redirect(url_for('admin_schedules'))

@app.route('/admin/schedules/<schedule_id>/edit', methods=['POST'])
@admin_required
def admin_schedule_edit(schedule_id):
    try:
        sched = db_get_schedule(schedule_id)
        if not sched:
            flash('Schedule not found.', 'danger')
            return redirect(url_for('admin_schedules'))
        
        # ✅ FIX: Check if schedule is within 5 minutes of start time
        today_dow = datetime.now().weekday()
        if int(sched['day_of_week']) == today_dow:
            current_time = datetime.now().strftime('%H:%M')
            start_time = sched['start_time']
            try:
                now_dt = datetime.strptime(current_time, '%H:%M')
                start_dt = datetime.strptime(start_time, '%H:%M')
                time_diff = (start_dt - now_dt).total_seconds() / 60
                if 0 <= time_diff <= 5:
                    flash('Cannot edit schedule within 5 minutes before start time. Delete and recreate if changes are needed.', 'warning')
                    return redirect(url_for('admin_schedules'))
            except:
                pass
        
        subject_id  = request.form.get('subject_id', sched['subject_id'])
        subj        = db_get_subject(subject_id)
        teacher_u   = request.form.get('teacher_username', sched['teacher_username'])
        teacher     = db_get_user(teacher_u)
        program      = request.form.get('program', '').strip()
        year_level   = request.form.get('year_level', '').strip()
        sec_letter   = (request.form.get('section', '') or request.form.get('section_letter', '')).strip()
        if program and year_level and sec_letter:
            new_section = normalize_section_key(f"{program}|{year_level}|{sec_letter}")
        else:
            new_section = normalize_section_key(request.form.get('section_key', sched['section_key']))
        new_start_time = request.form.get('start_time', sched['start_time'])
        new_end_time = request.form.get('end_time', sched['end_time'])
        start_mins = _time_mins(new_start_time)
        end_mins = _time_mins(new_end_time)
        if start_mins is None or end_mins is None or end_mins <= start_mins:
            flash('End time must be later than start time.', 'danger')
            return redirect(url_for('admin_schedules'))
            
        db_save_schedule({
            'schedule_id': schedule_id,
            'section_key': new_section,
            'subject_id':  subject_id,
            'subject_name': subj['name'] if subj else sched['subject_name'],
            'course_code': subj.get('course_code','') if subj else sched['course_code'],
            'teacher_username': teacher_u,
            'teacher_name': teacher.get('full_name', teacher_u) if teacher else sched['teacher_name'],
            'day_of_week': int(request.form.get('day_of_week', sched['day_of_week'])),
            'start_time':  new_start_time,
            'end_time':    new_end_time,
            'grace_minutes': int(request.form.get('grace_minutes', sched['grace_minutes'])),
            'created_by':  sched['created_by'],
        })
        flash('Schedule updated.', 'success')
    except Exception as e:
        flash(f'Error updating schedule: {e}', 'danger')
    return redirect(url_for('admin_schedules'))

@app.route('/admin/schedules/<schedule_id>/delete', methods=['POST'])
@admin_required
def admin_schedule_delete(schedule_id):
    db_delete_schedule(schedule_id)
    flash('Schedule removed.', 'success')
    return redirect(url_for('admin_schedules'))

@app.route('/api/schedules/today')
@login_required
def api_schedules_today():
    """Returns today's schedules for the current teacher (JSON)."""
    username  = session.get('username', '')
    schedules = get_todays_schedules(username)
    return jsonify({'schedules': schedules, 'dow_names': DOW_NAMES})


@app.route('/api/schedules/search')
@admin_required
def api_schedules_search():
    """Autocomplete search for teachers and subjects for schedule grid."""
    q = request.args.get('q', '').strip().lower()
    users = db_get_all_users()
    results = []
    for uname, u in users.items():
        if u.get('role') not in ('teacher', 'admin', 'super_admin'):
            continue
        name = u.get('full_name', uname)
        if not q or q in name.lower() or q in uname.lower():
            results.append({'value': uname, 'label': name, 'type': u.get('role', 'teacher')})
    subjects = db_get_all_subjects()
    subj_results = []
    for sid, s in subjects.items():
        label = s.get('name', '')
        code  = s.get('course_code', '')
        if not q or q in label.lower() or q in code.lower():
            subj_results.append({'value': sid, 'code': code, 'label': label})
    return jsonify({'teachers': results[:20], 'subjects': subj_results[:20]})

@app.route('/api/active_sessions_info')
@login_required
def api_active_sessions_info():
    """Return active sessions keyed by session id for schedule live indicators."""
    active = get_active_sessions()
    current_role = session.get('role', '')
    out = {}
    for sid, s in active.items():
        if current_role not in ADMIN_ROLES and not _is_my_session(s):
            continue
        out[sid] = {
            'sess_id': sid,
            'schedule_id': s.get('schedule_id') or '',
            'subject_id': s.get('subject_id', ''),
            'section_key': normalize_section_key(s.get('section_key', '')),
            'teacher_username': s.get('teacher_username', s.get('teacher', '')),
            'teacher_name': s.get('teacher_name', ''),
            'is_active': not bool(s.get('ended_at')),
        }
    return jsonify(out)


@app.route('/excuse/submit/<sess_id>/<nfc_id>', methods=['GET', 'POST'])
def excuse_submit(sess_id, nfc_id):
    sess = load_session(sess_id)
    if not sess:
        return jsonify({'status':'error', 'message':'Session not found'}), 404
    
    student = db_get_student(nfc_id)
    if not student:
        return jsonify({'status':'error', 'message':'Student not found'}), 404

    if request.method == 'GET':
        # Render the submission form for students
        return render_template('excuse_form.html', 
                               sess_id=sess_id, 
                               nfc_id=nfc_id, 
                               student=student, 
                               sess=sess,
                               excuse_reasons=EXCUSE_REASONS)

    reason_type   = request.form.get('reason_type', 'others')
    reason_detail = request.form.get('reason_detail', '').strip()
    
    if reason_type not in dict(EXCUSE_REASONS):
        return jsonify({'status':'error', 'message':'Invalid reason selected'}), 400
    if reason_type == 'others' and not reason_detail:
        return jsonify({'status':'error', 'message':'Please provide details for "Others"'}), 400

    attachment_file = ''
    uploaded = request.files.get('attachment')
    if not uploaded and 'attachments' in request.files:
        uploaded = request.files['attachments']
        
    if uploaded and uploaded.filename:
        try:
            attachment_file = _save_excuse_attachment(uploaded)
        except ValueError as ve:
            return jsonify({'status':'error', 'message':str(ve)}), 400

    excuse_id = db_save_excuse_request({
        'sess_id':       sess_id,
        'nfc_id':        nfc_id,
        'student_name':  student.get('name', ''),
        'student_id':    student.get('student_id', ''),
        'student_email': student.get('email', ''),
        'reason_type':   reason_type,
        'reason_detail': reason_detail,
        'attachment_file': attachment_file,
    })

    # Auto-approve if submitted by teacher for their own session or by any admin
    is_admin = session.get('role') in ADMIN_ROLES
    is_teacher = sess.get('teacher') == session.get('username', '')
    
    if is_admin or is_teacher:
        res = db_resolve_excuse(excuse_id, 'approved', session.get('username', 'system'))
        msg = 'Student marked as excused successfully.'
        reason_text = res.get('reason_detail', '') # fallback
        # Wait, db_resolve_excuse returns the row. 
        # Actually I can get the note from attendance_logs after resolve.
        with get_db() as conn:
            log_row = conn.execute("SELECT excuse_note FROM attendance_logs WHERE excuse_request_id=?", (excuse_id,)).fetchone()
            reason_text = log_row['excuse_note'] if log_row else ''
    else:
        # Notify student via email
        _send_excuse_received_email(
            student.get('email', ''), student.get('name', ''),
            sess.get('subject_name', ''), reason_type, excuse_id
        )
        msg = 'Excuse request submitted and is pending review.'
        reason_text = ''

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'json' in request.accept_mimetypes:
        return jsonify({'status': 'ok', 'message': msg, 'nfc_id': nfc_id, 'reason': reason_text})
    
    flash(msg, 'success')
    return redirect(url_for('admin_excuses') if is_admin else url_for('teacher_records'))

@app.route('/admin/excuses')
@admin_required
def admin_excuses():
    tab    = request.args.get('tab', 'pending')
    filter_status = tab if tab in ('pending','approved','rejected') else None
    excuses = db_get_all_excuse_requests(filter_status)
    counts  = {
        'pending':  len(db_get_all_excuse_requests('pending')),
        'approved': len(db_get_all_excuse_requests('approved')),
        'rejected': len(db_get_all_excuse_requests('rejected')),
    }
    return render_template('admin_excuses.html',
                           excuses=excuses, tab=tab, counts=counts,
                           excuse_reasons=dict(EXCUSE_REASONS))

@app.route('/admin/excuses/<int:excuse_id>/approve', methods=['POST'])
@admin_required
def admin_excuse_approve(excuse_id):
    row = db_resolve_excuse(excuse_id, 'approved', session.get('username',''))
    if row:
        _send_excuse_resolved_email(
            row.get('student_email',''), row.get('student_name',''),
            row.get('reason_type',''), 'approved'
        )
        flash('Excuse approved and attendance updated to Excused.', 'success')
    else:
        flash('Excuse request not found.', 'danger')
    return redirect(url_for('admin_excuses'))

@app.route('/admin/excuses/<int:excuse_id>/reject', methods=['POST'])
@admin_required
def admin_excuse_reject(excuse_id):
    row = db_resolve_excuse(excuse_id, 'rejected', session.get('username',''))
    if row:
        _send_excuse_resolved_email(
            row.get('student_email',''), row.get('student_name',''),
            row.get('reason_type',''), 'rejected'
        )
        flash('Excuse request rejected.', 'warning')
    else:
        flash('Excuse request not found.', 'danger')
    return redirect(url_for('admin_excuses'))

@app.route('/admin/excuses/<int:excuse_id>/attachment')
@admin_required
def admin_excuse_attachment(excuse_id):
    exc = db_get_excuse_request(excuse_id)
    if not exc or not exc.get('attachment_file'):
        flash('Attachment not found.', 'danger')
        return redirect(url_for('admin_excuses'))
    fpath = _resolve_excuse_attachment_path(exc['attachment_file'])
    if not fpath or not os.path.exists(fpath):
        flash('Attachment file missing from server.', 'danger')
        return redirect(url_for('admin_excuses'))
    from flask import send_file
    return send_file(fpath, as_attachment=False)

# ── Excuse email helpers ───────────────────────────────────────────────────

def _send_excuse_received_email(email, student_name, subject_name, reason_type, excuse_id):
    if not email or '@' not in email: return
    reason_label = dict(EXCUSE_REASONS).get(reason_type, reason_type.title())
    html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Calibri,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:32px 16px;">
    <table width="520" cellpadding="0" cellspacing="0"
           style="background:#fff;border-radius:12px;overflow:hidden;
                  box-shadow:0 2px 12px rgba(0,0,0,.1);">
      <tr><td style="background:#1E4A1A;padding:20px 28px;">
        <div style="font-size:18px;font-weight:700;color:#F5C518;">DAVS</div>
        <div style="font-size:11px;color:#94a3b8;">Decentralized Attendance Verification System</div>
      </td></tr>
      <tr><td style="padding:24px 28px;">
        <p style="font-size:15px;color:#1E4A1A;font-weight:700;">Excuse Request Received</p>
        <p style="font-size:13px;color:#444;">Dear <strong>{student_name}</strong>,</p>
        <p style="font-size:13px;color:#444;">
          Your excuse request for <strong>{subject_name}</strong> has been received and is
          <strong>pending review</strong> by the administrator.
        </p>
        <table style="border:1px solid #eee;border-radius:8px;width:100%;">
          <tr><td style="padding:8px 12px;font-size:12px;color:#666;width:130px;">Reason</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;">{reason_label}</td></tr>
          <tr><td style="padding:8px 12px;font-size:12px;color:#666;">Request #</td>
              <td style="padding:8px 12px;font-size:12px;font-family:monospace;color:#333;">#{excuse_id}</td></tr>
        </table>
        <p style="font-size:11px;color:#94a3b8;margin-top:20px;">
          You will receive another email when your request is reviewed.<br>Please do not reply to this email.
        </p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>'''
    _send_email([email], f'[DAVS] Excuse Request Received — {subject_name}', html)

def _send_excuse_resolved_email(email, student_name, reason_type, resolution):
    if not email or '@' not in email: return
    reason_label = dict(EXCUSE_REASONS).get(reason_type, reason_type.title())
    color   = '#2D6A27' if resolution == 'approved' else '#C0392B'
    badge   = 'APPROVED ✓' if resolution == 'approved' else 'REJECTED ✕'
    message = ('Your excuse has been <strong>approved</strong> and your attendance has been '
               'updated to <strong>Excused</strong>.' if resolution == 'approved'
               else 'Your excuse request has been <strong>rejected</strong>. '
                    'Please contact your instructor or administrator for more details.')
    html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Calibri,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0">
  <tr><td align="center" style="padding:32px 16px;">
    <table width="520" cellpadding="0" cellspacing="0"
           style="background:#fff;border-radius:12px;overflow:hidden;
                  box-shadow:0 2px 12px rgba(0,0,0,.1);">
      <tr><td style="background:#1E4A1A;padding:20px 28px;">
        <div style="font-size:18px;font-weight:700;color:#F5C518;">DAVS</div>
        <div style="font-size:11px;color:#94a3b8;">Decentralized Attendance Verification System</div>
      </td></tr>
      <tr><td style="padding:24px 28px;">
        <div style="font-size:22px;font-weight:700;color:{color};margin-bottom:12px;">{badge}</div>
        <p style="font-size:13px;color:#444;">Dear <strong>{student_name}</strong>,</p>
        <p style="font-size:13px;color:#444;">{message}</p>
        <table style="border:1px solid #eee;border-radius:8px;width:100%;">
          <tr><td style="padding:8px 12px;font-size:12px;color:#666;width:130px;">Reason Filed</td>
              <td style="padding:8px 12px;font-size:12px;color:#333;">{reason_label}</td></tr>
          <tr><td style="padding:8px 12px;font-size:12px;color:#666;">Decision</td>
              <td style="padding:8px 12px;font-size:12px;font-weight:700;color:{color};"
              >{resolution.title()}</td></tr>
        </table>
        <p style="font-size:11px;color:#94a3b8;margin-top:20px;">Please do not reply to this email.</p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>'''
    _send_email([email], f'[DAVS] Excuse Request {resolution.title()} — {reason_label}', html)

# ── helper: collect all existing section keys ────────────────────────────────
def _get_all_section_keys():
    with get_db() as conn:
        sess_rows = conn.execute(
            "SELECT DISTINCT section_key FROM sessions WHERE section_key != ''"
        ).fetchall()
        stu_rows  = conn.execute(
            "SELECT DISTINCT program, year_level, section FROM students "
            "WHERE program != '' AND year_level != '' AND section != ''"
        ).fetchall()
    keys = set()
    for r in sess_rows:
        keys.add(normalize_section_key(r['section_key']))
    for r in stu_rows:
        k = normalize_section_key(f"{r['program']}|{r['year_level']}|{r['section']}")
        if k: keys.add(k)
    return sorted(keys)

def _save_excuse_attachment(uploaded):
    """Saves an excuse attachment file and returns the secure filename."""
    if not uploaded or not uploaded.filename:
        return ''
    
    ext = uploaded.filename.rsplit('.', 1)[1].lower() if '.' in uploaded.filename else ''
    if ext not in ALLOWED_EXTENSIONS_EXCUSES:
        raise ValueError(f"File extension '.{ext}' not allowed.")
    
    fname = f"excuse_{uuid.uuid4().hex[:8]}_{secure_filename(uploaded.filename)}"
    fpath = os.path.join(UPLOAD_FOLDER_EXCUSES, fname)
    uploaded.save(fpath)
    return fname

@app.route('/teacher/schedule')
@staff_required
def teacher_schedule():
    username = session.get('username')
    # Filter only for the current teacher
    schedules = db_get_schedules_for_teacher(username)
    subjects = db_get_all_subjects()
    return render_template('teacher_schedule.html',
                           schedules=schedules,
                           subjects=subjects)

@app.route('/api/schedules/upcoming')
@login_required
def api_schedules_upcoming():
    """Upcoming schedules for notification (5-min lead time)."""
    username = session.get('username')
    now = datetime.now()
    today_dow = now.weekday()
    # 5 minutes from now
    target_time = (now + timedelta(minutes=5)).strftime('%H:%M')
    
    schedules = db_get_schedules_for_teacher(username)
    upcoming = [s for s in schedules if s['day_of_week'] == today_dow and s['start_time'] == target_time]
    
    return jsonify({'upcoming': upcoming})

def _row_to_dict(row):
    return dict(row) if row else {}

import subprocess as _sp
import os as _os
import sys as _sys

def _launch_nfc_listener():
    listener = _os.path.join(_os.path.dirname(__file__), 'nfc_listener.py')
    if not _os.path.exists(listener):
        print("[NFC] nfc_listener.py not found — skipping auto-launch.")
        return
    # Kill any existing nfc_listener process first
    if _sys.platform == 'win32':
        _sp.run(['taskkill', '/F', '/IM', 'python.exe', '/FI',
                 f'WINDOWTITLE eq nfc_listener'],
                capture_output=True)
        proc = _sp.Popen(
            [_sys.executable, listener],
            creationflags=0x00000008 | 0x08000000,  # DETACHED_PROCESS + CREATE_NO_WINDOW
            stdout=_sp.DEVNULL,
            stderr=_sp.DEVNULL,
            close_fds=True,
        )
    else:
        proc = _sp.Popen(
            [_sys.executable, listener],
            start_new_session=True,
            stdout=_sp.DEVNULL,
            stderr=_sp.DEVNULL,
        )
    print(f"[NFC] Listener started in background (PID {proc.pid})")
    print("[NFC] Check nfc_listener.log for tap activity.")

if __name__ == '__main__':
    ensure_automation_thread_running()
    _launch_nfc_listener()
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)